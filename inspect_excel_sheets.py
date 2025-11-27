import openpyxl
from openpyxl import load_workbook
path = r"C:\gestao\MapasDeFocoBncc_Unificados.xlsx"
wb = load_workbook(filename=path, read_only=True, data_only=True)
print('Sheets found:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    # get first non-empty row as header
    header = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if any(cell is not None for cell in row):
            header = row
            break
    print('\nSheet:', name)
    print('Header sample:', header)
print('\nDone')
