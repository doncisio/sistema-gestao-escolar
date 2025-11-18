from pathlib import Path
import openpyxl

orig = Path('C:/gestao/pendencias_por_bimestre.xlsx')
test = Path('C:/gestao/teste_pendencias.xlsx')

if not test.exists():
    print(f'Teste ausente: {test.resolve()}')
    raise SystemExit(1)

if not orig.exists():
    print(f'Arquivo original não encontrado: {orig.resolve()}')
    print('Nada para comparar — talvez o original estivesse corrompido ou ausente.')
    raise SystemExit(0)

wb_orig = openpyxl.load_workbook(orig, data_only=True)
wb_test = openpyxl.load_workbook(test, data_only=True)

sheets_orig = wb_orig.sheetnames
sheets_test = wb_test.sheetnames

print('Original sheets:', sheets_orig)
print('Test sheets    :', sheets_test)

common = [s for s in sheets_test if s in sheets_orig]
print('\nSheets em comum:', common)

any_diff = False
for s in common:
    ws_o = wb_orig[s]
    ws_t = wb_test[s]
    dims_o = {k:v.width for k,v in ws_o.column_dimensions.items()}
    dims_t = {k:v.width for k,v in ws_t.column_dimensions.items()}
    cols = sorted(set(list(dims_o.keys()) + list(dims_t.keys())))
    diffs = []
    for c in cols:
        wo = dims_o.get(c)
        wt = dims_t.get(c)
        if wo != wt:
            diffs.append((c, wo, wt))
    if diffs:
        any_diff = True
        print('\n-- Diferenças na sheet:', s)
        for c, wo, wt in diffs:
            print(f"Coluna {c}: original={wo!r}  test={wt!r}")
    else:
        print(f"\n-- Sheet '{s}': larguras preservadas (nenhuma diferença detectada)")

# Report sheets only in one file
only_orig = [s for s in sheets_orig if s not in sheets_test]
only_test = [s for s in sheets_test if s not in sheets_orig]
if only_orig:
    print('\nAbas apenas no original:', only_orig)
if only_test:
    print('\nAbas apenas no teste:', only_test)

wb_orig.close()
wb_test.close()

if not any_diff:
    print('\nResumo: não foram detectadas diferenças nas larguras entre original e teste nas abas em comum.')
else:
    print('\nResumo: foram detectadas diferenças nas larguras para algumas colunas indicadas acima.')
