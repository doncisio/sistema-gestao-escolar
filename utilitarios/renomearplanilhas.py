import openpyxl
import os
from config_logs import get_logger

logger = get_logger(__name__)

def process_excel_file(file_path):
    # Abrir o arquivo Excel
    wb = openpyxl.load_workbook(file_path)
    
    # Lista esperada de planilhas
    expected_sheets = [f"{i}º Ano" for i in range(1, 10)]
    alternative_sheets = [f"{i}º Ano 1" for i in range(1, 10)]
    
    # Verificar se todas as planilhas já estão no formato correto
    if set(wb.sheetnames) == set(expected_sheets):
        logger.info(f"O arquivo '{file_path}' já está no formato correto. Nenhuma modificação necessária.")
        return
    
    # Excluir todas as planilhas do tipo "nº ano" (de "1º ano" a "9º ano")
    for sheet_name in expected_sheets:
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
    
    # Renomear as planilhas do tipo "nº ano 1" para "nº ano"
    for alt_sheet, final_sheet in zip(alternative_sheets, expected_sheets):
        if alt_sheet in wb.sheetnames:
            # Renomear apenas se existir a alternativa
            wb[alt_sheet].title = final_sheet
    
    # Salvar o arquivo modificado
    wb.save(file_path)
    logger.info(f"O arquivo '{file_path}' foi processado e atualizado.")

# Caminho da pasta com os arquivos Excel
folder_path = "H:/Meu Drive/NADIR_2024/ATAS DIGITALIZADAS/dados historico escolar/FINALIZADAS - Copia"

# Processar todos os arquivos Excel na pasta
for file_name in os.listdir(folder_path):
    if file_name.endswith(".xlsx"):  # Verificar se é um arquivo Excel
        file_path = os.path.join(folder_path, file_name)
        process_excel_file(file_path)

logger.info("Processamento concluído!")
