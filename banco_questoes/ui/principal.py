"""
Interface Principal do Banco de Questões BNCC.

Janela principal que integra todas as funcionalidades do banco de questões.
"""

from config_logs import get_logger
logger = get_logger(__name__)
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Optional, Any
import config

from config import perfis_habilitados
from auth.usuario_logado import UsuarioLogado
from banco_questoes.texto_base_service import TextoBaseService, TipoTextoBase


class InterfaceBancoQuestoes:
    """Interface principal do Banco de Questões BNCC."""
    
    def __init__(self, root=None, janela_principal=None):
        """
        Inicializa a interface principal do banco de questões.
        
        Args:
            root: Janela Tk/Toplevel existente ou None para criar nova
            janela_principal: Referência à janela principal da aplicação
        """
        self.janela_principal = janela_principal
        
        if root is None:
            self.janela = tk.Toplevel()
            self.janela.title("Banco de Questões BNCC")
            self.janela.geometry("1100x700")
            self.janela.grab_set()
            self.janela.focus_force()
            self.janela.protocol("WM_DELETE_WINDOW", self.ao_fechar_janela)
        else:
            self.janela = root

        # Cores padrão do sistema
        self.co0 = "#F5F5F5"
        self.co1 = "#003A70"
        self.co2 = "#77B341"
        self.co3 = "#E2418E"
        self.co4 = "#4A86E8"
        self.co7 = "#333333"
        self.co8 = "#BF3036"
        self.co9 = "#999999"
        
        self.janela.configure(bg=self.co0)
        # Cache para referências de imagens (evita garbage collection do PhotoImage)
        self._image_refs = []
        # Cache para previews de imagens (evita recarregar)
        self._cache_imagens = {}
        
        # Verificar perfil do usuário
        self.perfil = None
        self.funcionario_id = None
        if perfis_habilitados():
            self.perfil = UsuarioLogado.get_perfil()
            self.funcionario_id = UsuarioLogado.get_funcionario_id()
        
        self.criar_interface()
    
    def ao_fechar_janela(self):
        """Trata o evento de fechamento da janela."""
        if self.janela_principal:
            self.janela_principal.deiconify()
        self.janela.destroy()
    
    def criar_interface(self):
        """Cria a interface principal com abas."""
        # Frame de título
        self.frame_titulo = tk.Frame(self.janela, bg=self.co1)
        self.frame_titulo.pack(side="top", fill="x")
        
        tk.Label(
            self.frame_titulo,
            text="📚 Banco de Questões BNCC",
            font=("Arial", 16, "bold"),
            bg=self.co1, fg="white"
        ).pack(fill="x", padx=10, pady=10)
        
        # Notebook (abas)
        self.notebook = ttk.Notebook(self.janela)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Aba 1: Buscar Questões
        self.frame_busca = tk.Frame(self.notebook, bg=self.co0)
        self.notebook.add(self.frame_busca, text="🔍 Buscar Questões")
        self.criar_aba_busca()
        
        # Aba 2: Cadastrar Questão (apenas se tem permissão)
        if self._pode_criar_questao():
            self.frame_cadastro = tk.Frame(self.notebook, bg=self.co0)
            self.notebook.add(self.frame_cadastro, text="➕ Nova Questão")
            self.criar_aba_cadastro()
        
        # Aba 3: Montar Avaliação
        if self._pode_criar_avaliacao():
            self.frame_avaliacao = tk.Frame(self.notebook, bg=self.co0)
            self.notebook.add(self.frame_avaliacao, text="📝 Montar Avaliação")
            self.criar_aba_avaliacao()
        
        # Aba 4: Minhas Questões (professor vê apenas suas)
        self.frame_minhas = tk.Frame(self.notebook, bg=self.co0)
        self.notebook.add(self.frame_minhas, text="📋 Minhas Questões")
        self.criar_aba_minhas_questoes()
        
        # Aba 5: Estatísticas (apenas coordenador/admin)
        if self.perfil in ['administrador', 'coordenador']:
            self.frame_estatisticas = tk.Frame(self.notebook, bg=self.co0)
            self.notebook.add(self.frame_estatisticas, text="📊 Estatísticas")
            self.criar_aba_estatisticas()
        
        # Aba 6: Textos Base (apenas se pode criar questões/avaliações)
        if self._pode_criar_questao():
            self.frame_textos_base = tk.Frame(self.notebook, bg=self.co0)
            self.notebook.add(self.frame_textos_base, text="📄 Textos Base")
            self.criar_aba_textos_base()
    
    def _pode_criar_questao(self) -> bool:
        """Verifica se o usuário pode criar questões."""
        if not perfis_habilitados():
            return True
        return self.perfil in ['administrador', 'coordenador', 'professor']
    
    def _pode_criar_avaliacao(self) -> bool:
        """Verifica se o usuário pode criar avaliações."""
        if not perfis_habilitados():
            return True
        return self.perfil in ['administrador', 'coordenador', 'professor']
    
    # =========================================================================
    # MAPEAMENTOS PARA HABILIDADES BNCC
    # =========================================================================
    
    def _get_componente_sigla(self, componente: str) -> Optional[str]:
        """Retorna a sigla BNCC do componente curricular."""
        mapa = {
            "Língua Portuguesa": "LP",
            "Matemática": "MA",
            "Ciências": "CI",
            "Geografia": "GE",
            "História": "HI",
            "Arte": "AR",
            "Educação Física": "EDF",
            "Língua Inglesa": "LI",
            "Ensino Religioso": "ER",
            "Computação": "CO"
        }
        return mapa.get(componente)
    
    def _get_ano_bloco(self, ano: str) -> list:
        """Retorna os códigos de ano_bloco correspondentes ao ano escolar selecionado."""
        # Mapeamento de ano escolar para possíveis valores de ano_bloco
        # ano_bloco pode ser: '01', '02', ..., '09', '12', '15', '35', '67', '69', '89'
        mapa = {
            "1º ano": ["01", "12", "15"],
            "2º ano": ["02", "12", "15"],
            "3º ano": ["03", "35", "15"],
            "4º ano": ["04", "35", "15"],
            "5º ano": ["05", "35", "15"],
            "6º ano": ["06", "67", "69"],
            "7º ano": ["07", "67", "69"],
            "8º ano": ["08", "89", "69"],
            "9º ano": ["09", "89", "69"]
        }
        return mapa.get(ano, [])
    
    def _carregar_habilidades_bncc(self, componente_sigla: Optional[str] = None, anos_bloco: Optional[list] = None) -> list:
        """Carrega habilidades BNCC do banco de dados filtradas por componente e ano."""
        try:
            from conexao import conectar_bd
            conn = conectar_bd()
            if not conn:
                return []
            
            cursor: Any = conn.cursor(dictionary=True)
            
            # Construir query com filtros opcionais
            query = "SELECT codigo, descricao FROM bncc_habilidades WHERE 1=1"
            params = []
            
            if componente_sigla:
                query += " AND componente_codigo = %s"
                params.append(componente_sigla)
            
            if anos_bloco:
                placeholders = ", ".join(["%s"] * len(anos_bloco))
                query += f" AND ano_bloco IN ({placeholders})"
                params.extend(anos_bloco)
            
            query += " ORDER BY codigo LIMIT 500"
            
            cursor.execute(query, params)
            resultados = cursor.fetchall()
            cursor.close()
            conn.close()
            
            # Formatar como "CODIGO - Descrição resumida"
            habilidades = []
            for row in resultados:
                codigo = row['codigo']
                descricao = row['descricao'] or ''
                # Truncar descrição se muito longa
                if len(descricao) > 80:
                    descricao = descricao[:77] + "..."
                habilidades.append(f"{codigo} - {descricao}")
            
            return habilidades
            
        except Exception as e:
            logger.error(f"Erro ao carregar habilidades BNCC: {e}")
            return []
    
    def _atualizar_habilidades_busca(self, event=None):
        """Atualiza o combobox de habilidades na aba de busca com base nos filtros."""
        componente = self.cb_componente.get()
        ano = self.cb_ano.get()
        
        componente_sigla = None
        anos_bloco = None
        
        if componente != "Todos":
            componente_sigla = self._get_componente_sigla(componente)
        
        if ano != "Todos":
            anos_bloco = self._get_ano_bloco(ano)
        
        # Carregar habilidades filtradas
        habilidades = self._carregar_habilidades_bncc(componente_sigla, anos_bloco)
        
        # Atualizar combobox
        self.cb_habilidade_busca['values'] = ["Todas"] + habilidades
        self.cb_habilidade_busca.current(0)
    
    def _atualizar_habilidades_cadastro(self, event=None):
        """Atualiza o combobox de habilidades na aba de cadastro com base nos filtros."""
        if not hasattr(self, 'cad_habilidade') or not hasattr(self, 'cad_componente') or not hasattr(self, 'cad_ano'):
            return
            
        componente = self.cad_componente.get()
        ano = self.cad_ano.get()
        
        componente_sigla = None
        anos_bloco = None
        
        if componente:
            componente_sigla = self._get_componente_sigla(componente)
        
        if ano:
            anos_bloco = self._get_ano_bloco(ano)
        
        # Carregar habilidades filtradas
        habilidades = self._carregar_habilidades_bncc(componente_sigla, anos_bloco)
        
        # Atualizar combobox
        self.cad_habilidade['values'] = [""] + habilidades
        self.cad_habilidade.set("")
    
    # =========================================================================
    # ABA: BUSCAR QUESTÕES
    # =========================================================================
    
    def criar_aba_busca(self):
        """Cria a aba de busca de questões."""
        # Frame de filtros
        frame_filtros = tk.LabelFrame(
            self.frame_busca, text="Filtros de Busca",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_filtros.pack(fill="x", padx=10, pady=10)
        
        # Linha 1: Componente e Ano
        tk.Label(frame_filtros, text="Componente:", bg=self.co0).grid(
            row=0, column=0, padx=5, pady=5, sticky="w"
        )
        self.cb_componente = ttk.Combobox(frame_filtros, width=25, state="readonly")
        self.cb_componente['values'] = [
            "Todos", "Língua Portuguesa", "Matemática", "Ciências",
            "Geografia", "História", "Arte", "Educação Física",
            "Língua Inglesa", "Ensino Religioso", "Computação"
        ]
        self.cb_componente.current(0)
        self.cb_componente.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.cb_componente.bind("<<ComboboxSelected>>", self._atualizar_habilidades_busca)
        
        tk.Label(frame_filtros, text="Ano:", bg=self.co0).grid(
            row=0, column=2, padx=5, pady=5, sticky="w"
        )
        self.cb_ano = ttk.Combobox(frame_filtros, width=15, state="readonly")
        self.cb_ano['values'] = [
            "Todos", "1º ano", "2º ano", "3º ano", "4º ano", "5º ano",
            "6º ano", "7º ano", "8º ano", "9º ano"
        ]
        self.cb_ano.current(0)
        self.cb_ano.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.cb_ano.bind("<<ComboboxSelected>>", self._atualizar_habilidades_busca)
        
        # Linha 2: Dificuldade e Tipo
        tk.Label(frame_filtros, text="Dificuldade:", bg=self.co0).grid(
            row=1, column=0, padx=5, pady=5, sticky="w"
        )
        self.cb_dificuldade = ttk.Combobox(frame_filtros, width=15, state="readonly")
        self.cb_dificuldade['values'] = ["Todas", "Fácil", "Média", "Difícil"]
        self.cb_dificuldade.current(0)
        self.cb_dificuldade.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        
        tk.Label(frame_filtros, text="Tipo:", bg=self.co0).grid(
            row=1, column=2, padx=5, pady=5, sticky="w"
        )
        self.cb_tipo = ttk.Combobox(frame_filtros, width=20, state="readonly")
        self.cb_tipo['values'] = [
            "Todos", "Múltipla Escolha", "Dissertativa", "V/F"
        ]
        self.cb_tipo.current(0)
        self.cb_tipo.grid(row=1, column=3, padx=5, pady=5, sticky="w")
        
        # Linha 3: Habilidade BNCC e texto
        tk.Label(frame_filtros, text="Habilidade BNCC:", bg=self.co0).grid(
            row=2, column=0, padx=5, pady=5, sticky="w"
        )
        self.cb_habilidade_busca = ttk.Combobox(frame_filtros, width=40, state="readonly")
        self.cb_habilidade_busca['values'] = ["Todas"]
        self.cb_habilidade_busca.current(0)
        self.cb_habilidade_busca.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="w")
        
        # Linha 4: Texto busca
        tk.Label(frame_filtros, text="Texto (busca):", bg=self.co0).grid(
            row=3, column=0, padx=5, pady=5, sticky="w"
        )
        self.entry_texto = ttk.Entry(frame_filtros, width=50)
        self.entry_texto.grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky="w")
        
        # Botão buscar
        btn_buscar = tk.Button(
            frame_filtros, text="🔍 Buscar",
            command=self.buscar_questoes,
            bg=self.co4, fg="white", font=("Arial", 10, "bold")
        )
        btn_buscar.grid(row=3, column=3, padx=10, pady=5)
        
        # Frame de resultados
        frame_resultados = tk.LabelFrame(
            self.frame_busca, text="Resultados",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_resultados.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Treeview para resultados
        colunas = ("id", "componente", "ano", "habilidade", "tipo", "dificuldade", "status")
        self.tree_busca = ttk.Treeview(
            frame_resultados, columns=colunas, show="headings", height=15
        )
        
        self.tree_busca.heading("id", text="ID")
        self.tree_busca.heading("componente", text="Componente")
        self.tree_busca.heading("ano", text="Ano")
        self.tree_busca.heading("habilidade", text="Habilidade")
        self.tree_busca.heading("tipo", text="Tipo")
        self.tree_busca.heading("dificuldade", text="Dificuldade")
        self.tree_busca.heading("status", text="Status")
        
        self.tree_busca.column("id", width=50)
        self.tree_busca.column("componente", width=150)
        self.tree_busca.column("ano", width=80)
        self.tree_busca.column("habilidade", width=100)
        self.tree_busca.column("tipo", width=120)
        self.tree_busca.column("dificuldade", width=80)
        self.tree_busca.column("status", width=80)
        
        scrollbar = ttk.Scrollbar(frame_resultados, orient="vertical", command=self.tree_busca.yview)
        self.tree_busca.configure(yscrollcommand=scrollbar.set)
        
        self.tree_busca.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind duplo clique para ver detalhes
        self.tree_busca.bind("<Double-1>", self.ver_questao_selecionada)
        
        # Frame de ações
        frame_acoes = tk.Frame(self.frame_busca, bg=self.co0)
        frame_acoes.pack(fill="x", padx=10, pady=5)
        
        tk.Button(
            frame_acoes, text="👁️ Ver Detalhes",
            command=self.ver_questao_selecionada,
            bg=self.co4, fg="white"
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_acoes, text="⭐ Favoritar",
            command=self.favoritar_questao,
            bg=self.co2, fg="white"
        ).pack(side="left", padx=5)
        
        if self._pode_criar_avaliacao():
            tk.Button(
                frame_acoes, text="➕ Adicionar à Avaliação",
                command=self.adicionar_avaliacao,
                bg=self.co3, fg="white"
            ).pack(side="left", padx=5)
    
    def buscar_questoes(self):
        """Executa a busca de questões."""
        from banco_questoes.services import QuestaoService
        from banco_questoes.models import FiltroQuestoes, ComponenteCurricular, AnoEscolar, DificuldadeQuestao, TipoQuestao, StatusQuestao
        
        # Limpar resultados anteriores
        for item in self.tree_busca.get_children():
            self.tree_busca.delete(item)
        
        # Montar filtros
        filtros = FiltroQuestoes()
        
        if self.cb_componente.get() != "Todos":
            try:
                filtros.componente_curricular = ComponenteCurricular(self.cb_componente.get())
            except ValueError:
                pass
        
        if self.cb_ano.get() != "Todos":
            try:
                filtros.ano_escolar = AnoEscolar(self.cb_ano.get())
            except ValueError:
                pass
        
        if self.cb_dificuldade.get() != "Todas":
            mapa_dif = {"Fácil": "facil", "Média": "media", "Difícil": "dificil"}
            val = mapa_dif.get(self.cb_dificuldade.get())
            if val:
                try:
                    filtros.dificuldade = DificuldadeQuestao(val)
                except ValueError:
                    pass
        
        if self.cb_tipo.get() != "Todos":
            mapa_tipo = {
                "Múltipla Escolha": "multipla_escolha",
                "Dissertativa": "dissertativa",
                "V/F": "verdadeiro_falso"
            }
            val = mapa_tipo.get(self.cb_tipo.get())
            if val:
                try:
                    filtros.tipo = TipoQuestao(val)
                except ValueError:
                    pass
        
        habilidade_selecionada = self.cb_habilidade_busca.get()
        if habilidade_selecionada and habilidade_selecionada != "Todas":
            # Extrair apenas o código da habilidade (antes do " - ")
            if " - " in habilidade_selecionada:
                habilidade = habilidade_selecionada.split(" - ")[0].strip()
            else:
                habilidade = habilidade_selecionada.strip()
            filtros.habilidade_bncc = habilidade
        
        texto = self.entry_texto.get().strip()
        if texto:
            filtros.texto_busca = texto
        
        try:
            questoes, total = QuestaoService.buscar(filtros, limite=100)
            
            for q in questoes:
                self.tree_busca.insert("", "end", values=(
                    q.id,
                    q.componente_curricular.value if q.componente_curricular else '',
                    q.ano_escolar.value if q.ano_escolar else '',
                    q.habilidade_bncc_codigo or '',
                    q.tipo.value if q.tipo else '',
                    q.dificuldade.value if q.dificuldade else '',
                    q.status.value if q.status else ''
                ))
            
            if not questoes:
                messagebox.showinfo("Busca", "Nenhuma questão encontrada com os filtros selecionados.")
                
        except Exception as e:
            logger.error(f"Erro na busca de questões: {e}")
            messagebox.showerror("Erro", f"Erro ao buscar questões: {e}")
    
    def ver_questao_selecionada(self, event=None):
        """Abre detalhes da questão selecionada."""
        selecao = self.tree_busca.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para ver os detalhes.")
            return
        
        item = self.tree_busca.item(selecao[0])
        questao_id = int(item['values'][0])
        
        # Abrir janela de detalhes
        self.abrir_detalhes_questao(questao_id)
    
    def favoritar_questao(self):
        """Adiciona questão aos favoritos."""
        selecao = self.tree_busca.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para favoritar.")
            return
        
        if not self.funcionario_id:
            messagebox.showwarning("Aviso", "É necessário estar logado para favoritar.")
            return
        
        item = self.tree_busca.item(selecao[0])
        questao_id = int(item['values'][0])
        
        try:
            from banco_questoes.services import QuestaoService
            sucesso = QuestaoService.favoritar(questao_id, self.funcionario_id)
            if sucesso:
                messagebox.showinfo("Sucesso", "Questão adicionada aos favoritos!")
            else:
                messagebox.showwarning("Aviso", "Não foi possível favoritar a questão.")
        except Exception as e:
            logger.error(f"Erro ao favoritar: {e}")
            messagebox.showerror("Erro", f"Erro ao favoritar questão: {e}")
    
    def adicionar_avaliacao(self):
        """Adiciona questão à avaliação em construção."""
        selecao = self.tree_busca.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para adicionar.")
            return
        
        item = self.tree_busca.item(selecao[0])
        questao_id = int(item['values'][0])
        
        # Mudar para aba de avaliação e adicionar
        if hasattr(self, 'frame_avaliacao'):
            self.notebook.select(self.frame_avaliacao)
            self.adicionar_questao_avaliacao(questao_id)
        else:
            messagebox.showwarning("Aviso", "Você não tem permissão para criar avaliações.")
    
    def abrir_detalhes_questao(self, questao_id: int):
        """Abre janela com detalhes da questão, incluindo imagens."""
        from banco_questoes.services import QuestaoService
        
        try:
            questao = QuestaoService.buscar_por_id(questao_id)
            
            if not questao:
                messagebox.showerror("Erro", "Questão não encontrada.")
                return
            
            # Criar janela de detalhes (maior para acomodar imagens)
            janela_detalhes = tk.Toplevel(self.janela)
            janela_detalhes.title(f"Questão #{questao_id}")
            janela_detalhes.geometry("900x700")
            janela_detalhes.grab_set()
            
            # Frame principal com scroll
            canvas = tk.Canvas(janela_detalhes, bg=self.co0)
            scrollbar = ttk.Scrollbar(janela_detalhes, orient="vertical", command=canvas.yview)
            frame = tk.Frame(canvas, bg=self.co0)
            
            frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
            scrollbar.pack(side="right", fill="y")
            
            # Carregar imagens da questão
            arquivos_questao = self._carregar_arquivos_questao(questao_id)
            imagem_enunciado = None
            imagens_alternativas = {}
            
            for arq in arquivos_questao:
                if arq.get('alternativa_id') is None:
                    imagem_enunciado = arq
                else:
                    # Buscar letra da alternativa
                    letra = arq.get('alternativa_letra', '')
                    if letra:
                        imagens_alternativas[letra] = arq
            
            # Informações básicas
            frame_info = tk.LabelFrame(frame, text="Informações", bg=self.co0, font=("Arial", 10, "bold"))
            frame_info.pack(fill="x", padx=10, pady=5)
            
            info_text = f"""Componente: {questao.componente_curricular.value if questao.componente_curricular else ''}
Ano: {questao.ano_escolar.value if questao.ano_escolar else ''}
Habilidade BNCC: {questao.habilidade_bncc_codigo or ''}
Tipo: {questao.tipo.value if questao.tipo else ''}
Dificuldade: {questao.dificuldade.value if questao.dificuldade else ''}
Status: {questao.status.value if questao.status else ''}"""
            
            tk.Label(frame_info, text=info_text, bg=self.co0, justify="left", font=("Arial", 10)).pack(anchor="w", padx=10, pady=5)
            
            # Enunciado com imagem
            frame_enunciado = tk.LabelFrame(frame, text="Enunciado", bg=self.co0, font=("Arial", 10, "bold"))
            frame_enunciado.pack(fill="x", padx=10, pady=5)
            
            posicao_img = imagem_enunciado.get('posicao', 'abaixo') if imagem_enunciado else 'abaixo'
            
            # Container para layout flexível
            frame_enunciado_container = tk.Frame(frame_enunciado, bg=self.co0)
            frame_enunciado_container.pack(fill="x", padx=5, pady=5)
            
            # Criar widgets conforme posição
            if imagem_enunciado and posicao_img == "acima":
                self._mostrar_imagem_questao(frame_enunciado_container, imagem_enunciado, posicao="top")
            
            if imagem_enunciado and posicao_img in ["esquerda", "direita"]:
                # Layout horizontal
                if posicao_img == "esquerda":
                    frame_img = tk.Frame(frame_enunciado_container, bg=self.co0)
                    frame_img.pack(side="left", padx=5)
                    self._mostrar_imagem_questao(frame_img, imagem_enunciado, max_size=250)
                
                frame_texto = tk.Frame(frame_enunciado_container, bg=self.co0)
                frame_texto.pack(side="left", fill="both", expand=True, padx=5)
                
                txt_enunciado = tk.Text(frame_texto, wrap="word", font=("Arial", 10), height=8)
                txt_enunciado.insert("1.0", questao.enunciado or '')
                txt_enunciado.config(state="disabled")
                txt_enunciado.pack(fill="both", expand=True)
                
                if posicao_img == "direita":
                    frame_img = tk.Frame(frame_enunciado_container, bg=self.co0)
                    frame_img.pack(side="left", padx=5)
                    self._mostrar_imagem_questao(frame_img, imagem_enunciado, max_size=250)
            else:
                # Layout vertical (padrão)
                txt_enunciado = tk.Text(frame_enunciado_container, wrap="word", font=("Arial", 10), height=6)
                txt_enunciado.insert("1.0", questao.enunciado or '')
                txt_enunciado.config(state="disabled")
                txt_enunciado.pack(fill="x")
            
            if imagem_enunciado and posicao_img == "abaixo":
                self._mostrar_imagem_questao(frame_enunciado_container, imagem_enunciado, posicao="bottom")
            
            # Texto de apoio (se houver)
            if questao.texto_apoio:
                frame_apoio = tk.LabelFrame(frame, text="Texto de Apoio", bg=self.co0, font=("Arial", 10, "bold"))
                frame_apoio.pack(fill="x", padx=10, pady=5)
                
                txt_apoio = tk.Text(frame_apoio, wrap="word", font=("Arial", 10), height=4)
                txt_apoio.insert("1.0", questao.texto_apoio)
                txt_apoio.config(state="disabled")
                txt_apoio.pack(fill="x", padx=5, pady=5)
            
            # Alternativas com imagens (se múltipla escolha)
            if questao.alternativas:
                frame_alternativas = tk.LabelFrame(frame, text="Alternativas", bg=self.co0, font=("Arial", 10, "bold"))
                frame_alternativas.pack(fill="x", padx=10, pady=5)
                
                for alt in questao.alternativas:
                    frame_alt = tk.Frame(frame_alternativas, bg=self.co0)
                    frame_alt.pack(fill="x", padx=5, pady=3)
                    
                    marca = "✓" if alt.correta else " "
                    cor_fundo = "#E8F5E9" if alt.correta else self.co0
                    
                    lbl_alt = tk.Label(
                        frame_alt, 
                        text=f"{alt.letra}) [{marca}] {alt.texto}",
                        bg=cor_fundo, font=("Arial", 10), anchor="w"
                    )
                    lbl_alt.pack(side="left", fill="x", expand=True)
                    
                    # Imagem da alternativa (se houver)
                    if alt.letra in imagens_alternativas:
                        self._mostrar_imagem_questao(frame_alt, imagens_alternativas[alt.letra], max_size=80)
            
            # Gabarito dissertativa
            if questao.gabarito_letra:
                frame_gab = tk.LabelFrame(frame, text="Gabarito", bg=self.co0, font=("Arial", 10, "bold"))
                frame_gab.pack(fill="x", padx=10, pady=5)
                tk.Label(frame_gab, text=f"Resposta correta: {questao.gabarito_letra}", bg=self.co0, font=("Arial", 11, "bold")).pack(padx=10, pady=5)
            
            if questao.gabarito_dissertativa:
                frame_gab = tk.LabelFrame(frame, text="Resposta Esperada", bg=self.co0, font=("Arial", 10, "bold"))
                frame_gab.pack(fill="x", padx=10, pady=5)
                
                txt_gab = tk.Text(frame_gab, wrap="word", font=("Arial", 10), height=3)
                txt_gab.insert("1.0", questao.gabarito_dissertativa)
                txt_gab.config(state="disabled")
                txt_gab.pack(fill="x", padx=5, pady=5)
            
            # Botão fechar
            tk.Button(
                frame, text="Fechar", command=janela_detalhes.destroy,
                bg=self.co9, fg="white", font=("Arial", 10)
            ).pack(pady=15)
            
        except Exception as e:
            logger.error(f"Erro ao abrir detalhes: {e}")
            messagebox.showerror("Erro", f"Erro ao carregar questão: {e}")
    
    def _carregar_arquivos_questao(self, questao_id: int) -> list:
        """Carrega arquivos (imagens) vinculados a uma questão."""
        try:
            from conexao import conectar_bd
            conn = conectar_bd()
            if not conn:
                return []
            
            cursor: Any = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT qa.*, qalt.letra as alternativa_letra
                FROM questoes_arquivos qa
                LEFT JOIN questoes_alternativas qalt ON qa.alternativa_id = qalt.id
                WHERE qa.questao_id = %s
                ORDER BY qa.posicao, qa.id
            """, (questao_id,))
            
            arquivos = cursor.fetchall()
            cursor.close()
            conn.close()
            
            return arquivos
            
        except Exception as e:
            logger.error(f"Erro ao carregar arquivos da questão: {e}")
            return []
    
    def _mostrar_imagem_questao(self, parent: tk.Widget, arquivo: dict, max_size: int = 300, posicao = None):
        """Mostra uma imagem da questão no widget pai. Busca do Drive se não encontrar localmente."""
        import os
        
        try:
            from PIL import Image, ImageTk
            
            caminho_relativo = arquivo.get('caminho_relativo', '')
            link_drive = arquivo.get('link_no_drive', '')
            drive_file_id = arquivo.get('drive_file_id', '')
            
            # Construir caminho absoluto
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            caminho_absoluto = os.path.join(base_dir, caminho_relativo.replace("/", os.sep)) if caminho_relativo else None
            
            img = None
            
            # Tentar carregar do arquivo local primeiro
            if caminho_absoluto and os.path.exists(caminho_absoluto):
                img = Image.open(caminho_absoluto)
            elif drive_file_id or link_drive:
                # Tentar baixar do Google Drive
                img = self._baixar_imagem_drive(drive_file_id, link_drive, caminho_absoluto)
            
            if not img:
                logger.warning(f"Imagem não encontrada localmente nem no Drive")
                return
            
            # Redimensionar mantendo proporção
            ratio = min(max_size / img.width, max_size / img.height)
            if ratio < 1:
                novo_w = int(img.width * ratio)
                novo_h = int(img.height * ratio)
                img = img.resize((novo_w, novo_h), Image.Resampling.LANCZOS)
            
            photo = ImageTk.PhotoImage(img)
            
            frame_img = tk.Frame(parent, bg=self.co0)
            if posicao == "top":
                frame_img.pack(fill="x", pady=5)
            elif posicao == "bottom":
                frame_img.pack(fill="x", pady=5)
            else:
                frame_img.pack(side="right", padx=5)
            
            lbl = tk.Label(frame_img, image=photo, bg=self.co0)
            # Manter referência ao PhotoImage para evitar garbage collection
            self._store_image_ref(photo)
            lbl.pack()
            
        except ImportError:
            logger.warning("PIL não disponível para exibir imagem")
        except Exception as e:
            logger.error(f"Erro ao mostrar imagem: {e}")
    
    def _baixar_imagem_drive(self, drive_file_id: str, link_drive: str, caminho_local: Optional[str] = None):
        """
        Baixa uma imagem do Google Drive e salva localmente como cache.
        
        Args:
            drive_file_id: ID do arquivo no Drive
            link_drive: Link alternativo do Drive
            caminho_local: Caminho onde salvar como cache (opcional)
            
        Returns:
            PIL.Image ou None em caso de erro
        """
        import os
        import io
        
        try:
            from PIL import Image
            from utilitarios.gerenciador_documentos import gerenciador
            
            # Garantir que o Drive esteja configurado
            if not gerenciador.service:
                gerenciador.setup_google_drive()
            
            file_id = drive_file_id
            if not file_id and link_drive:
                # Extrair file_id do link
                if '/d/' in link_drive:
                    file_id = link_drive.split('/d/')[1].split('/')[0]
                elif 'id=' in link_drive:
                    file_id = link_drive.split('id=')[1].split('&')[0]
            
            if not file_id:
                logger.warning("Não foi possível obter ID do arquivo no Drive")
                return None
            
            # Baixar arquivo do Drive
            svc = gerenciador.service
            if not svc:
                logger.warning("Drive service não inicializado ao tentar baixar arquivo")
                return None

            request = svc.files().get_media(fileId=file_id)
            file_data = io.BytesIO()

            from googleapiclient.http import MediaIoBaseDownload
            downloader = MediaIoBaseDownload(file_data, request)

            done = False
            while not done:
                status, done = downloader.next_chunk()
            
            file_data.seek(0)
            img = Image.open(file_data)
            
            # Salvar como cache local se caminho fornecido
            if caminho_local:
                try:
                    os.makedirs(os.path.dirname(caminho_local), exist_ok=True)
                    img.save(caminho_local)
                    logger.info(f"Imagem baixada do Drive e salva em cache: {caminho_local}")
                except Exception as e:
                    logger.warning(f"Não foi possível salvar cache local: {e}")
            
            return img
            
        except Exception as e:
            logger.error(f"Erro ao baixar imagem do Drive: {e}")
            return None
    
    # =========================================================================
    # ABA: CADASTRAR QUESTÃO
    # =========================================================================
    
    def criar_aba_cadastro(self):
        """Cria a aba de cadastro de nova questão."""
        # Frame com scroll
        canvas = tk.Canvas(self.frame_cadastro, bg=self.co0)
        scrollbar = ttk.Scrollbar(self.frame_cadastro, orient="vertical", command=canvas.yview)
        frame_scroll = tk.Frame(canvas, bg=self.co0)
        
        frame_scroll.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=frame_scroll, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Habilitar scroll com roda do mouse
        def _on_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                pass  # Canvas foi destruído, ignorar evento
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Seção: Dados básicos
        frame_dados = tk.LabelFrame(
            frame_scroll, text="Dados da Questão",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_dados.pack(fill="x", padx=10, pady=10)
        
        # Componente
        tk.Label(frame_dados, text="Componente Curricular:*", bg=self.co0).grid(
            row=0, column=0, padx=5, pady=5, sticky="w"
        )
        self.cad_componente = ttk.Combobox(frame_dados, width=25, state="readonly")
        self.cad_componente['values'] = [
            "Língua Portuguesa", "Matemática", "Ciências",
            "Geografia", "História", "Arte", "Educação Física",
            "Língua Inglesa", "Ensino Religioso", "Computação"
        ]
        self.cad_componente.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.cad_componente.bind("<<ComboboxSelected>>", self._atualizar_habilidades_cadastro)
        
        # Ano
        tk.Label(frame_dados, text="Ano Escolar:*", bg=self.co0).grid(
            row=0, column=2, padx=5, pady=5, sticky="w"
        )
        self.cad_ano = ttk.Combobox(frame_dados, width=15, state="readonly")
        self.cad_ano['values'] = [
            "1º ano", "2º ano", "3º ano", "4º ano", "5º ano",
            "6º ano", "7º ano", "8º ano", "9º ano"
        ]
        self.cad_ano.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.cad_ano.bind("<<ComboboxSelected>>", self._atualizar_habilidades_cadastro)
        
        # Habilidade BNCC
        tk.Label(frame_dados, text="Habilidade BNCC:*", bg=self.co0).grid(
            row=1, column=0, padx=5, pady=5, sticky="w"
        )
        self.cad_habilidade = ttk.Combobox(frame_dados, width=60, state="readonly")
        self.cad_habilidade['values'] = [""]
        self.cad_habilidade.grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky="w")
        tk.Label(frame_dados, text="Selecione Componente e Ano para filtrar", bg=self.co0, fg=self.co9).grid(
            row=1, column=4, padx=5, pady=5, sticky="w"
        )
        
        # Tipo
        tk.Label(frame_dados, text="Tipo:*", bg=self.co0).grid(
            row=2, column=0, padx=5, pady=5, sticky="w"
        )
        self.cad_tipo = ttk.Combobox(frame_dados, width=20, state="readonly")
        self.cad_tipo['values'] = [
            "multipla_escolha", "dissertativa", "verdadeiro_falso"
        ]
        self.cad_tipo.current(0)
        self.cad_tipo.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.cad_tipo.bind("<<ComboboxSelected>>", self.atualizar_campos_tipo)
        
        # Dificuldade
        tk.Label(frame_dados, text="Dificuldade:*", bg=self.co0).grid(
            row=2, column=2, padx=5, pady=5, sticky="w"
        )
        self.cad_dificuldade = ttk.Combobox(frame_dados, width=15, state="readonly")
        self.cad_dificuldade['values'] = ["facil", "media", "dificil"]
        self.cad_dificuldade.current(1)
        self.cad_dificuldade.grid(row=2, column=3, padx=5, pady=5, sticky="w")
        
        # Seção: Enunciado e Imagem
        frame_enunciado = tk.LabelFrame(
            frame_scroll, text="Enunciado*",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_enunciado.pack(fill="x", padx=10, pady=10)
        
        # Container para enunciado + imagem (layout flexível)
        self.frame_enunciado_container = tk.Frame(frame_enunciado, bg=self.co0)
        self.frame_enunciado_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Frame do texto do enunciado
        self.frame_texto_enunciado = tk.Frame(self.frame_enunciado_container, bg=self.co0)
        self.frame_texto_enunciado.pack(side="left", fill="both", expand=True)
        
        self.cad_enunciado = tk.Text(self.frame_texto_enunciado, height=5, wrap="word")
        self.cad_enunciado.pack(fill="both", expand=True)
        
        # Seção: Imagem do Enunciado
        frame_imagem_enunciado = tk.LabelFrame(
            frame_scroll, text="📷 Imagem do Enunciado (opcional)",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_imagem_enunciado.pack(fill="x", padx=10, pady=10)
        
        # Linha 1: Posição da imagem
        frame_pos_img = tk.Frame(frame_imagem_enunciado, bg=self.co0)
        frame_pos_img.pack(fill="x", padx=5, pady=5)
        
        tk.Label(frame_pos_img, text="Posição da imagem:", bg=self.co0).pack(side="left", padx=5)
        
        self.var_posicao_img = tk.StringVar(value="abaixo")
        
        for pos_val, pos_txt in [("acima", "⬆️ Acima"), ("abaixo", "⬇️ Abaixo"), 
                                  ("esquerda", "⬅️ Esquerda"), ("direita", "➡️ Direita"), ("entre", "↔️ Entre")]:
            tk.Radiobutton(
                frame_pos_img, text=pos_txt, variable=self.var_posicao_img,
                value=pos_val, bg=self.co0
            ).pack(side="left", padx=5)
        
        # Linha 2: Botão selecionar imagem e preview
        frame_btn_img = tk.Frame(frame_imagem_enunciado, bg=self.co0)
        frame_btn_img.pack(fill="x", padx=5, pady=5)
        
        tk.Button(
            frame_btn_img, text="📂 Selecionar Imagem",
            command=self.selecionar_imagem_enunciado,
            bg=self.co4, fg="white"
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_btn_img, text="✏️ Editar Imagem",
            command=self.editar_imagem_enunciado,
            bg="#FF9800", fg="white"
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_btn_img, text="🔍 Ampliar Preview",
            command=lambda: self.ampliar_preview(self.imagem_enunciado_path) if self.imagem_enunciado_path else None,
            bg=self.co4, fg="white"
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_btn_img, text="🗑️ Remover Imagem",
            command=self.remover_imagem_enunciado,
            bg=self.co8, fg="white"
        ).pack(side="left", padx=5)

        tk.Button(
            frame_btn_img, text="📎 Inserir no Texto",
            command=self.inserir_imagem_enunciado_entre,
            bg=self.co4, fg="white"
        ).pack(side="left", padx=5)
        
        self.lbl_imagem_enunciado = tk.Label(
            frame_btn_img, text="Nenhuma imagem selecionada",
            bg=self.co0, fg=self.co9, font=("Arial", 9, "italic")
        )
        self.lbl_imagem_enunciado.pack(side="left", padx=10)
        
        # Preview da imagem (thumbnail)
        self.frame_preview_enunciado = tk.Frame(frame_imagem_enunciado, bg=self.co0)
        self.frame_preview_enunciado.pack(fill="x", padx=5, pady=5)
        
        self.lbl_preview_enunciado = tk.Label(self.frame_preview_enunciado, bg=self.co0)
        self.lbl_preview_enunciado.pack()
        
        # Variável para armazenar o caminho da imagem
        self.imagem_enunciado_path = None
        self.imagem_enunciado_preview = None  # PhotoImage reference
        self.imagem_enunciado_inline = False  # Se a imagem foi inserida inline no Text
        
        # Seção: Texto de apoio (opcional)
        frame_apoio = tk.LabelFrame(
            frame_scroll, text="Texto de Apoio (opcional)",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_apoio.pack(fill="x", padx=10, pady=10)
        
        self.cad_texto_apoio = tk.Text(frame_apoio, height=3, wrap="word")
        self.cad_texto_apoio.pack(fill="x", padx=5, pady=5)
        
        # Seção: Alternativas (para múltipla escolha)
        self.frame_alternativas = tk.LabelFrame(
            frame_scroll, text="Alternativas",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        self.frame_alternativas.pack(fill="x", padx=10, pady=10)
        
        self.criar_campos_alternativas()
        
        # Seção: Imagens das Alternativas (para múltipla escolha)
        self.frame_imagens_alternativas = tk.LabelFrame(
            frame_scroll, text="📷 Imagens das Alternativas (opcional)",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        self.frame_imagens_alternativas.pack(fill="x", padx=10, pady=10)
        
        self.criar_campos_imagens_alternativas()
        
        # Seção: Verdadeiro/Falso (para questões V/F)
        self.frame_vf = tk.LabelFrame(
            frame_scroll, text="Afirmações (V ou F)",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        self.frame_vf.pack(fill="x", padx=10, pady=10)
        self.frame_vf.pack_forget()  # Esconder inicialmente
        
        self.criar_campos_vf()
        
        # Seção: Gabarito dissertativa
        self.frame_gabarito_diss = tk.LabelFrame(
            frame_scroll, text="Gabarito / Resposta Esperada",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        self.frame_gabarito_diss.pack(fill="x", padx=10, pady=10)
        self.frame_gabarito_diss.pack_forget()  # Esconder inicialmente
        
        self.cad_gabarito_diss = tk.Text(self.frame_gabarito_diss, height=3, wrap="word")
        self.cad_gabarito_diss.pack(fill="x", padx=5, pady=5)
        
        # Botões
        frame_botoes = tk.Frame(frame_scroll, bg=self.co0)
        frame_botoes.pack(fill="x", padx=10, pady=20)
        
        tk.Button(
            frame_botoes, text="💾 Salvar como Rascunho",
            command=lambda: self.salvar_questao("rascunho"),
            bg=self.co9, fg="white", font=("Arial", 10, "bold")
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_botoes, text="✅ Salvar e Enviar para Revisão",
            command=lambda: self.salvar_questao("revisao"),
            bg=self.co2, fg="white", font=("Arial", 10, "bold")
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_botoes, text="🧹 Limpar Campos",
            command=self.limpar_campos_cadastro,
            bg=self.co8, fg="white", font=("Arial", 10, "bold")
        ).pack(side="right", padx=5)
        
        tk.Button(
            frame_botoes, text="� Importar Excel",
            command=self.importar_questoes_excel,
            bg="#8B4513", fg="white", font=("Arial", 10, "bold")
        ).pack(side="right", padx=5)
        
        tk.Button(
            frame_botoes, text="�📋 Salvar Template",
            command=self.salvar_como_template,
            bg=self.co4, fg="white", font=("Arial", 10, "bold")
        ).pack(side="right", padx=5)
        
        tk.Button(
            frame_botoes, text="📂 Carregar Template",
            command=self.carregar_template,
            bg=self.co4, fg="white", font=("Arial", 10, "bold")
        ).pack(side="right", padx=5)
        
        # Habilitar drag & drop se disponível
        self.habilitar_drag_drop()
    
    def criar_campos_alternativas(self):
        """Cria os campos para alternativas de múltipla escolha."""
        self.cad_alternativas = {}
        self.var_gabarito = tk.StringVar(value="A")
        
        for letra in ["A", "B", "C", "D", "E"]:
            frame_alt = tk.Frame(self.frame_alternativas, bg=self.co0)
            frame_alt.pack(fill="x", padx=5, pady=2)
            
            rb = tk.Radiobutton(
                frame_alt, text=f"{letra})", variable=self.var_gabarito,
                value=letra, bg=self.co0
            )
            rb.pack(side="left")
            
            entry = ttk.Entry(frame_alt, width=70)
            entry.pack(side="left", fill="x", expand=True, padx=5)
            
            self.cad_alternativas[letra] = entry
        
        tk.Label(
            self.frame_alternativas,
            text="Selecione o botão da alternativa correta",
            bg=self.co0, fg=self.co9, font=("Arial", 9, "italic")
        ).pack(pady=5)
    
    def criar_campos_imagens_alternativas(self):
        """Cria campos para selecionar imagens das alternativas."""
        self.imagens_alternativas = {}  # Dict[letra, path]
        self.imagens_alternativas_preview = {}  # Dict[letra, PhotoImage]
        self.labels_imagem_alt = {}
        self.labels_preview_alt = {}
        
        # Info
        tk.Label(
            self.frame_imagens_alternativas,
            text="Adicione imagens ao lado de cada alternativa (as imagens aparecerão à direita do texto)",
            bg=self.co0, fg=self.co9, font=("Arial", 9, "italic")
        ).pack(pady=5)
        
        frame_grid = tk.Frame(self.frame_imagens_alternativas, bg=self.co0)
        frame_grid.pack(fill="x", padx=5, pady=5)
        
        for i, letra in enumerate(["A", "B", "C", "D", "E"]):
            frame_alt_img = tk.Frame(frame_grid, bg=self.co0)
            frame_alt_img.grid(row=i, column=0, sticky="w", padx=5, pady=2)
            
            tk.Label(frame_alt_img, text=f"{letra})", bg=self.co0, width=3).pack(side="left")
            
            tk.Button(
                frame_alt_img, text="📂",
                command=lambda l=letra: self.selecionar_imagem_alternativa(l),
                bg=self.co4, fg="white", width=3
            ).pack(side="left", padx=2)
            
            tk.Button(
                frame_alt_img, text="✏️",
                command=lambda l=letra: self.editar_imagem_alternativa(l),
                bg="#FF9800", fg="white", width=3
            ).pack(side="left", padx=2)
            
            tk.Button(
                frame_alt_img, text="🗑️",
                command=lambda l=letra: self.remover_imagem_alternativa(l),
                bg=self.co8, fg="white", width=3
            ).pack(side="left", padx=2)
            
            lbl = tk.Label(
                frame_alt_img, text="Sem imagem",
                bg=self.co0, fg=self.co9, font=("Arial", 8), width=20, anchor="w"
            )
            lbl.pack(side="left", padx=5)
            self.labels_imagem_alt[letra] = lbl
            
            # Mini preview
            lbl_prev = tk.Label(frame_alt_img, bg=self.co0)
            lbl_prev.pack(side="left", padx=5)
            self.labels_preview_alt[letra] = lbl_prev
            
            self.imagens_alternativas[letra] = None
    
    def selecionar_imagem_enunciado(self):
        """Abre diálogo para selecionar imagem do enunciado."""
        from tkinter import filedialog
        
        filetypes = [
            ("Imagens", "*.png *.jpg *.jpeg *.gif *.bmp"),
            ("PNG", "*.png"),
            ("JPEG", "*.jpg *.jpeg"),
            ("GIF", "*.gif"),
            ("Todos", "*.*")
        ]
        
        caminho = filedialog.askopenfilename(
            title="Selecionar Imagem do Enunciado",
            filetypes=filetypes,
            parent=self.janela
        )
        
        if caminho:
            # Validar tamanho da imagem
            if not self.validar_tamanho_imagem(caminho):
                return
            
            self.imagem_enunciado_path = caminho
            nome_arquivo = caminho.split("/")[-1].split("\\")[-1]
            self.lbl_imagem_enunciado.config(text=f"✅ {nome_arquivo[:30]}...")
            self.mostrar_preview_imagem(caminho, self.lbl_preview_enunciado, 150)

    def inserir_imagem_enunciado_entre(self):
        """Insere uma imagem no ponto atual do cursor dentro do enunciado (inline)."""
        from tkinter import filedialog

        filetypes = [
            ("Imagens", "*.png *.jpg *.jpeg *.gif *.bmp"),
            ("PNG", "*.png"),
            ("JPEG", "*.jpg *.jpeg"),
            ("GIF", "*.gif"),
            ("Todos", "*.*")
        ]

        caminho = filedialog.askopenfilename(
            title="Inserir Imagem no Enunciado",
            filetypes=filetypes,
            parent=self.janela
        )

        if not caminho:
            return

        try:
            from PIL import Image, ImageTk

            # Carregar imagem e reduzir largura se for muito larga para o editor
            img = Image.open(caminho)
            max_w = 600
            if img.width > max_w:
                ratio = max_w / img.width
                img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.Resampling.LANCZOS)

            photo = ImageTk.PhotoImage(img)

            # Inserir imagem no Text no índice atual do cursor
            try:
                self.cad_enunciado.image_create("insert", image=photo)
            except Exception:
                # fallback: append at the end
                self.cad_enunciado.image_create("end", image=photo)

            # Manter referência ao PhotoImage
            self._store_image_ref(photo)

            # Guardar informação para salvar posteriormente
            self.imagem_enunciado_path = caminho
            self.imagem_enunciado_inline = True
            nome_arquivo = caminho.split("/")[-1].split("\\")[-1]
            self.lbl_imagem_enunciado.config(text=f"✅ {nome_arquivo[:30]}... (inline)")
            # Também atualizar o preview menor
            self.mostrar_preview_imagem(caminho, self.lbl_preview_enunciado, 150)

            # Forçar opção de posição para 'entre'
            self.var_posicao_img.set("entre")

        except ImportError:
            messagebox.showwarning("Aviso", "PIL não disponível — não é possível inserir a imagem inline.")
        except Exception as e:
            logger.error(f"Erro ao inserir imagem inline: {e}")
            messagebox.showerror("Erro", f"Erro ao inserir imagem no enunciado: {e}")
    
    def remover_imagem_enunciado(self):
        """Remove a imagem do enunciado."""
        self.imagem_enunciado_path = None
        self.imagem_enunciado_preview = None
        self.lbl_imagem_enunciado.config(text="Nenhuma imagem selecionada")
        self.lbl_preview_enunciado.config(image='')
    
    def selecionar_imagem_alternativa(self, letra: str):
        """Abre diálogo para selecionar imagem de uma alternativa."""
        from tkinter import filedialog
        
        filetypes = [
            ("Imagens", "*.png *.jpg *.jpeg *.gif *.bmp"),
            ("PNG", "*.png"),
            ("JPEG", "*.jpg *.jpeg"),
            ("GIF", "*.gif"),
            ("Todos", "*.*")
        ]
        
        caminho = filedialog.askopenfilename(
            title=f"Selecionar Imagem da Alternativa {letra}",
            filetypes=filetypes,
            parent=self.janela
        )
        
        if caminho:
            # Validar tamanho da imagem
            if not self.validar_tamanho_imagem(caminho):
                return
            
            self.imagens_alternativas[letra] = caminho
            nome_arquivo = caminho.split("/")[-1].split("\\")[-1]
            self.labels_imagem_alt[letra].config(text=f"✅ {nome_arquivo[:15]}...")
            self.mostrar_preview_imagem(caminho, self.labels_preview_alt[letra], 40)
    
    def remover_imagem_alternativa(self, letra: str):
        """Remove a imagem de uma alternativa."""
        self.imagens_alternativas[letra] = None
        if letra in self.imagens_alternativas_preview:
            del self.imagens_alternativas_preview[letra]
        self.labels_imagem_alt[letra].config(text="Sem imagem")
        self.labels_preview_alt[letra].config(image='')
    
    def mostrar_preview_imagem(self, caminho: str, label: tk.Label, tamanho_max: int = 100):
        """Mostra preview de uma imagem em um label (com cache)."""
        try:
            # Verificar cache
            cache_key = f"{caminho}_{tamanho_max}"
            if cache_key in self._cache_imagens:
                photo = self._cache_imagens[cache_key]
                label.config(image=photo)
                return
            
            from PIL import Image, ImageTk
            
            img = Image.open(caminho)
            
            # Redimensionar mantendo proporção
            ratio = min(tamanho_max / img.width, tamanho_max / img.height)
            novo_w = int(img.width * ratio)
            novo_h = int(img.height * ratio)
            
            img_resized = img.resize((novo_w, novo_h), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img_resized)
            
            # Armazenar no cache
            self._cache_imagens[cache_key] = photo
            
            label.config(image=photo)
            # Manter referência ao PhotoImage para evitar garbage collection
            self._store_image_ref(photo)
            
        except ImportError:
            logger.warning("PIL não disponível para preview de imagem")
            label.config(text="[Preview indisponível]")
        except Exception as e:
            logger.error(f"Erro ao mostrar preview: {e}")
            label.config(text="[Erro no preview]")

    def _store_image_ref(self, photo):
        """Armazena referência ao PhotoImage para evitar garbage collection."""
        try:
            self._image_refs.append(photo)
        except Exception:
            # Caso a lista não exista por algum motivo, crie-a
            self._image_refs = [photo]
    
    def criar_campos_vf(self):
        """Cria os campos para questões de Verdadeiro/Falso."""
        self.vf_afirmacoes = []  # Lista de (Entry, var_resposta)
        self.frame_vf_itens = tk.Frame(self.frame_vf, bg=self.co0)
        self.frame_vf_itens.pack(fill="x", padx=5, pady=5)
        
        tk.Label(
            self.frame_vf,
            text="Adicione as afirmações e marque V (Verdadeiro) ou F (Falso)",
            bg=self.co0, fg=self.co9, font=("Arial", 9, "italic")
        ).pack(pady=5)
        
        # Criar 5 campos inicialmente
        for i in range(5):
            self._adicionar_campo_vf()
        
        # Botão para adicionar mais
        tk.Button(
            self.frame_vf, text="➕ Adicionar Afirmação",
            command=self._adicionar_campo_vf,
            bg=self.co4, fg="white"
        ).pack(pady=5)
    
    def _adicionar_campo_vf(self):
        """Adiciona um campo de afirmação V/F."""
        frame_item = tk.Frame(self.frame_vf_itens, bg=self.co0)
        frame_item.pack(fill="x", pady=2)
        
        idx = len(self.vf_afirmacoes) + 1
        tk.Label(frame_item, text=f"{idx}.", bg=self.co0, width=3).pack(side="left")
        
        entry = ttk.Entry(frame_item, width=50)
        entry.pack(side="left", fill="x", expand=True, padx=5)
        
        var_vf = tk.StringVar(value="V")
        tk.Radiobutton(frame_item, text="V", variable=var_vf, value="V", bg=self.co0).pack(side="left")
        tk.Radiobutton(frame_item, text="F", variable=var_vf, value="F", bg=self.co0).pack(side="left")
        
        self.vf_afirmacoes.append((entry, var_vf))
    
    def atualizar_campos_tipo(self, event=None):
        """Atualiza campos visíveis baseado no tipo de questão."""
        tipo = self.cad_tipo.get()
        
        # Esconder todos os frames específicos primeiro
        self.frame_alternativas.pack_forget()
        self.frame_imagens_alternativas.pack_forget()
        self.frame_vf.pack_forget()
        self.frame_gabarito_diss.pack_forget()
        
        if tipo == "multipla_escolha":
            self.frame_alternativas.pack(fill="x", padx=10, pady=10)
            self.frame_imagens_alternativas.pack(fill="x", padx=10, pady=10)
        elif tipo == "dissertativa":
            self.frame_gabarito_diss.pack(fill="x", padx=10, pady=10)
        elif tipo == "verdadeiro_falso":
            self.frame_vf.pack(fill="x", padx=10, pady=10)
    
    def salvar_questao(self, status: str):
        """Salva a questão no banco de dados (cria nova ou atualiza existente)."""
        from banco_questoes.services import QuestaoService
        from banco_questoes.models import (
            Questao, QuestaoAlternativa, QuestaoArquivo, ComponenteCurricular, AnoEscolar,
            TipoQuestao, DificuldadeQuestao, StatusQuestao, VisibilidadeQuestao, TipoArquivo
        )
        
        # Verificar se está editando (se tem ID armazenado)
        editando = hasattr(self, '_questao_id_edicao') and self._questao_id_edicao
        
        # Validar campos obrigatórios
        if not self.cad_componente.get():
            messagebox.showwarning("Aviso", "Selecione o componente curricular.")
            return
        
        if not self.cad_ano.get():
            messagebox.showwarning("Aviso", "Selecione o ano escolar.")
            return
        
        # VALIDAÇÃO OBRIGATÓRIA: Habilidade BNCC (conforme especificação BNCC)
        if not self.cad_habilidade.get().strip():
            messagebox.showerror(
                "Campo Obrigatório",
                "⚠️ Habilidade BNCC é obrigatória!\n\n"
                "Todas as questões devem estar vinculadas a pelo menos uma habilidade da BNCC.\n\n"
                "Selecione o Componente e o Ano primeiro para filtrar as habilidades disponíveis."
            )
            return
        
        # Extrair código da habilidade (antes de " - " se houver descrição)
        habilidade_selecionada = self.cad_habilidade.get().strip()
        if " - " in habilidade_selecionada:
            habilidade_codigo = habilidade_selecionada.split(" - ")[0].strip().upper()
        else:
            habilidade_codigo = habilidade_selecionada.upper()
        
        enunciado = self.cad_enunciado.get("1.0", "end").strip()
        if not enunciado:
            messagebox.showwarning("Aviso", "Informe o enunciado da questão.")
            return
        
        if not self.funcionario_id:
            messagebox.showwarning("Aviso", "É necessário estar logado para salvar questões.")
            return
        
        try:
            # Mapear status
            status_map = {
                'rascunho': StatusQuestao.RASCUNHO,
                'revisao': StatusQuestao.REVISAO
            }
            
            # Criar objeto Questao
            questao = Questao(
                componente_curricular=ComponenteCurricular(self.cad_componente.get()),
                ano_escolar=AnoEscolar(self.cad_ano.get()),
                habilidade_bncc_codigo=habilidade_codigo,
                tipo=TipoQuestao(self.cad_tipo.get()),
                dificuldade=DificuldadeQuestao(self.cad_dificuldade.get()),
                enunciado=enunciado,
                texto_apoio=self.cad_texto_apoio.get("1.0", "end").strip() or None,
                status=status_map.get(status, StatusQuestao.RASCUNHO),
                visibilidade=VisibilidadeQuestao.ESCOLA,
                autor_id=self.funcionario_id,
                escola_id=config.ESCOLA_ID
            )
            
            # Preparar arquivos (imagens)
            arquivos_questao = []
            
            # Imagem do enunciado
            if self.imagem_enunciado_path:
                # Se a posição selecionada for 'entre' (inline), o DB pode não ter enum
                # correspondente; salvo como 'abaixo' no banco e marco 'inline' no metadado.
                pos_sel = self.var_posicao_img.get()
                pos_for_db = 'abaixo' if pos_sel == 'entre' else pos_sel
                arquivo_enunciado = self._preparar_arquivo_imagem(
                    self.imagem_enunciado_path,
                    posicao_imagem=pos_for_db
                )
                if arquivo_enunciado and pos_sel == 'entre':
                    arquivo_enunciado['inline'] = True
                if arquivo_enunciado:
                    arquivos_questao.append(arquivo_enunciado)
            
            # Gabarito para múltipla escolha
            if self.cad_tipo.get() == "multipla_escolha":
                questao.gabarito_letra = self.var_gabarito.get()
                
                # Coletar alternativas
                alternativas = []
                for letra, entry in self.cad_alternativas.items():
                    texto = entry.get().strip()
                    if texto:
                        alt = QuestaoAlternativa(
                            letra=letra,
                            texto=texto,
                            correta=(letra == self.var_gabarito.get()),
                            ordem=ord(letra) - ord('A') + 1
                        )
                        alternativas.append(alt)
                
                if len(alternativas) < 2:
                    messagebox.showwarning("Aviso", "Informe pelo menos 2 alternativas.")
                    return
                
                questao.alternativas = alternativas
                
                # Imagens das alternativas
                for letra, caminho in self.imagens_alternativas.items():
                    if caminho:
                        arquivo_alt = self._preparar_arquivo_imagem(
                            caminho,
                            alternativa_letra=letra,
                            posicao_imagem="direita"
                        )
                        if arquivo_alt:
                            arquivos_questao.append(arquivo_alt)
            
            # Gabarito para dissertativa
            elif self.cad_tipo.get() == "dissertativa":
                questao.gabarito_dissertativa = self.cad_gabarito_diss.get("1.0", "end").strip()
            
            questao.arquivos = arquivos_questao
            
            # CRIAR OU ATUALIZAR
            if editando:
                questao.id = self._questao_id_edicao
                
                # VERSIONAMENTO: Registrar histórico antes de atualizar
                try:
                    QuestaoService.registrar_historico(
                        questao_id=self._questao_id_edicao,
                        usuario_id=self.funcionario_id,
                        motivo="Edição manual via interface"
                    )
                except Exception as e:
                    logger.warning(f"Não foi possível registrar histórico: {e}")
                
                sucesso = QuestaoService.atualizar(questao)
                
                if sucesso:
                    # Atualizar arquivos se houver
                    if arquivos_questao:
                        self._salvar_arquivos_questao(self._questao_id_edicao, arquivos_questao)
                    
                    messagebox.showinfo(
                        "Sucesso",
                        f"✅ Questão #{self._questao_id_edicao} atualizada com sucesso!\n\nO histórico de alterações foi registrado."
                    )
                    self.limpar_campos_cadastro()
                    self._questao_id_edicao = None  # Limpar flag de edição
                else:
                    messagebox.showerror("Erro", "Não foi possível atualizar a questão.")
            else:
                questao_id = QuestaoService.criar(questao)
                
                if questao_id:
                    # Salvar arquivos de imagem
                    if arquivos_questao:
                        self._salvar_arquivos_questao(questao_id, arquivos_questao)
                    
                    messagebox.showinfo(
                        "Sucesso",
                        f"Questão #{questao_id} salva com sucesso!\nStatus: {status}"
                    )
                    self.limpar_campos_cadastro()
                else:
                    messagebox.showerror("Erro", "Não foi possível salvar a questão.")
            
        except Exception as e:
            logger.error(f"Erro ao salvar questão: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar questão: {e}")
    
    def _preparar_arquivo_imagem(self, caminho: str, alternativa_letra = None, posicao_imagem: str = "abaixo"):
        """Prepara dados de um arquivo de imagem para salvar."""
        import os
        
        if not caminho or not os.path.exists(caminho):
            return None
        
        try:
            nome_original = os.path.basename(caminho)
            tamanho = os.path.getsize(caminho)
            
            # Detectar mime type
            ext = os.path.splitext(caminho)[1].lower()
            mime_types = {
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.gif': 'image/gif',
                '.bmp': 'image/bmp'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')
            
            # Obter dimensões
            largura, altura = None, None
            try:
                from PIL import Image
                with Image.open(caminho) as img:
                    largura, altura = img.size
            except ImportError:
                pass
            except Exception as e:
                logger.warning(f"Não foi possível obter dimensões da imagem: {e}")
            
            return {
                'caminho_origem': caminho,
                'nome_original': nome_original,
                'tamanho_bytes': tamanho,
                'mime_type': mime_type,
                'largura': largura,
                'altura': altura,
                'alternativa_letra': alternativa_letra,
                'posicao_imagem': posicao_imagem
            }
            
        except Exception as e:
            logger.error(f"Erro ao preparar arquivo de imagem: {e}")
            return None
    
    def _salvar_arquivos_questao(self, questao_id: int, arquivos: list):
        """Salva os arquivos de uma questão no sistema de arquivos, Google Drive e banco de dados."""
        import os
        import shutil
        import hashlib
        from datetime import datetime
        
        try:
            # Diretório base para uploads (backup local)
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            uploads_dir = os.path.join(base_dir, "uploads", "questoes", str(datetime.now().year), str(datetime.now().month).zfill(2))
            
            if not os.path.exists(uploads_dir):
                os.makedirs(uploads_dir)
            
            from conexao import conectar_bd
            conn = conectar_bd()
            if not conn:
                logger.error("Não foi possível conectar ao BD para salvar arquivos")
                return
            
            cursor = conn.cursor()
            
            # Configurar Google Drive
            drive_service = None
            pasta_drive_id = None
            try:
                from utilitarios.gerenciador_documentos import gerenciador
                if not gerenciador.service:
                    gerenciador.setup_google_drive()
                drive_service = gerenciador.service
                
                # Criar estrutura de pastas no Drive: Sistema Escolar - Documentos / Banco de Questões / YYYY / MM
                ano_atual = str(datetime.now().year)
                mes_atual = str(datetime.now().month).zfill(2)
                
                pasta_banco = gerenciador.get_or_create_folder("Banco de Questões - Imagens", gerenciador.pasta_raiz_id)
                pasta_ano = gerenciador.get_or_create_folder(ano_atual, pasta_banco)
                pasta_drive_id = gerenciador.get_or_create_folder(mes_atual, pasta_ano)
                
                logger.info(f"Google Drive configurado para upload de imagens de questões")
            except Exception as e:
                logger.warning(f"Google Drive não disponível, salvando apenas localmente: {e}")
                drive_service = None
            
            for i, arq in enumerate(arquivos, 1):
                try:
                    caminho_origem = arq.get('caminho_origem')
                    if not caminho_origem or not os.path.exists(caminho_origem):
                        continue
                    
                    # Gerar nome único
                    nome_original = arq.get('nome_original', 'imagem.jpg')
                    ext = os.path.splitext(nome_original)[1]
                    hash_nome = hashlib.md5(f"{questao_id}_{i}_{datetime.now().timestamp()}".encode()).hexdigest()[:12]
                    nome_armazenado = f"q{questao_id}_{hash_nome}{ext}"
                    
                    # Copiar arquivo localmente (backup)
                    caminho_destino = os.path.join(uploads_dir, nome_armazenado)
                    shutil.copy2(caminho_origem, caminho_destino)
                    
                    # Caminho relativo para o banco
                    caminho_relativo = os.path.join("uploads", "questoes", str(datetime.now().year), 
                                                     str(datetime.now().month).zfill(2), nome_armazenado)
                    
                    # Upload para Google Drive
                    link_drive = None
                    drive_file_id = None
                    if drive_service and pasta_drive_id:
                        try:
                            link_drive, drive_file_id = self._upload_imagem_drive(
                                drive_service, caminho_destino, nome_armazenado, pasta_drive_id
                            )
                            if link_drive:
                                logger.info(f"Imagem enviada para Google Drive: {nome_armazenado}")
                        except Exception as e:
                            logger.warning(f"Falha no upload para Drive, mantendo apenas local: {e}")
                    
                    # Buscar ID da alternativa se aplicável
                    alternativa_id = None
                    alternativa_letra = arq.get('alternativa_letra')
                    if alternativa_letra:
                        cursor.execute("""
                            SELECT id FROM questoes_alternativas 
                            WHERE questao_id = %s AND letra = %s
                        """, (questao_id, alternativa_letra))
                        row = cursor.fetchone()
                        if row:
                            alternativa_id = row[0]
                    
                    # Normalizar/converter valores para tipos primitivos (int/str/None)
                    # Usar conversão segura para int para evitar chamadas int() sobre tipos incompatíveis
                    from decimal import Decimal

                    def _safe_int(v):
                        if v is None:
                            return None
                        if isinstance(v, int):
                            return v
                        if isinstance(v, (float, str, bytes, Decimal)):
                            try:
                                return int(v)
                            except Exception:
                                return None
                        return None

                    questao_id_param = int(questao_id)
                    alternativa_id_param = _safe_int(alternativa_id)
                    tipo_arquivo_param = 'imagem'
                    nome_original_param = str(nome_original)
                    nome_armazenado_param = str(nome_armazenado)
                    caminho_relativo_param = caminho_relativo.replace("\\", "/")
                    tamanho_bytes_param = _safe_int(arq.get('tamanho_bytes', 0)) or 0
                    mime_type_param = str(arq.get('mime_type', 'image/jpeg') or 'image/jpeg')
                    largura_param = _safe_int(arq.get('largura'))
                    altura_param = _safe_int(arq.get('altura'))
                    posicao_param = str(arq.get('posicao_imagem', 'abaixo'))
                    uploaded_by_param = _safe_int(self.funcionario_id)
                    link_drive_param = link_drive if link_drive else None
                    drive_file_id_param = drive_file_id if drive_file_id else None

                    # Inserir no banco (com campos do Google Drive se disponíveis)
                    cursor.execute("""
                        INSERT INTO questoes_arquivos 
                        (questao_id, alternativa_id, tipo_arquivo, nome_original, nome_arquivo, 
                         caminho_relativo, link_no_drive, drive_file_id, tamanho_bytes, mime_type, 
                         largura, altura, posicao, uploaded_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        questao_id_param,
                        alternativa_id_param,
                        tipo_arquivo_param,
                        nome_original_param,
                        nome_armazenado_param,
                        caminho_relativo_param,
                        link_drive_param,
                        drive_file_id_param,
                        tamanho_bytes_param,
                        mime_type_param,
                        largura_param,
                        altura_param,
                        posicao_param,
                        uploaded_by_param
                    ))
                    
                except Exception as e:
                    logger.error(f"Erro ao salvar arquivo {i}: {e}")
                    continue
            
            conn.commit()
            cursor.close()
            conn.close()
            
        except Exception as e:
            logger.error(f"Erro ao salvar arquivos da questão: {e}")
    
    def _upload_imagem_drive(self, drive_service, caminho_arquivo: str, nome_arquivo: str, pasta_id: str):
        """
        Faz upload de uma imagem para o Google Drive.
        
        Args:
            drive_service: Serviço do Google Drive já autenticado
            caminho_arquivo: Caminho local do arquivo
            nome_arquivo: Nome do arquivo para salvar no Drive
            pasta_id: ID da pasta no Drive onde salvar
            
        Returns:
            tuple: (link_do_drive, file_id) ou (None, None) em caso de erro
        """
        import mimetypes
        from googleapiclient.http import MediaFileUpload
        
        try:
            # Detectar mime type
            mime_type = mimetypes.guess_type(caminho_arquivo)[0] or 'image/jpeg'
            
            # Metadados do arquivo
            file_metadata = {
                'name': nome_arquivo,
                'parents': [pasta_id]
            }
            
            # Upload do arquivo
            media = MediaFileUpload(caminho_arquivo, mimetype=mime_type, resumable=True)
            
            file = drive_service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id, webViewLink'
            ).execute()
            
            # Configurar permissão de visualização pública
            permission = {
                'type': 'anyone',
                'role': 'reader',
                'allowFileDiscovery': False
            }
            drive_service.permissions().create(
                fileId=file['id'],
                body=permission
            ).execute()
            
            return file.get('webViewLink'), file.get('id')
            
        except Exception as e:
            logger.error(f"Erro ao fazer upload para Google Drive: {e}")
            return None, None
    
    def limpar_campos_cadastro(self):
        """Limpa todos os campos do cadastro."""
        self.cad_componente.set("")
        self.cad_ano.set("")
        self.cad_habilidade.set("")
        self.cad_habilidade['values'] = [""]  # Resetar valores do combobox
        self.cad_tipo.current(0)
        self.cad_dificuldade.current(1)
        self.cad_enunciado.delete("1.0", "end")
        self.cad_texto_apoio.delete("1.0", "end")
        self.cad_gabarito_diss.delete("1.0", "end")
        
        for entry in self.cad_alternativas.values():
            entry.delete(0, "end")
        
        self.var_gabarito.set("A")
        
        # Limpar imagem do enunciado
        self.remover_imagem_enunciado()
        self.var_posicao_img.set("abaixo")
        
        # Limpar imagens das alternativas
        for letra in ["A", "B", "C", "D", "E"]:
            self.remover_imagem_alternativa(letra)
        
        # Limpar campos V/F
        for entry, var in self.vf_afirmacoes:
            entry.delete(0, "end")
            var.set("V")
        
        # Limpar flag de edição
        if hasattr(self, '_questao_id_edicao'):
            self._questao_id_edicao = None
        
        self.atualizar_campos_tipo()
    
    def carregar_questao_para_edicao(self, questao):
        """Carrega dados de uma questão nos campos para edição."""
        try:
            # Armazenar ID para modo edição
            self._questao_id_edicao = questao.id
            
            # Limpar campos primeiro
            self.limpar_campos_cadastro()
            
            # Preencher dados básicos
            if questao.componente_curricular:
                self.cad_componente.set(questao.componente_curricular.value)
                self.cad_componente.event_generate("<<ComboboxSelected>>")
            
            if questao.ano_escolar:
                self.cad_ano.set(questao.ano_escolar.value)
                self.cad_ano.event_generate("<<ComboboxSelected>>")
            
            if questao.habilidade_bncc_codigo:
                # Aguardar atualização de habilidades
                self.janela.after(100, lambda: self.cad_habilidade.set(questao.habilidade_bncc_codigo))
            
            if questao.tipo:
                self.cad_tipo.set(questao.tipo.value)
                self.atualizar_campos_tipo()
            
            if questao.dificuldade:
                self.cad_dificuldade.set(questao.dificuldade.value)
            
            # Enunciado
            if questao.enunciado:
                self.cad_enunciado.insert("1.0", questao.enunciado)
            
            # Texto de apoio
            if questao.texto_apoio:
                self.cad_texto_apoio.insert("1.0", questao.texto_apoio)
            
            # Carregar alternativas se múltipla escolha
            if questao.tipo and questao.tipo.value == "multipla_escolha":
                from banco_questoes.services import QuestaoService
                alternativas = QuestaoService.buscar_alternativas(questao.id)
                
                for alt in alternativas:
                    letra = alt.letra.upper()
                    if letra in self.cad_alternativas:
                        self.cad_alternativas[letra].delete(0, tk.END)
                        self.cad_alternativas[letra].insert(0, alt.texto or '')
                        if alt.correta:
                            self.var_gabarito.set(letra)
            
            # Gabarito dissertativa
            if questao.tipo and questao.tipo.value == "dissertativa" and questao.gabarito_dissertativa:
                self.cad_gabarito_diss.insert("1.0", questao.gabarito_dissertativa)
            
            messagebox.showinfo(
                "Modo Edição",
                f"Editando questão #{questao.id}\n\n"
                "Faça as alterações necessárias e clique em 'Salvar Rascunho' ou 'Enviar para Revisão'."
            )
            
        except Exception as e:
            logger.error(f"Erro ao carregar questão para edição: {e}")
            messagebox.showerror("Erro", f"Erro ao carregar questão: {e}")
    
    # =========================================================================
    # ABA: MONTAR AVALIAÇÃO
    # =========================================================================
    
    def criar_aba_avaliacao(self):
        """Cria a aba de montagem de avaliação."""
        # Frame esquerdo - dados da avaliação
        frame_esq = tk.Frame(self.frame_avaliacao, bg=self.co0, width=400)
        frame_esq.pack(side="left", fill="y", padx=10, pady=10)
        frame_esq.pack_propagate(False)
        
        # Dados da avaliação
        frame_dados = tk.LabelFrame(
            frame_esq, text="Dados da Avaliação",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_dados.pack(fill="x", pady=5)
        
        tk.Label(frame_dados, text="Título:*", bg=self.co0).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.aval_titulo = ttk.Entry(frame_dados, width=30)
        self.aval_titulo.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(frame_dados, text="Componente:*", bg=self.co0).grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.aval_componente = ttk.Combobox(frame_dados, width=25, state="readonly")
        self.aval_componente['values'] = [
            "Língua Portuguesa", "Matemática", "Ciências",
            "Geografia", "História", "Arte", "Educação Física",
            "Língua Inglesa", "Ensino Religioso"
        ]
        self.aval_componente.grid(row=1, column=1, padx=5, pady=5)
        
        tk.Label(frame_dados, text="Ano:*", bg=self.co0).grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.aval_ano = ttk.Combobox(frame_dados, width=15, state="readonly")
        self.aval_ano['values'] = [
            "1º ano", "2º ano", "3º ano", "4º ano", "5º ano",
            "6º ano", "7º ano", "8º ano", "9º ano"
        ]
        self.aval_ano.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        
        tk.Label(frame_dados, text="Bimestre:", bg=self.co0).grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.aval_bimestre = ttk.Combobox(frame_dados, width=15, state="readonly")
        self.aval_bimestre['values'] = ["1º bimestre", "2º bimestre", "3º bimestre", "4º bimestre"]
        self.aval_bimestre.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        
        tk.Label(frame_dados, text="Tipo:*", bg=self.co0).grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.aval_tipo = ttk.Combobox(frame_dados, width=20, state="readonly")
        self.aval_tipo['values'] = [
            "diagnostica", "formativa", "somativa", "recuperacao", "simulado", "exercicio"
        ]
        self.aval_tipo.current(2)
        self.aval_tipo.grid(row=4, column=1, padx=5, pady=5, sticky="w")
        
        # Buscar Questões
        frame_busca_q = tk.LabelFrame(
            frame_esq, text="Buscar Questões",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_busca_q.pack(fill="x", expand=False, pady=5)  # MUDADO: fill="x" e expand=False
        
        # Filtros de busca - LINHA 1
        frame_filtros1 = tk.Frame(frame_busca_q, bg=self.co0)
        frame_filtros1.pack(fill="x", padx=5, pady=2)
        
        tk.Label(frame_filtros1, text="ID:", bg=self.co0, font=("Arial", 8)).grid(row=0, column=0, padx=2, pady=2, sticky="w")
        self.busca_aval_id = ttk.Entry(frame_filtros1, width=8)
        self.busca_aval_id.grid(row=0, column=1, padx=2, pady=2)
        
        tk.Label(frame_filtros1, text="Componente:", bg=self.co0, font=("Arial", 8)).grid(row=0, column=2, padx=2, pady=2, sticky="w")
        self.busca_aval_comp = ttk.Combobox(frame_filtros1, width=14, state="readonly", font=("Arial", 8))
        self.busca_aval_comp['values'] = ["Todos", "Língua Portuguesa", "Matemática", "Ciências", "Geografia", "História"]
        self.busca_aval_comp.current(0)
        self.busca_aval_comp.grid(row=0, column=3, padx=2, pady=2)
        
        # LINHA 2
        frame_filtros2 = tk.Frame(frame_busca_q, bg=self.co0)
        frame_filtros2.pack(fill="x", padx=5, pady=2)
        
        tk.Label(frame_filtros2, text="Ano:", bg=self.co0, font=("Arial", 8)).grid(row=0, column=0, padx=2, pady=2, sticky="w")
        self.busca_aval_ano = ttk.Combobox(frame_filtros2, width=8, state="readonly", font=("Arial", 8))
        self.busca_aval_ano['values'] = ["Todos", "6º ano", "7º ano", "8º ano", "9º ano"]
        self.busca_aval_ano.current(0)
        self.busca_aval_ano.grid(row=0, column=1, padx=2, pady=2)
        
        tk.Label(frame_filtros2, text="Tipo:", bg=self.co0, font=("Arial", 8)).grid(row=0, column=2, padx=2, pady=2, sticky="w")
        self.busca_aval_tipo = ttk.Combobox(frame_filtros2, width=12, state="readonly", font=("Arial", 8))
        self.busca_aval_tipo['values'] = ["Todos", "Dissertativa", "Múltipla Escolha"]
        self.busca_aval_tipo.current(0)
        self.busca_aval_tipo.grid(row=0, column=3, padx=2, pady=2)
        
        # LINHA 3 - Busca por texto
        frame_filtros3 = tk.Frame(frame_busca_q, bg=self.co0)
        frame_filtros3.pack(fill="x", padx=5, pady=2)
        
        tk.Label(frame_filtros3, text="Palavras-chave:", bg=self.co0, font=("Arial", 8)).pack(side="left", padx=2)
        self.busca_aval_texto = ttk.Entry(frame_filtros3, width=25, font=("Arial", 8))
        self.busca_aval_texto.pack(side="left", padx=2, fill="x", expand=True)
        
        tk.Button(
            frame_filtros3,
            text="🔍 Buscar",
            command=self.buscar_questoes_para_avaliacao,
            bg=self.co4, fg="white",
            font=("Arial", 8, "bold"),
            relief="flat"
        ).pack(side="left", padx=5)
        
        # Lista de questões encontradas - COM PREVIEW
        frame_lista_q = tk.Frame(frame_busca_q, bg=self.co0)
        frame_lista_q.pack(fill="both", expand=False, padx=5, pady=5)  # MUDADO: expand=False
        
        scroll_q = tk.Scrollbar(frame_lista_q, orient="vertical")
        scroll_q.pack(side="right", fill="y")
        
        # Treeview com colunas: ID, Tipo, Preview do Enunciado
        self.tree_questoes_busca = ttk.Treeview(
            frame_lista_q,
            columns=("id", "tipo", "enunciado"),
            show="headings",
            yscrollcommand=scroll_q.set,
            height=5  # REDUZIDO DE 8 PARA 5
        )
        
        self.tree_questoes_busca.heading("id", text="ID")
        self.tree_questoes_busca.heading("tipo", text="Tipo")
        self.tree_questoes_busca.heading("enunciado", text="Enunciado (preview)")
        
        self.tree_questoes_busca.column("id", width=40, anchor="center")
        self.tree_questoes_busca.column("tipo", width=80, anchor="w")
        self.tree_questoes_busca.column("enunciado", width=250, anchor="w")
        
        self.tree_questoes_busca.pack(fill="both", expand=False)  # MUDADO: expand=False
        scroll_q.config(command=self.tree_questoes_busca.yview)
        
        # Duplo clique para adicionar OU ver detalhes
        self.tree_questoes_busca.bind("<Double-Button-1>", lambda e: self.adicionar_questao_da_busca())
        self.tree_questoes_busca.bind("<Return>", lambda e: self.adicionar_questao_da_busca())
        
        # Frame de botões
        frame_btns_busca = tk.Frame(frame_busca_q, bg=self.co0)
        frame_btns_busca.pack(fill="x", padx=5, pady=5)
        
        tk.Button(
            frame_btns_busca,
            text="➕ Adicionar Selecionada",
            command=self.adicionar_questao_da_busca,
            bg=self.co2, fg="white",
            font=("Arial", 8, "bold"),
            relief="flat"
        ).pack(side="left", padx=2)
        
        tk.Button(
            frame_btns_busca,
            text="👁️ Ver Detalhes",
            command=self.ver_detalhes_questao_busca,
            bg=self.co4, fg="white",
            font=("Arial", 8, "bold"),
            relief="flat"
        ).pack(side="left", padx=2)
        
        # Textos Base da Avaliação
        frame_textos = tk.LabelFrame(
            frame_esq, text="Textos Base (opcional)",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_textos.pack(fill="both", expand=True, pady=5)  # ESTE SIM USA expand=True
        
        # Lista de textos base vinculados
        self.textos_base_avaliacao = []  # Lista de dicts: {id, ordem, questoes_vinculadas}
        
        # Listbox para textos base
        scroll_textos = tk.Scrollbar(frame_textos, orient="vertical")
        scroll_textos.pack(side="right", fill="y")
        
        self.listbox_textos = tk.Listbox(
            frame_textos,
            yscrollcommand=scroll_textos.set,
            height=5,  # REDUZIDO DE 8 PARA 5
            font=("Arial", 9)
        )
        self.listbox_textos.pack(fill="both", expand=True, padx=5, pady=(5, 2))  # REDUZIDO pady inferior
        scroll_textos.config(command=self.listbox_textos.yview)
        
        # Botões para gerenciar textos base
        frame_btn_textos = tk.Frame(frame_textos, bg=self.co0)
        frame_btn_textos.pack(fill="x", padx=5, pady=(2, 5))  # REDUZIDO pady superior
        
        tk.Button(
            frame_btn_textos,
            text="➕ Adicionar",
            command=self.adicionar_texto_base_avaliacao,
            bg=self.co2, fg="white",
            font=("Arial", 8, "bold"),
            relief="flat"
        ).pack(side="left", padx=2)
        
        tk.Button(
            frame_btn_textos,
            text="⬆️",
            command=lambda: self.mover_texto_base(-1),
            bg=self.co9, fg="white",
            font=("Arial", 8, "bold"),
            relief="flat",
            width=3
        ).pack(side="left", padx=2)
        
        tk.Button(
            frame_btn_textos,
            text="⬇️",
            command=lambda: self.mover_texto_base(1),
            bg=self.co9, fg="white",
            font=("Arial", 8, "bold"),
            relief="flat",
            width=3
        ).pack(side="left", padx=2)
        
        tk.Button(
            frame_btn_textos,
            text="🗑️",
            command=self.remover_texto_base_avaliacao,
            bg=self.co8, fg="white",
            font=("Arial", 8, "bold"),
            relief="flat",
            width=3
        ).pack(side="left", padx=2)
        
        # Frame direito - questões selecionadas
        frame_dir = tk.Frame(self.frame_avaliacao, bg=self.co0)
        frame_dir.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        frame_questoes = tk.LabelFrame(
            frame_dir, text="Questões da Avaliação",
            bg=self.co0, font=("Arial", 10, "bold")
        )
        frame_questoes.pack(fill="both", expand=True)
        
        # Lista de questões na avaliação
        self.questoes_avaliacao = []  # Lista de IDs
        
        colunas = ("ordem", "id", "habilidade", "tipo", "pontos")
        self.tree_avaliacao = ttk.Treeview(
            frame_questoes, columns=colunas, show="headings", height=12
        )
        
        self.tree_avaliacao.heading("ordem", text="#")
        self.tree_avaliacao.heading("id", text="ID")
        self.tree_avaliacao.heading("habilidade", text="Habilidade")
        self.tree_avaliacao.heading("tipo", text="Tipo")
        self.tree_avaliacao.heading("pontos", text="Pontos")
        
        self.tree_avaliacao.column("ordem", width=40)
        self.tree_avaliacao.column("id", width=60)
        self.tree_avaliacao.column("habilidade", width=100)
        self.tree_avaliacao.column("tipo", width=120)
        self.tree_avaliacao.column("pontos", width=70)
        
        self.tree_avaliacao.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Botões de ação
        frame_btns = tk.Frame(frame_questoes, bg=self.co0)
        frame_btns.pack(fill="x", padx=5, pady=5)
        
        tk.Button(
            frame_btns, text="🗑️ Remover", command=self.remover_questao_avaliacao,
            bg=self.co8, fg="white"
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_btns, text="⬆️ Subir", command=lambda: self.mover_questao(-1),
            bg=self.co9, fg="white"
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_btns, text="⬇️ Descer", command=lambda: self.mover_questao(1),
            bg=self.co9, fg="white"
        ).pack(side="left", padx=5)
        
        # Total de pontos
        frame_total = tk.Frame(frame_dir, bg=self.co0)
        frame_total.pack(fill="x", pady=5)
        
        tk.Label(frame_total, text="Total de Questões:", bg=self.co0).pack(side="left", padx=5)
        self.lbl_total_questoes = tk.Label(frame_total, text="0", bg=self.co0, font=("Arial", 10, "bold"))
        self.lbl_total_questoes.pack(side="left")
        
        tk.Label(frame_total, text="  |  Total de Pontos:", bg=self.co0).pack(side="left", padx=5)
        self.lbl_total_pontos = tk.Label(frame_total, text="0.0", bg=self.co0, font=("Arial", 10, "bold"))
        self.lbl_total_pontos.pack(side="left")
        
        # Botões finais
        frame_final = tk.Frame(frame_dir, bg=self.co0)
        frame_final.pack(fill="x", pady=10)
        
        tk.Button(
            frame_final, text="💾 Salvar Avaliação",
            command=self.salvar_avaliacao,
            bg=self.co2, fg="white", font=("Arial", 10, "bold")
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_final, text="🖨️ Gerar PDF",
            command=self.gerar_pdf_avaliacao,
            bg=self.co4, fg="white", font=("Arial", 10, "bold")
        ).pack(side="left", padx=5)
    
    def adicionar_questao_avaliacao(self, questao_id: int):
        """Adiciona uma questão à avaliação."""
        if questao_id in self.questoes_avaliacao:
            messagebox.showinfo("Aviso", "Esta questão já está na avaliação.")
            return
        
        try:
            from banco_questoes.services import QuestaoService
            questao = QuestaoService.buscar_por_id(questao_id)
            
            if not questao:
                messagebox.showerror("Erro", "Questão não encontrada.")
                return
            
            self.questoes_avaliacao.append(questao_id)
            
            ordem = len(self.questoes_avaliacao)
            self.tree_avaliacao.insert("", "end", values=(
                ordem,
                questao_id,
                questao.habilidade_bncc_codigo or '',
                questao.tipo.value if questao.tipo else '',
                "1.0"  # Pontuação padrão
            ))
            
            self.atualizar_totais_avaliacao()
            messagebox.showinfo("Sucesso", f"Questão #{questao_id} adicionada!")
            
        except Exception as e:
            logger.error(f"Erro ao adicionar questão: {e}")
            messagebox.showerror("Erro", f"Erro ao adicionar questão: {e}")
    
    def remover_questao_avaliacao(self):
        """Remove questão selecionada da avaliação."""
        selecao = self.tree_avaliacao.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para remover.")
            return
        
        item = self.tree_avaliacao.item(selecao[0])
        questao_id = item['values'][1]
        
        self.questoes_avaliacao.remove(questao_id)
        self.tree_avaliacao.delete(selecao[0])
        
        # Renumerar
        for i, item_id in enumerate(self.tree_avaliacao.get_children(), 1):
            valores = list(self.tree_avaliacao.item(item_id)['values'])
            valores[0] = str(i)
            self.tree_avaliacao.item(item_id, values=valores)
        
        self.atualizar_totais_avaliacao()
    
    def mover_questao(self, direcao: int):
        """Move questão para cima (-1) ou para baixo (+1)."""
        selecao = self.tree_avaliacao.selection()
        if not selecao:
            return
        
        item_id = selecao[0]
        index = self.tree_avaliacao.index(item_id)
        novo_index = index + direcao
        
        if novo_index < 0 or novo_index >= len(self.tree_avaliacao.get_children()):
            return
        
        self.tree_avaliacao.move(item_id, "", novo_index)
        
        # Renumerar
        for i, item in enumerate(self.tree_avaliacao.get_children(), 1):
            valores = list(self.tree_avaliacao.item(item)['values'])
            valores[0] = str(i)
            self.tree_avaliacao.item(item, values=valores)
    
    def adicionar_texto_base_avaliacao(self):
        """Abre dialog para adicionar texto base à avaliação."""
        dialog = tk.Toplevel(self.janela)
        dialog.title("Adicionar Texto Base")
        dialog.geometry("700x650")  # AUMENTADO DE 600x450
        dialog.configure(bg=self.co0)
        dialog.grab_set()
        
        # Frame principal com scroll
        main_frame = tk.Frame(dialog, bg=self.co0)
        main_frame.pack(fill="both", expand=True)
        
        # Canvas e scrollbar
        canvas = tk.Canvas(main_frame, bg=self.co0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        
        container = tk.Frame(canvas, bg=self.co0)
        
        container.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=container, anchor="nw", width=680)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        # Habilitar scroll com mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        tk.Label(
            container,
            text="Selecione um Texto Base",
            font=("Arial", 12, "bold"),
            bg=self.co0, fg=self.co1
        ).pack(anchor="w", pady=(10, 10), padx=10)
        
        # Treeview para selecionar texto
        frame_tree = tk.Frame(container, bg=self.co0)
        frame_tree.pack(fill="both", expand=False, pady=(0, 10), padx=10)  # expand=False para controlar altura
        
        scroll_y = tk.Scrollbar(frame_tree, orient="vertical")
        scroll_y.pack(side="right", fill="y")
        
        tree_textos = ttk.Treeview(
            frame_tree,
            columns=("id", "titulo", "tipo"),
            show="headings",
            yscrollcommand=scroll_y.set,
            height=8  # ALTURA FIXA
        )
        
        tree_textos.heading("id", text="ID")
        tree_textos.heading("titulo", text="Título")
        tree_textos.heading("tipo", text="Tipo")
        
        tree_textos.column("id", width=50, anchor="center")
        tree_textos.column("titulo", width=350, anchor="w")
        tree_textos.column("tipo", width=100, anchor="center")
        
        tree_textos.pack(fill="both", expand=True)
        scroll_y.config(command=tree_textos.yview)
        
        # Carregar textos base disponíveis
        try:
            textos = TextoBaseService.listar(escola_id=config.ESCOLA_ID)
            for texto in textos:
                tipo_display = "📝 Texto" if texto.tipo == TipoTextoBase.TEXTO else "🖼️ Imagem"
                tree_textos.insert("", "end", values=(texto.id, texto.titulo, tipo_display))
        except Exception as e:
            logger.error(f"Erro ao carregar textos base: {e}")
        
        # Campo para números de questões vinculadas - CRIAR ANTES DO PREVIEW
        frame_questoes = tk.Frame(container, bg=self.co0)
        frame_questoes.pack(fill="x", pady=(0, 10), padx=10)
        
        tk.Label(
            frame_questoes,
            text="Questões vinculadas a este texto (IDs separados por vírgula):",
            font=("Arial", 9),
            bg=self.co0, fg=self.co7
        ).pack(anchor="w")
        
        entry_questoes = tk.Entry(frame_questoes, font=("Arial", 10), relief="solid", bd=1)
        entry_questoes.pack(fill="x", pady=(5, 0))
        entry_questoes.insert(0, "")
        
        # Preview do texto selecionado + sugestão de questões
        frame_preview = tk.LabelFrame(container, text="Preview e Sugestões", bg=self.co0, font=("Arial", 9, "bold"))
        frame_preview.pack(fill="x", expand=False, pady=(0, 10), padx=10)  # NÃO expand para controlar tamanho
        
        # Texto do conteúdo
        preview_texto = tk.Text(frame_preview, height=5, wrap="word", font=("Arial", 9), state="disabled")  # ALTURA REDUZIDA
        preview_texto.pack(fill="both", expand=False, padx=5, pady=5)
        
        # Label de sugestões
        lbl_sugestoes = tk.Label(
            frame_preview,
            text="💡 Selecione um texto para ver sugestões de questões",
            bg=self.co0, fg=self.co9, font=("Arial", 8, "italic"),
            wraplength=500, justify="left"
        )
        lbl_sugestoes.pack(fill="x", padx=5, pady=5)
        
        def atualizar_preview(event=None):
            """Atualiza preview do texto e sugere questões."""
            selecionado = tree_textos.selection()
            if not selecionado:
                preview_texto.config(state="normal")
                preview_texto.delete("1.0", tk.END)
                preview_texto.config(state="disabled")
                lbl_sugestoes.config(text="💡 Selecione um texto para ver sugestões de questões")
                return
            
            item = tree_textos.item(selecionado[0])
            texto_id = item["values"][0]
            
            try:
                texto_obj = TextoBaseService.buscar_por_id(texto_id)
                if texto_obj:
                    # Mostrar preview
                    preview_texto.config(state="normal")
                    preview_texto.delete("1.0", tk.END)
                    if texto_obj.tipo == TipoTextoBase.TEXTO:
                        conteudo_preview = texto_obj.conteudo[:300] if texto_obj.conteudo else ""
                        if len(texto_obj.conteudo or "") > 300:
                            conteudo_preview += "..."
                        preview_texto.insert("1.0", conteudo_preview)
                    else:
                        preview_texto.insert("1.0", f"[IMAGEM] {texto_obj.caminho_imagem or 'Sem caminho'}")
                    preview_texto.config(state="disabled")
                    
                    # Sugerir questões - buscar questões já na avaliação
                    questoes_aval = [int(self.tree_avaliacao.item(item)['values'][1]) for item in self.tree_avaliacao.get_children()]
                    if questoes_aval:
                        sugestao_ids = ", ".join(map(str, questoes_aval))
                        lbl_sugestoes.config(
                            text=f"💡 Sugestão: Você já adicionou as questões {sugestao_ids} à avaliação. "
                                 f"Vincule as que se relacionam a este texto.",
                            fg=self.co2
                        )
                        entry_questoes.delete(0, tk.END)
                        entry_questoes.insert(0, sugestao_ids)
                    else:
                        lbl_sugestoes.config(
                            text="💡 Dica: Primeiro adicione questões à avaliação (seção 'Buscar Questões'), "
                                 "depois vincule-as aos textos base.",
                            fg=self.co9
                        )
            except Exception as e:
                logger.error(f"Erro ao carregar preview: {e}")
        
        tree_textos.bind("<<TreeviewSelect>>", atualizar_preview)
        
        # Botões - SEMPRE VISÍVEIS NO FINAL
        frame_btns = tk.Frame(container, bg=self.co0)
        frame_btns.pack(fill="x", pady=(15, 10), padx=10)
        
        def confirmar():
            selecionado = tree_textos.selection()
            if not selecionado:
                messagebox.showwarning("Atenção", "Selecione um texto base.")
                return
            
            item = tree_textos.item(selecionado[0])
            texto_id = item["values"][0]
            titulo = item["values"][1]
            
            # Parse questões vinculadas
            questoes_str = entry_questoes.get().strip()
            questoes_vinculadas = []
            if questoes_str:
                try:
                    questoes_vinculadas = [int(q.strip()) for q in questoes_str.split(",") if q.strip()]
                except ValueError:
                    messagebox.showwarning("Atenção", "Formato inválido. Use números separados por vírgula (ex: 1,2,3)")
                    return
            
            # Verificar se já existe
            if any(t["id"] == texto_id for t in self.textos_base_avaliacao):
                messagebox.showinfo("Aviso", "Este texto base já foi adicionado.")
                return
            
            # Adicionar à lista
            ordem = len(self.textos_base_avaliacao) + 1
            self.textos_base_avaliacao.append({
                "id": texto_id,
                "titulo": titulo,
                "ordem": ordem,
                "questoes_vinculadas": questoes_vinculadas
            })
            
            # Atualizar listbox
            self.atualizar_listbox_textos()
            
            messagebox.showinfo("Sucesso", f"Texto base '{titulo}' adicionado!")
            dialog.destroy()
        
        tk.Button(
            frame_btns,
            text="✓ Adicionar",
            command=confirmar,
            bg=self.co2, fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20, pady=8
        ).pack(side="left", padx=(0, 10))
        
        def fechar():
            canvas.unbind_all("<MouseWheel>")
            dialog.destroy()
        
        tk.Button(
            frame_btns,
            text="✗ Cancelar",
            command=fechar,
            bg=self.co9, fg="white",
            font=("Arial", 10, "bold"),
            relief="flat",
            padx=20, pady=8
        ).pack(side="left")
        
        dialog.protocol("WM_DELETE_WINDOW", fechar)
    
    def atualizar_listbox_textos(self):
        """Atualiza a listbox de textos base da avaliação."""
        self.listbox_textos.delete(0, tk.END)
        for texto in self.textos_base_avaliacao:
            questoes = ",".join(str(q) for q in texto["questoes_vinculadas"]) if texto["questoes_vinculadas"] else "-"
            display = f"{texto['ordem']}. {texto['titulo']} (Q: {questoes})"
            self.listbox_textos.insert(tk.END, display)
    
    def remover_texto_base_avaliacao(self):
        """Remove texto base selecionado da avaliação."""
        selecao = self.listbox_textos.curselection()
        if not selecao:
            messagebox.showwarning("Atenção", "Selecione um texto base para remover.")
            return
        
        index = selecao[0]
        texto = self.textos_base_avaliacao[index]
        
        confirmacao = messagebox.askyesno(
            "Confirmar",
            f"Remover texto base '{texto['titulo']}'?"
        )
        
        if confirmacao:
            self.textos_base_avaliacao.pop(index)
            # Renumerar
            for i, t in enumerate(self.textos_base_avaliacao, 1):
                t["ordem"] = i
            self.atualizar_listbox_textos()
    
    def mover_texto_base(self, direcao: int):
        """Move texto base para cima (-1) ou para baixo (+1)."""
        selecao = self.listbox_textos.curselection()
        if not selecao:
            return
        
        index = selecao[0]
        novo_index = index + direcao
        
        if novo_index < 0 or novo_index >= len(self.textos_base_avaliacao):
            return
        
        # Trocar posições
        self.textos_base_avaliacao[index], self.textos_base_avaliacao[novo_index] = \
            self.textos_base_avaliacao[novo_index], self.textos_base_avaliacao[index]
        
        # Renumerar
        for i, t in enumerate(self.textos_base_avaliacao, 1):
            t["ordem"] = i
        
        self.atualizar_listbox_textos()
        self.listbox_textos.selection_set(novo_index)
    
    def buscar_questoes_para_avaliacao(self):
        """Busca questões para adicionar na avaliação com FILTROS AVANÇADOS."""
        from banco_questoes.services import QuestaoService
        from banco_questoes.models import FiltroQuestoes, StatusQuestao, ComponenteCurricular, AnoEscolar, TipoQuestao
        
        # Limpar treeview
        for item in self.tree_questoes_busca.get_children():
            self.tree_questoes_busca.delete(item)
        
        try:
            # Montar filtros
            filtros = FiltroQuestoes(
                escola_id=config.ESCOLA_ID,
                status=StatusQuestao.APROVADA
            )
            
            # Filtro por ID específico
            id_busca = self.busca_aval_id.get().strip()
            if id_busca:
                try:
                    questao_id = int(id_busca)
                    questao = QuestaoService.buscar_por_id(questao_id)
                    if questao:
                        tipo_str = {"multipla_escolha": "Múlt. Escolha", "dissertativa": "Dissertativa", "verdadeiro_falso": "V/F"}.get(
                            questao.tipo.value if questao.tipo else "", "?"
                        )
                        enunciado_preview = (questao.enunciado[:80] + "...") if len(questao.enunciado or "") > 80 else (questao.enunciado or "")
                        
                        self.tree_questoes_busca.insert("", "end", values=(
                            questao.id,
                            tipo_str,
                            enunciado_preview
                        ))
                    else:
                        messagebox.showinfo("Não encontrado", f"Questão ID {questao_id} não encontrada.")
                except ValueError:
                    messagebox.showwarning("Erro", "ID deve ser um número.")
                return
            
            # Filtro por componente
            componente = self.busca_aval_comp.get()
            if componente != "Todos":
                try:
                    filtros.componente_curricular = ComponenteCurricular(componente)
                except ValueError:
                    pass
            
            # Filtro por ano
            ano = self.busca_aval_ano.get()
            if ano != "Todos":
                try:
                    filtros.ano_escolar = AnoEscolar(ano)
                except ValueError:
                    pass
            
            # Filtro por tipo
            tipo = self.busca_aval_tipo.get()
            if tipo != "Todos":
                tipo_map = {"Dissertativa": "dissertativa", "Múltipla Escolha": "multipla_escolha"}
                tipo_valor = tipo_map.get(tipo)
                if tipo_valor:
                    try:
                        filtros.tipo = TipoQuestao(tipo_valor)
                    except ValueError:
                        pass
            
            # Filtro por texto
            texto = self.busca_aval_texto.get().strip()
            if texto:
                filtros.texto_busca = texto
            
            # Buscar questões
            questoes, total = QuestaoService.buscar(filtros, limite=100)
            
            if not questoes:
                messagebox.showinfo("Aviso", "Nenhuma questão encontrada com os filtros aplicados.")
                return
            
            # Popular treeview
            for questao in questoes:
                tipo_str = {"multipla_escolha": "Múlt. Escolha", "dissertativa": "Dissertativa", "verdadeiro_falso": "V/F"}.get(
                    questao.tipo.value if questao.tipo else "", "?"
                )
                enunciado_preview = (questao.enunciado[:80] + "...") if len(questao.enunciado or "") > 80 else (questao.enunciado or "")
                
                self.tree_questoes_busca.insert("", "end", values=(
                    questao.id,
                    tipo_str,
                    enunciado_preview
                ))
            
            # Informar total encontrado
            if total > len(questoes):
                messagebox.showinfo("Resultados", f"Encontradas {total} questões. Exibindo primeiras {len(questoes)}.")
            
        except Exception as e:
            logger.error(f"Erro ao buscar questões: {e}")
            messagebox.showerror("Erro", f"Erro ao buscar questões: {e}")
    
    def ver_detalhes_questao_busca(self):
        """Abre janela de detalhes da questão selecionada na busca."""
        selecao = self.tree_questoes_busca.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para ver detalhes.")
            return
        
        item = self.tree_questoes_busca.item(selecao[0])
        questao_id = item["values"][0]
        self.abrir_detalhes_questao(questao_id)
    
    def adicionar_questao_da_busca(self):
        """Adiciona a questão selecionada na busca para a avaliação."""
        selecao = self.tree_questoes_busca.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para adicionar.")
            return
        
        item = self.tree_questoes_busca.item(selecao[0])
        questao_id = item["values"][0]
        self.adicionar_questao_avaliacao(questao_id)
    
    def atualizar_totais_avaliacao(self):
        """Atualiza totais de questões e pontos."""
        total_questoes = len(self.questoes_avaliacao)
        total_pontos = total_questoes * 1.0  # Pontuação fixa por enquanto
        
        self.lbl_total_questoes.config(text=str(total_questoes))
        self.lbl_total_pontos.config(text=f"{total_pontos:.1f}")
    
    def salvar_avaliacao(self):
        """Salva a avaliação no banco de dados."""
        from banco_questoes.services import AvaliacaoService
        from banco_questoes.models import Avaliacao, ComponenteCurricular, AnoEscolar, TipoAvaliacao, StatusAvaliacao
        
        if not self.aval_titulo.get().strip():
            messagebox.showwarning("Aviso", "Informe o título da avaliação.")
            return
        
        if not self.aval_componente.get():
            messagebox.showwarning("Aviso", "Selecione o componente curricular.")
            return
        
        if not self.aval_ano.get():
            messagebox.showwarning("Aviso", "Selecione o ano escolar.")
            return
        
        if not self.questoes_avaliacao:
            messagebox.showwarning("Aviso", "Adicione pelo menos uma questão.")
            return
        
        if not self.funcionario_id:
            messagebox.showwarning("Aviso", "É necessário estar logado.")
            return
        
        try:
            # Bimestre como string (ex: "1" ou "1º bimestre")
            bimestre = self.aval_bimestre.get() if self.aval_bimestre.get() else None
            
            avaliacao = Avaliacao(
                titulo=self.aval_titulo.get().strip(),
                componente_curricular=ComponenteCurricular(self.aval_componente.get()),
                ano_escolar=AnoEscolar(self.aval_ano.get()),
                bimestre=bimestre,
                tipo=TipoAvaliacao(self.aval_tipo.get()),
                professor_id=self.funcionario_id,
                escola_id=config.ESCOLA_ID,
                status=StatusAvaliacao.RASCUNHO
            )
            
            avaliacao_id = AvaliacaoService.criar(avaliacao)
            
            if avaliacao_id:
                # Adicionar questões à avaliação
                for ordem, questao_id in enumerate(self.questoes_avaliacao, 1):
                    AvaliacaoService.adicionar_questao(avaliacao_id, questao_id, ordem=ordem, pontuacao=1.0)
                
                # Adicionar textos base à avaliação
                for texto in self.textos_base_avaliacao:
                    try:
                        TextoBaseService.vincular_texto_avaliacao(
                            avaliacao_id=avaliacao_id,
                            texto_base_id=texto["id"],
                            ordem=texto["ordem"]
                        )
                    except Exception as e:
                        logger.error(f"Erro ao vincular texto base {texto['id']}: {e}")
                
                msg_textos = f"\n{len(self.textos_base_avaliacao)} textos base vinculados." if self.textos_base_avaliacao else ""
                messagebox.showinfo(
                    "Sucesso",
                    f"Avaliação #{avaliacao_id} salva com sucesso!\n"
                    f"{len(self.questoes_avaliacao)} questões adicionadas.{msg_textos}"
                )
            else:
                messagebox.showerror("Erro", "Não foi possível salvar a avaliação.")
            
        except Exception as e:
            logger.error(f"Erro ao salvar avaliação: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar avaliação: {e}")
    
    def gerar_pdf_avaliacao(self):
        """Gera PDF da avaliação."""
        # Validar campos obrigatórios
        if not self.aval_titulo.get().strip():
            messagebox.showwarning("Aviso", "Informe o título da avaliação.")
            return
        
        if not self.aval_componente.get():
            messagebox.showwarning("Aviso", "Selecione o componente curricular.")
            return
        
        if not self.aval_ano.get():
            messagebox.showwarning("Aviso", "Selecione o ano escolar.")
            return
        
        if not self.questoes_avaliacao:
            messagebox.showwarning("Aviso", "Adicione pelo menos uma questão à avaliação.")
            return
        
        try:
            from tkinter import filedialog
            import os
            
            # Solicitar local para salvar
            nome_sugerido = f"Avaliacao_{self.aval_componente.get().replace(' ', '_')}_{self.aval_ano.get().replace(' ', '_')}.pdf"
            
            caminho_pdf = filedialog.asksaveasfilename(
                title="Salvar Avaliação em PDF",
                defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")],
                initialfile=nome_sugerido,
                parent=self.janela
            )
            
            if not caminho_pdf:
                return
            
            # Gerar o PDF
            self._criar_pdf_avaliacao(caminho_pdf)
            
            # Perguntar se quer abrir
            resposta = messagebox.askyesno(
                "PDF Gerado",
                f"PDF gerado com sucesso!\n\n{caminho_pdf}\n\nDeseja abrir o arquivo?",
                icon='info'
            )
            
            if resposta:
                os.startfile(caminho_pdf)
            
        except Exception as e:
            logger.error(f"Erro ao gerar PDF: {e}")
            messagebox.showerror("Erro", f"Erro ao gerar PDF:\n{e}")
    
    def _criar_pdf_avaliacao(self, caminho_pdf: str):
        """Cria o arquivo PDF da avaliação."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import Paragraph, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
        from datetime import datetime
        
        # Criar canvas
        c = canvas.Canvas(caminho_pdf, pagesize=A4)
        width, height = A4
        
        # Margens
        margin_left = 2 * cm
        margin_right = 2 * cm
        margin_top = 2 * cm
        margin_bottom = 2 * cm
        
        y_position = height - margin_top
        
        # ====================================================================
        # CABEÇALHO
        # ====================================================================
        
        # Logo da escola (centralizado, preservando proporção)
        logo_path = r"C:\gestao\imagens\logopaco.png"
        header_block_height = 0
        if os.path.exists(logo_path):
            try:
                # Obter tamanho da imagem e calcular escala para manter proporção
                from PIL import Image
                img_pil = Image.open(logo_path)
                img_w, img_h = img_pil.size

                # Limites máximos para exibição do logo
                max_w = 6 * cm
                max_h = 3 * cm

                scale = min(max_w / img_w, max_h / img_h, 1.0)
                disp_w = img_w * scale
                disp_h = img_h * scale

                x_img = (width - disp_w) / 2
                y_img = y_position - disp_h
                c.drawImage(logo_path, x_img, y_img, width=disp_w, height=disp_h, preserveAspectRatio=True, mask='auto')

                header_block_height = disp_h
                y_position = y_img - 0.2 * cm
            except Exception as e:
                logger.warning(f"Erro ao carregar logo: {e}")
                header_block_height = 0
        else:
            header_block_height = 0

        # Texto do cabeçalho (centralizado e depois justificado no corpo)
        estilos = getSampleStyleSheet()
        estilo_cab = ParagraphStyle(
            'cabecalho', parent=estilos['Normal'], alignment=TA_JUSTIFY, fontName='Helvetica-Bold', fontSize=12)
        estilo_info = ParagraphStyle(
            'info', parent=estilos['Normal'], alignment=TA_JUSTIFY, fontName='Helvetica', fontSize=11)

        # Construir conteúdo do cabeçalho
        # Espaçamento uniforme entre as três linhas do cabeçalho
        gap = 0.5 * cm

        cab_html = 'E.M. PROFª. NADIR NASCIMENTO MORAES'
        p = Paragraph(cab_html, estilo_cab)
        w_p, h_p = p.wrap(width - margin_left - margin_right, y_position)
        p.drawOn(c, margin_left, y_position - h_p)
        y_position -= h_p + gap

        # Adicionar '- MA' conforme documentação
        info_text = 'PAÇO DO LUMIAR - MA, ______de___________________ de 2025.'
        p2 = Paragraph(info_text, estilo_info)
        w2, h2 = p2.wrap(width - margin_left - margin_right, y_position)
        p2.drawOn(c, margin_left, y_position - h2)
        y_position -= h2 + gap

        # Linha com ESTUDANTE (esquerda) e TURMA (direita) usando Table para alinhar e manter espaçamento
        estudante_text = 'ESTUDANTE:____________________________________________'
        turma_text = 'TURMA:________'
        content_width = width - margin_left - margin_right
        col_w1 = content_width * 0.75
        col_w2 = content_width - col_w1
        table = Table([[estudante_text, turma_text]], colWidths=[col_w1, col_w2])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
        ]))

        w_tab, h_tab = table.wrap(content_width, y_position)
        table.drawOn(c, margin_left, y_position - h_tab)
        y_position -= h_tab + gap

        # Linha separadora
        c.line(margin_left, y_position, width - margin_right, y_position)
        y_position -= 0.8 * cm
        
        # ====================================================================
        # TÍTULO DA ATIVIDADE
        # ====================================================================
        
        componente = self.aval_componente.get()
        ano = self.aval_ano.get()
        bimestre = self.aval_bimestre.get() if self.aval_bimestre.get() else "____"
        
        titulo_atividade = f"Atividade Avaliativa de {componente} do {ano} - {bimestre}"
        
        c.setFont("Helvetica-Bold", 14)
        # Centralizar título
        text_width = c.stringWidth(titulo_atividade, "Helvetica-Bold", 14)
        x_centered = (width - text_width) / 2
        c.drawString(x_centered, y_position, titulo_atividade)
        
        # Remover subtítulo "Lista de Questões" — o formato das questões terá numeração direta
        y_position -= 1.2 * cm
        
        # ====================================================================
        # TEXTOS BASE (se houver)
        # ====================================================================
        
        if self.textos_base_avaliacao:
            # Coletar questões vinculadas para o enunciado
            todas_questoes = set()
            for texto in self.textos_base_avaliacao:
                todas_questoes.update(texto.get("questoes_vinculadas", []))
            
            questoes_vinculadas_str = ", ".join(str(q) for q in sorted(todas_questoes))
            
            # Verificar tipos dos textos para gerar enunciado apropriado
            tipos = set()
            textos_obj = []
            for texto in self.textos_base_avaliacao:
                try:
                    texto_obj = TextoBaseService.buscar_por_id(texto["id"])
                    if texto_obj:
                        textos_obj.append(texto_obj)
                        tipos.add(texto_obj.tipo)
                except Exception as e:
                    logger.error(f"Erro ao carregar texto base {texto['id']}: {e}")
            
            # Gerar enunciado introdutório
            if tipos == {TipoTextoBase.TEXTO}:
                enunciado_intro = f"Com base nos textos, responda as questões {questoes_vinculadas_str}."
            elif tipos == {TipoTextoBase.IMAGEM}:
                enunciado_intro = f"Analise as imagens e responda as questões {questoes_vinculadas_str}."
            else:
                enunciado_intro = f"Com base nos textos e imagens, responda as questões {questoes_vinculadas_str}."
            
            # Renderizar enunciado
            estilo_enunciado = ParagraphStyle(
                'enunciado_intro',
                parent=estilos['Normal'],
                alignment=TA_JUSTIFY,
                fontName='Helvetica-Bold',
                fontSize=11,
                leading=13
            )
            p_enunciado = Paragraph(enunciado_intro, estilo_enunciado)
            w_en, h_en = p_enunciado.wrap(width - margin_left - margin_right, y_position)
            
            if h_en > y_position - margin_bottom:
                c.showPage()
                y_position = height - margin_top
            
            p_enunciado.drawOn(c, margin_left, y_position - h_en)
            y_position -= h_en + 0.6 * cm
            
            # Renderizar textos base
            if len(textos_obj) == 1:
                # Um único texto - usar largura completa
                texto = textos_obj[0]
                
                if texto.tipo == TipoTextoBase.TEXTO:
                    # Título do texto
                    estilo_titulo = ParagraphStyle('titulo_texto', parent=estilos['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=11)
                    p_tit = Paragraph(f"<b>{texto.titulo}</b>", estilo_titulo)
                    w_tit, h_tit = p_tit.wrap(width - margin_left - margin_right, y_position)
                    
                    if h_tit > y_position - margin_bottom:
                        c.showPage()
                        y_position = height - margin_top
                    
                    p_tit.drawOn(c, margin_left, y_position - h_tit)
                    y_position -= h_tit + 0.3 * cm
                    
                    # Conteúdo do texto
                    estilo_texto = ParagraphStyle('texto_base', parent=estilos['Normal'], alignment=TA_JUSTIFY, fontName='Helvetica', fontSize=10, leading=12)
                    p_txt = Paragraph(texto.conteudo or "", estilo_texto)
                    w_txt, h_txt = p_txt.wrap(width - margin_left - margin_right, y_position)
                    
                    if h_txt > y_position - margin_bottom:
                        c.showPage()
                        y_position = height - margin_top
                    
                    p_txt.drawOn(c, margin_left, y_position - h_txt)
                    y_position -= h_txt + 0.8 * cm
                
                else:  # Imagem
                    if texto.caminho_imagem and os.path.exists(texto.caminho_imagem):
                        # Título
                        estilo_titulo = ParagraphStyle('titulo_img', parent=estilos['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=11)
                        p_tit = Paragraph(f"<b>{texto.titulo}</b>", estilo_titulo)
                        w_tit, h_tit = p_tit.wrap(width - margin_left - margin_right, y_position)
                        
                        if h_tit > y_position - margin_bottom:
                            c.showPage()
                            y_position = height - margin_top
                        
                        p_tit.drawOn(c, margin_left, y_position - h_tit)
                        y_position -= h_tit + 0.3 * cm
                        
                        # Imagem centralizada
                        from PIL import Image
                        img_pil = Image.open(texto.caminho_imagem)
                        img_w, img_h = img_pil.size
                        
                        max_w = width - margin_left - margin_right
                        max_h = 8 * cm
                        scale = min(max_w / img_w, max_h / img_h, 1.0)
                        disp_w = img_w * scale
                        disp_h = img_h * scale
                        
                        if disp_h > y_position - margin_bottom:
                            c.showPage()
                            y_position = height - margin_top
                        
                        x_img = margin_left + (max_w - disp_w) / 2
                        c.drawImage(texto.caminho_imagem, x_img, y_position - disp_h, width=disp_w, height=disp_h, preserveAspectRatio=True)
                        y_position -= disp_h + 0.8 * cm
            
            elif len(textos_obj) == 2:
                # Dois textos - layout lado a lado usando Table
                col_width = (width - margin_left - margin_right - 0.5 * cm) / 2
                
                estilo_titulo_texto = ParagraphStyle('tit_texto', parent=estilos['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=11)
                estilo_texto = ParagraphStyle('texto', parent=estilos['Normal'], alignment=TA_JUSTIFY, fontName='Helvetica', fontSize=10, leading=12)
                
                # Preparar células para a tabela
                celulas_titulo = []
                celulas_conteudo = []
                
                for texto in textos_obj:
                    # Título
                    p_tit = Paragraph(f"<b>{texto.titulo}</b>", estilo_titulo_texto)
                    celulas_titulo.append(p_tit)
                    
                    # Conteúdo
                    if texto.tipo == TipoTextoBase.TEXTO:
                        p_txt = Paragraph(texto.conteudo or "", estilo_texto)
                        celulas_conteudo.append(p_txt)
                    else:  # Imagem como placeholder (simplificado)
                        p_txt = Paragraph(f"<i>[Imagem: {texto.titulo}]</i>", estilo_texto)
                        celulas_conteudo.append(p_txt)
                
                # Criar tabela
                table_textos = Table([celulas_titulo, celulas_conteudo], colWidths=[col_width, col_width])
                table_textos.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0,0), (-1,-1), 5),
                    ('RIGHTPADDING', (0,0), (-1,-1), 5),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                    ('TOPPADDING', (0,0), (-1,-1), 5),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ]))
                
                w_table, h_table = table_textos.wrap(width - margin_left - margin_right, y_position)
                
                if h_table > y_position - margin_bottom:
                    c.showPage()
                    y_position = height - margin_top
                
                table_textos.drawOn(c, margin_left, y_position - h_table)
                y_position -= h_table + 0.8 * cm
            
            else:
                # Mais de 2 textos - renderizar sequencialmente
                for texto in textos_obj:
                    if texto.tipo == TipoTextoBase.TEXTO:
                        estilo_titulo = ParagraphStyle('titulo_texto', parent=estilos['Normal'], alignment=TA_LEFT, fontName='Helvetica-Bold', fontSize=10)
                        p_tit = Paragraph(f"<b>{texto.titulo}</b>", estilo_titulo)
                        w_tit, h_tit = p_tit.wrap(width - margin_left - margin_right, y_position)
                        
                        if h_tit > y_position - margin_bottom:
                            c.showPage()
                            y_position = height - margin_top
                        
                        p_tit.drawOn(c, margin_left, y_position - h_tit)
                        y_position -= h_tit + 0.2 * cm
                        
                        estilo_texto = ParagraphStyle('texto_base', parent=estilos['Normal'], alignment=TA_JUSTIFY, fontName='Helvetica', fontSize=9, leading=11)
                        p_txt = Paragraph(texto.conteudo or "", estilo_texto)
                        w_txt, h_txt = p_txt.wrap(width - margin_left - margin_right, y_position)
                        
                        if h_txt > y_position - margin_bottom:
                            c.showPage()
                            y_position = height - margin_top
                        
                        p_txt.drawOn(c, margin_left, y_position - h_txt)
                        y_position -= h_txt + 0.5 * cm
        
        # ====================================================================
        # QUESTÕES
        # ====================================================================
        
        from banco_questoes.services import QuestaoService
        
        for ordem, questao_id in enumerate(self.questoes_avaliacao, 1):
            try:
                questao = QuestaoService.buscar_por_id(questao_id)
                
                if not questao:
                    continue
                
                # Verificar se precisa de nova página
                # Reservar espaço estimado: se espaço insuficiente, nova página
                if y_position < margin_bottom + 6*cm:
                    c.showPage()
                    y_position = height - margin_top

                # Número da questão seguido do enunciado (ex: "1. Texto")
                enunciado = questao.enunciado or ""
                estilo_enun = ParagraphStyle('enunciado', parent=getSampleStyleSheet()['Normal'], alignment=TA_JUSTIFY, fontName='Helvetica', fontSize=12, leading=14)
                p_enun = Paragraph(f"<b>{ordem}.</b> {enunciado.replace('\n', '<br/>')}", estilo_enun)
                w_enun, h_enun = p_enun.wrap(width - margin_left - margin_right, y_position)
                if h_enun > y_position - margin_bottom:
                    c.showPage()
                    y_position = height - margin_top
                p_enun.drawOn(c, margin_left, y_position - h_enun)
                y_position -= h_enun + 0.2 * cm

                # Alternativas (se múltipla escolha)
                if questao.tipo and getattr(questao.tipo, 'value', '') == "multipla_escolha":
                    alternativas = QuestaoService.buscar_alternativas(questao_id)
                    estilo_alt = ParagraphStyle('alternativa', parent=getSampleStyleSheet()['Normal'], alignment=TA_JUSTIFY, fontName='Helvetica', fontSize=12, leftIndent=12, leading=14)
                    for alt in alternativas:
                        texto_alt = f"{alt.letra}) {alt.texto or ''}"
                        p_alt = Paragraph(texto_alt, estilo_alt)
                        w_alt, h_alt = p_alt.wrap(width - margin_left - margin_right - 0.5*cm, y_position)
                        if h_alt > y_position - margin_bottom:
                            c.showPage()
                            y_position = height - margin_top
                        p_alt.drawOn(c, margin_left + 0.5*cm, y_position - h_alt)
                        y_position -= h_alt

                # Espaço para resposta (se dissertativa)
                elif questao.tipo and getattr(questao.tipo, 'value', '') == "dissertativa":
                    # Espaço para resposta (5 linhas)
                    for _ in range(5):
                        if y_position < margin_bottom + 0.5*cm:
                            c.showPage()
                            y_position = height - margin_top
                        c.line(margin_left + 0.5*cm, y_position, width - margin_right, y_position)
                        y_position -= 0.6 * cm

                # Espaçamento entre questões
                y_position -= 0.8 * cm
                
            except Exception as e:
                logger.error(f"Erro ao processar questão {questao_id}: {e}")
                continue
        
        # Salvar PDF
        c.save()
        logger.info(f"PDF gerado: {caminho_pdf}")
    
    def _quebrar_texto(self, texto: str, largura_max: float, canvas_obj, font_name: str, font_size: int) -> list:
        """Quebra texto em múltiplas linhas para caber na largura especificada."""
        palavras = texto.split()
        linhas = []
        linha_atual = ""
        
        for palavra in palavras:
            teste = f"{linha_atual} {palavra}".strip()
            largura_teste = canvas_obj.stringWidth(teste, font_name, font_size)
            
            if largura_teste <= largura_max:
                linha_atual = teste
            else:
                if linha_atual:
                    linhas.append(linha_atual)
                linha_atual = palavra
        
        if linha_atual:
            linhas.append(linha_atual)
        
        return linhas if linhas else [""]
    
    # =========================================================================
    # ABA: MINHAS QUESTÕES
    # =========================================================================
    
    def criar_aba_minhas_questoes(self):
        """Cria a aba de questões do usuário."""
        # Botão atualizar
        frame_top = tk.Frame(self.frame_minhas, bg=self.co0)
        frame_top.pack(fill="x", padx=10, pady=10)
        
        tk.Button(
            frame_top, text="🔄 Atualizar Lista",
            command=self.carregar_minhas_questoes,
            bg=self.co4, fg="white"
        ).pack(side="left")
        
        # Treeview
        frame_tree = tk.Frame(self.frame_minhas, bg=self.co0)
        frame_tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        colunas = ("id", "componente", "ano", "habilidade", "tipo", "status", "data")
        self.tree_minhas = ttk.Treeview(
            frame_tree, columns=colunas, show="headings", height=15
        )
        
        self.tree_minhas.heading("id", text="ID")
        self.tree_minhas.heading("componente", text="Componente")
        self.tree_minhas.heading("ano", text="Ano")
        self.tree_minhas.heading("habilidade", text="Habilidade")
        self.tree_minhas.heading("tipo", text="Tipo")
        self.tree_minhas.heading("status", text="Status")
        self.tree_minhas.heading("data", text="Criada em")
        
        self.tree_minhas.column("id", width=50)
        self.tree_minhas.column("componente", width=130)
        self.tree_minhas.column("ano", width=70)
        self.tree_minhas.column("habilidade", width=100)
        self.tree_minhas.column("tipo", width=110)
        self.tree_minhas.column("status", width=80)
        self.tree_minhas.column("data", width=100)
        
        scrollbar = ttk.Scrollbar(frame_tree, orient="vertical", command=self.tree_minhas.yview)
        self.tree_minhas.configure(yscrollcommand=scrollbar.set)
        
        self.tree_minhas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Duplo clique para editar
        self.tree_minhas.bind("<Double-1>", lambda e: self.editar_minha_questao())
        
        # Botões de ação
        frame_acoes = tk.Frame(self.frame_minhas, bg=self.co0)
        frame_acoes.pack(fill="x", padx=10, pady=5)
        
        tk.Button(
            frame_acoes, text="✏️ Editar",
            command=self.editar_minha_questao,
            bg=self.co4, fg="white"
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_acoes, text="❌ Excluir",
            command=self.excluir_minha_questao,
            bg=self.co8, fg="white"
        ).pack(side="left", padx=5)
        
        # WORKFLOW DE APROVAÇÃO
        tk.Label(frame_acoes, text="|", bg=self.co0).pack(side="left", padx=5)
        
        tk.Button(
            frame_acoes, text="📤 Enviar p/ Revisão",
            command=self.enviar_questao_revisao,
            bg="#FF8C00", fg="white"
        ).pack(side="left", padx=5)
        
        if self.perfil in ['administrador', 'coordenador']:
            tk.Button(
                frame_acoes, text="✅ Aprovar",
                command=self.aprovar_questao,
                bg=self.co2, fg="white"
            ).pack(side="left", padx=5)
            
            tk.Button(
                frame_acoes, text="↩️ Devolver",
                command=self.devolver_questao,
                bg="#FFA500", fg="white"
            ).pack(side="left", padx=5)
        
        # Carregar lista ao criar aba
        self.janela.after(500, self.carregar_minhas_questoes)
        
        # Carregar automaticamente
        self.janela.after(500, self.carregar_minhas_questoes)
    
    def carregar_minhas_questoes(self):
        """Carrega as questões do usuário logado."""
        from banco_questoes.services import QuestaoService
        from banco_questoes.models import FiltroQuestoes
        
        # Limpar
        for item in self.tree_minhas.get_children():
            self.tree_minhas.delete(item)
        
        # Se perfis não estão habilitados, mostrar TODAS as questões da escola
        if not perfis_habilitados():
            logger.info("Perfis desabilitados - carregando todas as questões da escola")
            try:
                import config
                filtros = FiltroQuestoes(escola_id=config.ESCOLA_ID)
                questoes, total = QuestaoService.buscar(filtros, limite=100)
                
                logger.info(f"Carregadas {len(questoes)} questões da escola (total: {total})")
                
                for q in questoes:
                    data_str = ''
                    if q.created_at:
                        data_str = q.created_at.strftime('%Y-%m-%d') if hasattr(q.created_at, 'strftime') else str(q.created_at)[:10]
                    
                    self.tree_minhas.insert("", "end", values=(
                        q.id,
                        q.componente_curricular.value if q.componente_curricular else '',
                        q.ano_escolar.value if q.ano_escolar else '',
                        q.habilidade_bncc_codigo or '',
                        q.tipo.value if q.tipo else '',
                        q.status.value if q.status else '',
                        data_str
                    ))
                    
                if not questoes:
                    messagebox.showinfo(
                        "Info", 
                        "Nenhuma questão cadastrada na escola.\n\n"
                        "Vá na aba '➕ Cadastrar Questão' para criar a primeira questão!"
                    )
                    
            except Exception as e:
                logger.error(f"Erro ao carregar questões da escola: {e}")
                messagebox.showerror("Erro", f"Erro ao carregar questões: {e}")
            return
        
        # Perfis habilitados - verificar autenticação
        if not self.funcionario_id:
            logger.warning("Perfis habilitados mas nenhum funcionário logado")
            messagebox.showwarning(
                "Aviso", 
                "Sistema de perfis habilitado mas usuário não autenticado.\n\n"
                "Mostrando todas as questões da escola como fallback."
            )
            # Fallback: carregar todas da escola
            try:
                import config
                filtros = FiltroQuestoes(escola_id=config.ESCOLA_ID)
                questoes, total = QuestaoService.buscar(filtros, limite=100)
                
                logger.info(f"Carregadas {len(questoes)} questões da escola (fallback - total: {total})")
                
                for q in questoes:
                    data_str = ''
                    if q.created_at:
                        data_str = q.created_at.strftime('%Y-%m-%d') if hasattr(q.created_at, 'strftime') else str(q.created_at)[:10]
                    
                    self.tree_minhas.insert("", "end", values=(
                        q.id,
                        q.componente_curricular.value if q.componente_curricular else '',
                        q.ano_escolar.value if q.ano_escolar else '',
                        q.habilidade_bncc_codigo or '',
                        q.tipo.value if q.tipo else '',
                        q.status.value if q.status else '',
                        data_str
                    ))
                    
                if not questoes:
                    messagebox.showinfo("Info", "Nenhuma questão encontrada no banco de dados.")
                    
            except Exception as e:
                logger.error(f"Erro ao carregar questões da escola: {e}")
                messagebox.showerror("Erro", f"Erro ao carregar questões: {e}")
            return
        
        # Perfis habilitados E usuário autenticado - filtrar por autor
        try:
            logger.info(f"Carregando questões do autor ID: {self.funcionario_id}")
            filtros = FiltroQuestoes(autor_id=self.funcionario_id)
            questoes, total = QuestaoService.buscar(filtros, limite=100)
            
            logger.info(f"Encontradas {len(questoes)} questões do autor (total: {total})")
            
            for q in questoes:
                data_str = ''
                if q.created_at:
                    data_str = q.created_at.strftime('%Y-%m-%d') if hasattr(q.created_at, 'strftime') else str(q.created_at)[:10]
                
                self.tree_minhas.insert("", "end", values=(
                    q.id,
                    q.componente_curricular.value if q.componente_curricular else '',
                    q.ano_escolar.value if q.ano_escolar else '',
                    q.habilidade_bncc_codigo or '',
                    q.tipo.value if q.tipo else '',
                    q.status.value if q.status else '',
                    data_str
                ))
            
            if not questoes:
                messagebox.showinfo(
                    "Info", 
                    f"Você ainda não criou nenhuma questão.\n\n"
                    f"Vá na aba '➕ Cadastrar Questão' para criar sua primeira questão!"
                )
                
        except Exception as e:
            logger.error(f"Erro ao carregar questões: {e}")
            messagebox.showerror("Erro", f"Erro ao carregar questões: {e}")
    
    def editar_minha_questao(self):
        """Abre questão para edição."""
        selecao = self.tree_minhas.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para editar.")
            return
        
        item = self.tree_minhas.item(selecao[0])
        questao_id = int(item['values'][0])
        
        # Buscar dados da questão
        from banco_questoes.services import QuestaoService
        questao = QuestaoService.buscar_por_id(questao_id)
        
        if not questao:
            messagebox.showerror("Erro", "Questão não encontrada.")
            return
        
        # CONTROLE DE PERMISSÕES GRANULAR
        if perfis_habilitados():
            # Verificar se pode editar esta questão
            pode_editar_todas = self.perfil in ['administrador', 'coordenador']
            e_autor = (questao.autor_id == self.funcionario_id)
            
            if not pode_editar_todas and not e_autor:
                messagebox.showerror(
                    "Sem Permissão",
                    "❌ Você não tem permissão para editar esta questão.\n\n"
                    "Você só pode editar questões criadas por você."
                )
                return
        
        # Verificar se tem aba de cadastro
        if not hasattr(self, 'frame_cadastro'):
            messagebox.showwarning("Aviso", "A aba de cadastro não está disponível para o seu perfil.")
            return
        
        # Mudar para aba de cadastro
        self.notebook.select(self.frame_cadastro)
        
        # Preencher campos com dados da questão
        self.carregar_questao_para_edicao(questao)
    
    def excluir_minha_questao(self):
        """Exclui a questão selecionada."""
        selecao = self.tree_minhas.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para excluir.")
            return
        
        item = self.tree_minhas.item(selecao[0])
        questao_id = int(item['values'][0])
        
        # CONTROLE DE PERMISSÕES GRANULAR
        if perfis_habilitados():
            try:
                from banco_questoes.services import QuestaoService
                questao = QuestaoService.buscar_por_id(questao_id)
                if questao:
                    pode_excluir_todas = self.perfil in ['administrador', 'coordenador']
                    e_autor = (questao.autor_id == self.funcionario_id)
                    
                    if not pode_excluir_todas and not e_autor:
                        messagebox.showerror(
                            "Sem Permissão",
                            "❌ Você não tem permissão para excluir esta questão.\n\n"
                            "Você só pode excluir questões criadas por você."
                        )
                        return
            except Exception as e:
                logger.error(f"Erro ao verificar permissões: {e}")
        
        if not messagebox.askyesno("Confirmar", f"Deseja realmente excluir a questão #{questao_id}?\n\nEsta ação não pode ser desfeita."):
            return
        
        try:
            from banco_questoes.services import QuestaoService
            sucesso = QuestaoService.excluir(questao_id)
            
            if sucesso:
                self.tree_minhas.delete(selecao[0])
                messagebox.showinfo("Sucesso", "Questão excluída com sucesso!")
            else:
                messagebox.showwarning(
                    "Aviso", 
                    "Não foi possível excluir a questão.\n"
                    "Ela pode estar sendo usada em alguma avaliação."
                )
            
        except Exception as e:
            logger.error(f"Erro ao excluir: {e}")
            messagebox.showerror("Erro", f"Erro ao excluir questão: {e}")
    
    # =========================================================================
    # WORKFLOW DE APROVAÇÃO DE QUESTÕES
    # =========================================================================
    
    def enviar_questao_revisao(self):
        """Envia questão para revisão (rascunho → revisão)."""
        selecao = self.tree_minhas.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para enviar para revisão.")
            return
        
        item = self.tree_minhas.item(selecao[0])
        questao_id = int(item['values'][0])
        status_atual = item['values'][5]  # Coluna status
        
        if status_atual != 'rascunho':
            messagebox.showinfo(
                "Info",
                f"Esta questão já está em status '{status_atual}'.\n\n"
                "Apenas questões em rascunho podem ser enviadas para revisão."
            )
            return
        
        confirmacao = messagebox.askyesno(
            "Confirmar Envio",
            f"Enviar questão #{questao_id} para revisão?\n\n"
            "A questão ficará disponível para análise do coordenador/administrador."
        )
        
        if not confirmacao:
            return
        
        try:
            from banco_questoes.services import QuestaoService
            sucesso = QuestaoService.alterar_status(questao_id, 'revisao', self.funcionario_id)
            
            if sucesso:
                messagebox.showinfo(
                    "Sucesso",
                    "✅ Questão enviada para revisão!\n\n"
                    "Aguarde análise da coordenação."
                )
                self.carregar_minhas_questoes()
            else:
                messagebox.showerror("Erro", "Não foi possível alterar o status da questão.")
                
        except Exception as e:
            logger.error(f"Erro ao enviar questão para revisão: {e}")
            messagebox.showerror("Erro", f"Erro ao processar solicitação: {e}")
    
    def aprovar_questao(self):
        """Aprova questão (revisão → aprovada). Apenas coordenador/admin."""
        if self.perfil not in ['administrador', 'coordenador']:
            messagebox.showerror("Sem Permissão", "Apenas coordenadores e administradores podem aprovar questões.")
            return
        
        selecao = self.tree_minhas.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para aprovar.")
            return
        
        item = self.tree_minhas.item(selecao[0])
        questao_id = int(item['values'][0])
        status_atual = item['values'][5]
        
        if status_atual not in ['revisao', 'rascunho']:
            messagebox.showinfo(
                "Info",
                f"Esta questão está em status '{status_atual}'.\n\n"
                "Apenas questões em revisão ou rascunho podem ser aprovadas."
            )
            return
        
        # Janela de aprovação com comentário opcional
        from tkinter import simpledialog
        comentario = simpledialog.askstring(
            "Aprovar Questão",
            f"Aprovando questão #{questao_id}\n\n"
            "Comentário (opcional):",
            parent=self.janela
        )
        
        try:
            from banco_questoes.services import QuestaoService
            sucesso = QuestaoService.aprovar_questao(
                questao_id,
                aprovador_id=self.funcionario_id,
                comentario=comentario
            )
            
            if sucesso:
                messagebox.showinfo(
                    "Sucesso",
                    f"✅ Questão #{questao_id} aprovada!\n\n"
                    "A questão está pronta para uso em avaliações."
                )
                self.carregar_minhas_questoes()
            else:
                messagebox.showerror("Erro", "Não foi possível aprovar a questão.")
                
        except Exception as e:
            logger.error(f"Erro ao aprovar questão: {e}")
            messagebox.showerror("Erro", f"Erro ao processar aprovação: {e}")
    
    def devolver_questao(self):
        """Devolve questão para revisão (aprovada/revisão → rascunho). Apenas coordenador/admin."""
        if self.perfil not in ['administrador', 'coordenador']:
            messagebox.showerror("Sem Permissão", "Apenas coordenadores e administradores podem devolver questões.")
            return
        
        selecao = self.tree_minhas.selection()
        if not selecao:
            messagebox.showwarning("Aviso", "Selecione uma questão para devolver.")
            return
        
        item = self.tree_minhas.item(selecao[0])
        questao_id = int(item['values'][0])
        status_atual = item['values'][5]
        
        if status_atual == 'rascunho':
            messagebox.showinfo("Info", "Esta questão já está em rascunho.")
            return
        
        # Janela de devolução com motivo obrigatório
        from tkinter import simpledialog
        motivo = simpledialog.askstring(
            "Devolver Questão",
            f"Devolvendo questão #{questao_id} para revisão\n\n"
            "Motivo da devolução (obrigatório):",
            parent=self.janela
        )
        
        if not motivo or not motivo.strip():
            messagebox.showwarning("Aviso", "É necessário informar o motivo da devolução.")
            return
        
        try:
            from banco_questoes.services import QuestaoService
            sucesso = QuestaoService.devolver_questao(
                questao_id,
                revisor_id=self.funcionario_id,
                motivo=motivo
            )
            
            if sucesso:
                messagebox.showinfo(
                    "Sucesso",
                    f"↩️ Questão #{questao_id} devolvida para revisão!\n\n"
                    "O autor será notificado sobre as correções necessárias."
                )
                self.carregar_minhas_questoes()
            else:
                messagebox.showerror("Erro", "Não foi possível devolver a questão.")
                
        except Exception as e:
            logger.error(f"Erro ao devolver questão: {e}")
            messagebox.showerror("Erro", f"Erro ao processar devolução: {e}")
    
    # =========================================================================
    # ABA: ESTATÍSTICAS
    # =========================================================================
    
    def criar_aba_estatisticas(self):
        """Cria a aba de estatísticas (admin/coordenador)."""
        tk.Label(
            self.frame_estatisticas,
            text="📊 Estatísticas do Banco de Questões",
            font=("Arial", 14, "bold"),
            bg=self.co0
        ).pack(pady=20)
        
        frame_stats = tk.Frame(self.frame_estatisticas, bg=self.co0)
        frame_stats.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Placeholder - estatísticas serão carregadas aqui
        self.lbl_stats = tk.Label(
            frame_stats,
            text="Carregando estatísticas...",
            bg=self.co0, font=("Arial", 11)
        )
        self.lbl_stats.pack(pady=20)
        
        tk.Button(
            frame_stats, text="🔄 Atualizar Estatísticas",
            command=self.carregar_estatisticas,
            bg=self.co4, fg="white"
        ).pack(pady=10)
        
        self.janela.after(500, self.carregar_estatisticas)
    
    def carregar_estatisticas(self):
        """Carrega estatísticas do banco de questões com visualização aprimorada."""
        try:
            from banco_questoes.services import EstatisticasService
            import config
            
            # Obter estatísticas gerais
            stats = EstatisticasService.obter_estatisticas_gerais(
                escola_id=config.ESCOLA_ID,
                autor_id=self.funcionario_id if perfis_habilitados() else None
            )
            
            if not stats:
                self.lbl_stats.config(text="Nenhuma estatística disponível.")
                return
            
            # Criar interface rica com frames
            for widget in self.frame_estatisticas.winfo_children():
                if widget != self.lbl_stats.master:
                    widget.destroy()
            
            # Frame principal com scroll
            canvas = tk.Canvas(self.frame_estatisticas, bg=self.co0)
            scrollbar = ttk.Scrollbar(self.frame_estatisticas, orient="vertical", command=canvas.yview)
            frame_scroll = tk.Frame(canvas, bg=self.co0)
            
            frame_scroll.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=frame_scroll, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Título
            tk.Label(
                frame_scroll,
                text="📊 Estatísticas do Banco de Questões",
                font=("Arial", 16, "bold"),
                bg=self.co0, fg=self.co1
            ).pack(pady=10)
            
            # Botão atualizar
            tk.Button(
                frame_scroll, text="🔄 Atualizar",
                command=self.carregar_estatisticas,
                bg=self.co4, fg="white"
            ).pack(pady=5)
            
            # CARD: Total de Questões
            frame_total = tk.LabelFrame(
                frame_scroll, text="📚 Total de Questões",
                bg=self.co0, font=("Arial", 11, "bold")
            )
            frame_total.pack(fill="x", padx=20, pady=10)
            
            tk.Label(
                frame_total,
                text=str(stats['total_questoes']),
                font=("Arial", 32, "bold"),
                bg=self.co0, fg=self.co2
            ).pack(pady=10)
            
            # CARD: Por Status
            frame_status = tk.LabelFrame(
                frame_scroll, text="📋 Por Status",
                bg=self.co0, font=("Arial", 11, "bold")
            )
            frame_status.pack(fill="x", padx=20, pady=10)
            
            cores_status = {
                'rascunho': '#999999',
                'revisao': '#FF8C00',
                'aprovada': '#77B341',
                'arquivada': '#666666'
            }
            
            for status, qtd in stats['por_status'].items():
                frame_item = tk.Frame(frame_status, bg=self.co0)
                frame_item.pack(fill="x", padx=10, pady=5)
                
                tk.Label(
                    frame_item,
                    text=f"{status.capitalize()}:",
                    bg=self.co0,
                    font=("Arial", 10)
                ).pack(side="left")
                
                tk.Label(
                    frame_item,
                    text=str(qtd),
                    bg=cores_status.get(status, self.co9),
                    fg="white",
                    font=("Arial", 10, "bold"),
                    padx=10
                ).pack(side="right")
            
            # CARD: Por Tipo
            frame_tipo = tk.LabelFrame(
                frame_scroll, text="📝 Por Tipo",
                bg=self.co0, font=("Arial", 11, "bold")
            )
            frame_tipo.pack(fill="x", padx=20, pady=10)
            
            for tipo, qtd in stats['por_tipo'].items():
                frame_item = tk.Frame(frame_tipo, bg=self.co0)
                frame_item.pack(fill="x", padx=10, pady=5)
                
                tk.Label(
                    frame_item,
                    text=f"{tipo.replace('_', ' ').title()}:",
                    bg=self.co0,
                    font=("Arial", 10)
                ).pack(side="left")
                
                tk.Label(
                    frame_item,
                    text=str(qtd),
                    bg=self.co4,
                    fg="white",
                    font=("Arial", 10, "bold"),
                    padx=10
                ).pack(side="right")
            
            # CARD: Por Dificuldade
            frame_dif = tk.LabelFrame(
                frame_scroll, text="⚡ Por Dificuldade",
                bg=self.co0, font=("Arial", 11, "bold")
            )
            frame_dif.pack(fill="x", padx=20, pady=10)
            
            cores_dif = {
                'facil': '#77B341',
                'media': '#FF8C00',
                'dificil': '#BF3036'
            }
            
            for dif, qtd in stats['por_dificuldade'].items():
                frame_item = tk.Frame(frame_dif, bg=self.co0)
                frame_item.pack(fill="x", padx=10, pady=5)
                
                tk.Label(
                    frame_item,
                    text=f"{dif.capitalize()}:",
                    bg=self.co0,
                    font=("Arial", 10)
                ).pack(side="left")
                
                tk.Label(
                    frame_item,
                    text=str(qtd),
                    bg=cores_dif.get(dif, self.co9),
                    fg="white",
                    font=("Arial", 10, "bold"),
                    padx=10
                ).pack(side="right")
            
            # CARD: Questões Mais Utilizadas
            if stats['mais_utilizadas']:
                frame_top = tk.LabelFrame(
                    frame_scroll, text="⭐ Questões Mais Utilizadas",
                    bg=self.co0, font=("Arial", 11, "bold")
                )
                frame_top.pack(fill="x", padx=20, pady=10)
                
                for item in stats['mais_utilizadas'][:5]:
                    frame_item = tk.Frame(frame_top, bg=self.co0)
                    frame_item.pack(fill="x", padx=10, pady=3)
                    
                    tk.Label(
                        frame_item,
                        text=f"#{item['id']} - {item['enunciado']}",
                        bg=self.co0,
                        font=("Arial", 9),
                        anchor="w"
                    ).pack(side="left", fill="x", expand=True)
                    
                    tk.Label(
                        frame_item,
                        text=f"{item['vezes']}x",
                        bg=self.co2,
                        fg="white",
                        font=("Arial", 9, "bold"),
                        padx=8
                    ).pack(side="right")
            
        except Exception as e:
            logger.error(f"Erro ao carregar estatísticas: {e}")
            messagebox.showerror("Erro", f"Erro ao carregar estatísticas:\n{e}")
    
    # =========================================================================
    # ABA TEXTOS BASE - GERENCIAMENTO DE TEXTOS/IMAGENS
    # =========================================================================
    
    def criar_aba_textos_base(self):
        """Cria a aba de gerenciamento de textos base."""
        # Container principal com scroll
        container = tk.Frame(self.frame_textos_base, bg=self.co0)
        container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Título
        tk.Label(
            container,
            text="Gerenciamento de Textos Base",
            font=("Arial", 14, "bold"),
            bg=self.co0, fg=self.co1
        ).pack(anchor="w", pady=(0, 10))
        
        tk.Label(
            container,
            text="Crie textos ou selecione imagens que podem ser reutilizados em múltiplas questões e avaliações.",
            font=("Arial", 10),
            bg=self.co0, fg=self.co7
        ).pack(anchor="w", pady=(0, 10))
        
        # Frame de ações no topo
        frame_acoes = tk.Frame(container, bg=self.co0)
        frame_acoes.pack(fill="x", pady=(0, 10))
        
        tk.Button(
            frame_acoes,
            text="➕ Novo Texto Base",
            command=self.novo_texto_base,
            bg=self.co2, fg="white",
            font=("Arial", 10, "bold"),
            cursor="hand2",
            relief="flat",
            padx=15, pady=8
        ).pack(side="left", padx=(0, 5))
        
        tk.Button(
            frame_acoes,
            text="✏️ Editar",
            command=self.editar_texto_base,
            bg=self.co4, fg="white",
            font=("Arial", 10, "bold"),
            cursor="hand2",
            relief="flat",
            padx=15, pady=8
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_acoes,
            text="🗑️ Excluir",
            command=self.excluir_texto_base,
            bg=self.co8, fg="white",
            font=("Arial", 10, "bold"),
            cursor="hand2",
            relief="flat",
            padx=15, pady=8
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_acoes,
            text="🔄 Atualizar Lista",
            command=self.carregar_lista_textos_base,
            bg=self.co9, fg="white",
            font=("Arial", 10, "bold"),
            cursor="hand2",
            relief="flat",
            padx=15, pady=8
        ).pack(side="left", padx=5)
        
        # Treeview para listagem
        frame_tree = tk.Frame(container, bg=self.co0)
        frame_tree.pack(fill="both", expand=True)
        
        # Scrollbars
        scroll_y = tk.Scrollbar(frame_tree, orient="vertical")
        scroll_y.pack(side="right", fill="y")
        
        scroll_x = tk.Scrollbar(frame_tree, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")
        
        # Treeview
        self.tree_textos_base = ttk.Treeview(
            frame_tree,
            columns=("id", "titulo", "tipo", "criado_em", "autor"),
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            height=15
        )
        
        self.tree_textos_base.heading("id", text="ID")
        self.tree_textos_base.heading("titulo", text="Título")
        self.tree_textos_base.heading("tipo", text="Tipo")
        self.tree_textos_base.heading("criado_em", text="Criado em")
        self.tree_textos_base.heading("autor", text="Autor")
        
        self.tree_textos_base.column("id", width=50, anchor="center")
        self.tree_textos_base.column("titulo", width=300, anchor="w")
        self.tree_textos_base.column("tipo", width=100, anchor="center")
        self.tree_textos_base.column("criado_em", width=150, anchor="center")
        self.tree_textos_base.column("autor", width=200, anchor="w")
        
        self.tree_textos_base.pack(fill="both", expand=True)
        
        scroll_y.config(command=self.tree_textos_base.yview)
        scroll_x.config(command=self.tree_textos_base.xview)
        
        # Doubleclick para editar
        self.tree_textos_base.bind("<Double-1>", lambda e: self.editar_texto_base())
        
        # Carregar dados
        self.carregar_lista_textos_base()
    
    def carregar_lista_textos_base(self):
        """Carrega a lista de textos base no treeview."""
        try:
            # Limpar treeview
            for item in self.tree_textos_base.get_children():
                self.tree_textos_base.delete(item)
            
            # Buscar textos base
            escola_id = config.ESCOLA_ID
            textos = TextoBaseService.listar(escola_id=escola_id)
            
            for texto in textos:
                tipo_display = "📝 Texto" if texto.tipo == TipoTextoBase.TEXTO else "🖼️ Imagem"
                autor = texto.autor_id or "N/A"
                criado = texto.created_at.strftime("%d/%m/%Y %H:%M") if texto.created_at else "N/A"
                
                self.tree_textos_base.insert(
                    "",
                    "end",
                    values=(texto.id, texto.titulo, tipo_display, criado, autor)
                )
            
            logger.info(f"{len(textos)} textos base carregados")
            
        except Exception as e:
            logger.error(f"Erro ao carregar textos base: {e}")
            messagebox.showerror("Erro", f"Erro ao carregar textos base:\n{e}")
    
    def novo_texto_base(self):
        """Abre janela para criar novo texto base."""
        self.abrir_janela_texto_base(modo="criar")
    
    def editar_texto_base(self):
        """Abre janela para editar texto base selecionado."""
        selecionado = self.tree_textos_base.selection()
        if not selecionado:
            messagebox.showwarning("Atenção", "Selecione um texto base para editar.")
            return
        
        item = self.tree_textos_base.item(selecionado[0])
        texto_id = item["values"][0]
        self.abrir_janela_texto_base(modo="editar", texto_id=texto_id)
    
    def excluir_texto_base(self):
        """Exclui o texto base selecionado."""
        selecionado = self.tree_textos_base.selection()
        if not selecionado:
            messagebox.showwarning("Atenção", "Selecione um texto base para excluir.")
            return
        
        item = self.tree_textos_base.item(selecionado[0])
        texto_id = item["values"][0]
        titulo = item["values"][1]
        
        confirmacao = messagebox.askyesno(
            "Confirmar Exclusão",
            f"Deseja realmente excluir o texto base:\n\n'{titulo}'?\n\nEsta ação não pode ser desfeita."
        )
        
        if confirmacao:
            try:
                TextoBaseService.excluir(texto_id)
                messagebox.showinfo("Sucesso", "Texto base excluído com sucesso!")
                self.carregar_lista_textos_base()
            except Exception as e:
                logger.error(f"Erro ao excluir texto base: {e}")
                messagebox.showerror("Erro", f"Erro ao excluir texto base:\n{e}")
    
    def abrir_janela_texto_base(self, modo="criar", texto_id=None):
        """
        Abre janela para criar/editar texto base.
        
        Args:
            modo: 'criar' ou 'editar'
            texto_id: ID do texto (para modo editar)
        """
        janela = tk.Toplevel(self.janela)
        janela.title(f"{'Novo' if modo == 'criar' else 'Editar'} Texto Base")
        janela.geometry("700x600")
        janela.configure(bg=self.co0)
        janela.grab_set()
        
        # Variáveis
        var_tipo = tk.StringVar(value="texto")
        var_titulo = tk.StringVar()
        var_conteudo = tk.StringVar()
        var_caminho_imagem = tk.StringVar()
        
        # Carregar dados se editar
        if modo == "editar" and texto_id:
            try:
                texto = TextoBaseService.buscar_por_id(texto_id)
                if texto:
                    var_titulo.set(texto.titulo)
                    var_tipo.set(texto.tipo.value)
                    if texto.tipo == TipoTextoBase.TEXTO:
                        var_conteudo.set(texto.conteudo or "")
                    else:
                        var_caminho_imagem.set(texto.caminho_imagem or "")
            except Exception as e:
                logger.error(f"Erro ao carregar texto base: {e}")
                messagebox.showerror("Erro", f"Erro ao carregar dados:\n{e}")
                janela.destroy()
                return
        
        # Container com scroll
        container = tk.Frame(janela, bg=self.co0)
        container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        tk.Label(
            container,
            text=f"{'Criar Novo' if modo == 'criar' else 'Editar'} Texto Base",
            font=("Arial", 14, "bold"),
            bg=self.co0, fg=self.co1
        ).pack(anchor="w", pady=(0, 20))
        
        # Campo: Título
        frame_titulo = tk.Frame(container, bg=self.co0)
        frame_titulo.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            frame_titulo,
            text="Título:",
            font=("Arial", 10, "bold"),
            bg=self.co0, fg=self.co7
        ).pack(anchor="w")
        
        tk.Entry(
            frame_titulo,
            textvariable=var_titulo,
            font=("Arial", 11),
            relief="solid",
            bd=1
        ).pack(fill="x", pady=(5, 0))
        
        # Campo: Tipo
        frame_tipo = tk.Frame(container, bg=self.co0)
        frame_tipo.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            frame_tipo,
            text="Tipo:",
            font=("Arial", 10, "bold"),
            bg=self.co0, fg=self.co7
        ).pack(anchor="w")
        
        frame_radio = tk.Frame(frame_tipo, bg=self.co0)
        frame_radio.pack(anchor="w", pady=(5, 0))
        
        def atualizar_campos():
            """Mostra/oculta campos baseado no tipo selecionado."""
            if var_tipo.get() == "texto":
                frame_conteudo.pack(fill="both", expand=True, pady=(0, 15))
                frame_imagem.pack_forget()
            else:
                frame_conteudo.pack_forget()
                frame_imagem.pack(fill="x", pady=(0, 15))
        
        tk.Radiobutton(
            frame_radio,
            text="📝 Texto",
            variable=var_tipo,
            value="texto",
            font=("Arial", 10),
            bg=self.co0,
            command=atualizar_campos
        ).pack(side="left", padx=(0, 20))
        
        tk.Radiobutton(
            frame_radio,
            text="🖼️ Imagem",
            variable=var_tipo,
            value="imagem",
            font=("Arial", 10),
            bg=self.co0,
            command=atualizar_campos
        ).pack(side="left")
        
        # Campo: Conteúdo (para texto)
        frame_conteudo = tk.Frame(container, bg=self.co0)
        
        tk.Label(
            frame_conteudo,
            text="Conteúdo do Texto:",
            font=("Arial", 10, "bold"),
            bg=self.co0, fg=self.co7
        ).pack(anchor="w")
        
        scroll_texto = tk.Scrollbar(frame_conteudo)
        scroll_texto.pack(side="right", fill="y")
        
        texto_widget = tk.Text(
            frame_conteudo,
            font=("Arial", 11),
            relief="solid",
            bd=1,
            wrap="word",
            yscrollcommand=scroll_texto.set,
            height=12
        )
        texto_widget.pack(fill="both", expand=True, pady=(5, 0))
        scroll_texto.config(command=texto_widget.yview)
        
        if var_conteudo.get():
            texto_widget.insert("1.0", var_conteudo.get())
        
        # Campo: Imagem
        frame_imagem = tk.Frame(container, bg=self.co0)
        
        tk.Label(
            frame_imagem,
            text="Arquivo de Imagem:",
            font=("Arial", 10, "bold"),
            bg=self.co0, fg=self.co7
        ).pack(anchor="w")
        
        frame_img_input = tk.Frame(frame_imagem, bg=self.co0)
        frame_img_input.pack(fill="x", pady=(5, 0))
        
        tk.Entry(
            frame_img_input,
            textvariable=var_caminho_imagem,
            font=("Arial", 11),
            relief="solid",
            bd=1,
            state="readonly"
        ).pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        def selecionar_imagem():
            """Abre dialog para selecionar imagem."""
            caminho = filedialog.askopenfilename(
                title="Selecionar Imagem",
                filetypes=[
                    ("Imagens", "*.png;*.jpg;*.jpeg;*.gif;*.bmp"),
                    ("Todos os arquivos", "*.*")
                ]
            )
            if caminho:
                var_caminho_imagem.set(caminho)
        
        tk.Button(
            frame_img_input,
            text="📂 Escolher",
            command=selecionar_imagem,
            bg=self.co4, fg="white",
            font=("Arial", 9, "bold"),
            cursor="hand2",
            relief="flat",
            padx=15, pady=5
        ).pack(side="left")
        
        # Exibir campo correto inicialmente
        atualizar_campos()
        
        # Botões de ação
        frame_botoes = tk.Frame(container, bg=self.co0)
        frame_botoes.pack(fill="x", pady=(20, 0))
        
        def salvar():
            """Salva o texto base."""
            titulo = var_titulo.get().strip()
            tipo = var_tipo.get()
            
            if not titulo:
                messagebox.showwarning("Atenção", "O título é obrigatório.")
                return
            
            try:
                if tipo == "texto":
                    conteudo = texto_widget.get("1.0", "end-1c").strip()
                    if not conteudo:
                        messagebox.showwarning("Atenção", "O conteúdo do texto é obrigatório.")
                        return
                    
                    if modo == "criar":
                        TextoBaseService.criar(
                            titulo=titulo,
                            tipo=TipoTextoBase.TEXTO,
                            conteudo=conteudo,
                            escola_id=config.ESCOLA_ID,
                            autor_id=self.funcionario_id
                        )
                    else:
                        # Buscar texto existente e atualizar
                        texto = TextoBaseService.buscar_por_id(texto_id)
                        if texto:
                            texto.titulo = titulo
                            texto.conteudo = conteudo
                            # Atualizar no banco (será necessário método de update)
                            # Por enquanto, recriar
                            TextoBaseService.excluir(texto_id)
                            TextoBaseService.criar(
                                titulo=titulo,
                                tipo=TipoTextoBase.TEXTO,
                                conteudo=conteudo,
                                escola_id=config.ESCOLA_ID,
                                autor_id=self.funcionario_id
                            )
                else:
                    caminho = var_caminho_imagem.get().strip()
                    if not caminho:
                        messagebox.showwarning("Atenção", "Selecione uma imagem.")
                        return
                    
                    # Copiar imagem para pasta uploads
                    import os
                    import shutil
                    from datetime import datetime
                    
                    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")
                    os.makedirs(uploads_dir, exist_ok=True)
                    
                    ext = os.path.splitext(caminho)[1]
                    nome_arquivo = f"texto_base_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
                    destino = os.path.join(uploads_dir, nome_arquivo)
                    
                    shutil.copy2(caminho, destino)
                    
                    # Obter dimensões da imagem
                    from PIL import Image
                    with Image.open(destino) as img:
                        largura, altura = img.size
                    
                    if modo == "criar":
                        TextoBaseService.criar(
                            titulo=titulo,
                            tipo=TipoTextoBase.IMAGEM,
                            caminho_imagem=destino,
                            largura=largura,
                            altura=altura,
                            escola_id=config.ESCOLA_ID,
                            autor_id=self.funcionario_id
                        )
                    else:
                        # Atualizar (recriar)
                        TextoBaseService.excluir(texto_id)
                        TextoBaseService.criar(
                            titulo=titulo,
                            tipo=TipoTextoBase.IMAGEM,
                            caminho_imagem=destino,
                            largura=largura,
                            altura=altura,
                            escola_id=config.ESCOLA_ID,
                            autor_id=self.funcionario_id
                        )
                
                messagebox.showinfo("Sucesso", f"Texto base {'criado' if modo == 'criar' else 'atualizado'} com sucesso!")
                janela.destroy()
                self.carregar_lista_textos_base()
                
            except Exception as e:
                logger.error(f"Erro ao salvar texto base: {e}")
                messagebox.showerror("Erro", f"Erro ao salvar:\n{e}")
        
        tk.Button(
            frame_botoes,
            text="💾 Salvar",
            command=salvar,
            bg=self.co2, fg="white",
            font=("Arial", 11, "bold"),
            cursor="hand2",
            relief="flat",
            padx=30, pady=10
        ).pack(side="left", padx=(0, 10))
        
        tk.Button(
            frame_botoes,
            text="❌ Cancelar",
            command=janela.destroy,
            bg=self.co9, fg="white",
            font=("Arial", 11, "bold"),
            cursor="hand2",
            relief="flat",
            padx=30, pady=10
        ).pack(side="left")
    
    # =========================================================================
    # FASE 1 - MELHORIAS CRÍTICAS: EDITOR DE IMAGENS E VALIDAÇÕES
    # =========================================================================
    
    def abrir_editor_imagem(self, caminho_imagem: str, tipo: str = 'enunciado', letra_alt: str = None):
        """
        Abre o editor de imagens integrado.
        
        Args:
            caminho_imagem: Caminho da imagem a editar
            tipo: 'enunciado' ou 'alternativa'
            letra_alt: Letra da alternativa (se tipo='alternativa')
        """
        import os
        if not caminho_imagem or not os.path.exists(caminho_imagem):
            messagebox.showwarning("Aviso", "Imagem não encontrada.")
            return
        
        try:
            from banco_questoes.ui.editor_imagem import EditorImagem
            
            editor = EditorImagem(
                parent=self.janela,
                caminho_imagem=caminho_imagem,
                callback=lambda caminho_editado: self._aplicar_imagem_editada(
                    caminho_editado, tipo, letra_alt
                )
            )
            editor.abrir()
        except ImportError as e:
            logger.error(f"Erro ao importar editor: {e}")
            messagebox.showerror("Erro", "Módulo de edição de imagens não disponível.")
        except Exception as e:
            logger.error(f"Erro ao abrir editor de imagem: {e}")
            messagebox.showerror("Erro", f"Erro ao abrir editor:\n{e}")
    
    def _aplicar_imagem_editada(self, caminho_editado: str, tipo: str, letra_alt: str = None):
        """Aplica a imagem editada ao campo apropriado."""
        import os
        if tipo == 'enunciado':
            self.imagem_enunciado_path = caminho_editado
            nome_arquivo = os.path.basename(caminho_editado)
            self.lbl_imagem_enunciado.config(text=f"✅ {nome_arquivo[:30]}... (editada)")
            self.mostrar_preview_imagem(caminho_editado, self.lbl_preview_enunciado, 150)
        elif tipo == 'alternativa' and letra_alt:
            self.imagens_alternativas[letra_alt] = caminho_editado
            nome_arquivo = os.path.basename(caminho_editado)
            self.labels_imagem_alt[letra_alt].config(text=f"✅ {nome_arquivo[:15]}... (ed)")
            self.mostrar_preview_imagem(caminho_editado, self.labels_preview_alt[letra_alt], 40)
    
    def editar_imagem_enunciado(self):
        """Abre editor para editar imagem do enunciado."""
        if not self.imagem_enunciado_path:
            messagebox.showinfo("Aviso", "Selecione uma imagem primeiro.")
            return
        
        import os
        if not os.path.exists(self.imagem_enunciado_path):
            messagebox.showerror("Erro", "Arquivo de imagem não encontrado.")
            return
        
        self.abrir_editor_imagem(self.imagem_enunciado_path, tipo='enunciado')
    
    def editar_imagem_alternativa(self, letra: str):
        """Abre editor para editar imagem de uma alternativa."""
        caminho = self.imagens_alternativas.get(letra)
        if not caminho:
            messagebox.showinfo("Aviso", f"Selecione uma imagem para a alternativa {letra} primeiro.")
            return
        
        import os
        if not os.path.exists(caminho):
            messagebox.showerror("Erro", "Arquivo de imagem não encontrado.")
            return
        
        self.abrir_editor_imagem(caminho, tipo='alternativa', letra_alt=letra)
    
    def validar_tamanho_imagem(self, caminho: str, tamanho_max_mb: int = 5) -> bool:
        """
        Valida se a imagem não excede o tamanho máximo.
        
        Args:
            caminho: Caminho do arquivo
            tamanho_max_mb: Tamanho máximo em MB
            
        Returns:
            bool: True se válido, False caso contrário
        """
        import os
        try:
            tamanho_bytes = os.path.getsize(caminho)
            tamanho_mb = tamanho_bytes / (1024 * 1024)
            
            if tamanho_mb > tamanho_max_mb:
                resposta = messagebox.askyesno(
                    "Arquivo Grande",
                    f"A imagem tem {tamanho_mb:.1f}MB (máximo recomendado: {tamanho_max_mb}MB).\\n\\n"
                    "Deseja redimensionar automaticamente?",
                    icon='warning'
                )
                
                if resposta:
                    return self._redimensionar_automatico(caminho, tamanho_max_mb)
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"Erro ao validar tamanho: {e}")
            return False
    
    def _redimensionar_automatico(self, caminho: str, tamanho_max_mb: int) -> bool:
        """Redimensiona imagem automaticamente para não exceder tamanho máximo."""
        import os
        try:
            from PIL import Image
            
            img = Image.open(caminho)
            
            # Reduzir qualidade/tamanho progressivamente
            qualidade = 85
            largura_atual = img.width
            
            while True:
                # Salvar temporariamente
                temp_path = caminho + ".temp.jpg"
                
                if largura_atual < img.width:
                    ratio = largura_atual / img.width
                    nova_altura = int(img.height * ratio)
                    img_redim = img.resize((largura_atual, nova_altura), Image.Resampling.LANCZOS)
                else:
                    img_redim = img
                
                # Converter para RGB se necessário (para salvar como JPEG)
                if img_redim.mode in ('RGBA', 'LA', 'P'):
                    img_rgb = Image.new('RGB', img_redim.size, (255, 255, 255))
                    if img_redim.mode == 'P':
                        img_redim = img_redim.convert('RGBA')
                    img_rgb.paste(img_redim, mask=img_redim.split()[-1] if img_redim.mode == 'RGBA' else None)
                    img_redim = img_rgb
                
                img_redim.save(temp_path, "JPEG", quality=qualidade, optimize=True)
                
                # Verificar tamanho
                tamanho_mb = os.path.getsize(temp_path) / (1024 * 1024)
                
                if tamanho_mb <= tamanho_max_mb:
                    # Substituir original
                    os.replace(temp_path, caminho)
                    messagebox.showinfo(
                        "Redimensionamento",
                        f"Imagem redimensionada para {tamanho_mb:.1f}MB"
                    )
                    return True
                
                # Reduzir mais
                qualidade -= 10
                largura_atual = int(largura_atual * 0.8)
                
                if qualidade < 30 or largura_atual < 200:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    messagebox.showerror(
                        "Erro",
                        "Não foi possível reduzir o tamanho da imagem adequadamente."
                    )
                    return False
            
        except Exception as e:
            logger.error(f"Erro ao redimensionar: {e}")
            return False
    
    def ampliar_preview(self, caminho: str):
        """Abre janela com visualização ampliada da imagem."""
        import os
        if not caminho or not os.path.exists(caminho):
            messagebox.showwarning("Aviso", "Nenhuma imagem selecionada.")
            return
        
        try:
            janela_preview = tk.Toplevel(self.janela)
            janela_preview.title("📷 Visualização da Imagem")
            janela_preview.geometry("800x600")
            janela_preview.grab_set()
            
            from PIL import Image, ImageTk
            
            img = Image.open(caminho)
            
            # Redimensionar para caber na janela
            img_display = img.copy()
            img_display.thumbnail((780, 520), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img_display)
            
            label = tk.Label(janela_preview, image=photo)
            label.image = photo  # Manter referência
            label.pack(expand=True, pady=10)
            
            # Info da imagem
            tamanho_kb = os.path.getsize(caminho) / 1024
            info_text = f"Dimensões: {img.width}x{img.height}px | Tamanho: {tamanho_kb:.1f}KB"
            tk.Label(janela_preview, text=info_text, font=("Arial", 10)).pack(pady=5)
            
            tk.Button(
                janela_preview, text="Fechar",
                command=janela_preview.destroy,
                bg=self.co8, fg="white", padx=20
            ).pack(pady=10)
            
        except Exception as e:
            logger.error(f"Erro ao visualizar imagem: {e}")
            messagebox.showerror("Erro", f"Erro ao visualizar imagem:\\n{e}")
    
    # =========================================================================
    # FASE 2 - MELHORIAS IMPORTANTES: DRAG & DROP, CACHE, TEMPLATES
    # =========================================================================
    
    def habilitar_drag_drop(self):
        """Habilita arrastar e soltar imagens."""
        try:
            from tkinterdnd2 import DND_FILES
            
            # Enunciado
            self.frame_preview_enunciado.drop_target_register(DND_FILES)
            self.frame_preview_enunciado.dnd_bind(
                '<<Drop>>',
                lambda e: self._processar_drop(e.data, 'enunciado')
            )
            
            # Alternativas
            for letra in ["A", "B", "C", "D", "E"]:
                if hasattr(self, 'labels_preview_alt') and letra in self.labels_preview_alt:
                    self.labels_preview_alt[letra].drop_target_register(DND_FILES)
                    self.labels_preview_alt[letra].dnd_bind(
                        '<<Drop>>',
                        lambda e, l=letra: self._processar_drop(e.data, 'alternativa', l)
                    )
                    
        except ImportError:
            logger.info("tkinterdnd2 não disponível - drag & drop desabilitado")
        except Exception as e:
            logger.warning(f"Erro ao habilitar drag & drop: {e}")
    
    def _processar_drop(self, data: str, tipo: str, letra: str = None):
        """Processa arquivo arrastado."""
        import os
        # Limpar string do caminho
        caminho = data.strip('{}').strip()
        
        # Verificar se é imagem
        extensoes_validas = ['.png', '.jpg', '.jpeg', '.gif', '.bmp']
        if not any(caminho.lower().endswith(ext) for ext in extensoes_validas):
            messagebox.showwarning("Aviso", "Arraste apenas arquivos de imagem.")
            return
        
        # Validar tamanho
        if not self.validar_tamanho_imagem(caminho):
            return
        
        if tipo == 'enunciado':
            self.imagem_enunciado_path = caminho
            nome_arquivo = os.path.basename(caminho)
            self.lbl_imagem_enunciado.config(text=f"✅ {nome_arquivo[:30]}...")
            self.mostrar_preview_imagem(caminho, self.lbl_preview_enunciado, 150)
        elif tipo == 'alternativa' and letra:
            self.imagens_alternativas[letra] = caminho
            nome_arquivo = os.path.basename(caminho)
            self.labels_imagem_alt[letra].config(text=f"✅ {nome_arquivo[:15]}...")
            self.mostrar_preview_imagem(caminho, self.labels_preview_alt[letra], 40)
    
    def salvar_como_template(self):
        """Salva questão atual como template."""
        import os
        import json
        from tkinter import simpledialog
        
        nome = simpledialog.askstring(
            "Nome do Template",
            "Digite um nome para o template:",
            parent=self.janela
        )
        
        if not nome:
            return
        
        try:
            template = {
                'nome': nome,
                'componente': self.cad_componente.get() if hasattr(self, 'cad_componente') else '',
                'ano': self.cad_ano.get() if hasattr(self, 'cad_ano') else '',
                'tipo': self.cad_tipo.get() if hasattr(self, 'cad_tipo') else '',
                'dificuldade': self.cad_dificuldade.get() if hasattr(self, 'cad_dificuldade') else '',
                'alternativas': {}
            }
            
            # Salvar alternativas se existirem
            if hasattr(self, 'cad_alternativas'):
                for letra, entry in self.cad_alternativas.items():
                    template['alternativas'][letra] = entry.get()
            
            # Salvar em arquivo JSON
            templates_dir = os.path.join("config", "templates_questoes")
            os.makedirs(templates_dir, exist_ok=True)
            
            arquivo = os.path.join(templates_dir, f"{nome}.json")
            with open(arquivo, 'w', encoding='utf-8') as f:
                json.dump(template, f, indent=2, ensure_ascii=False)
            
            messagebox.showinfo("Sucesso", f"Template '{nome}' salvo!")
            
        except Exception as e:
            logger.error(f"Erro ao salvar template: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar template:\\n{e}")
    
    def carregar_template(self):
        """Carrega um template salvo."""
        import os
        
        templates_dir = os.path.join("config", "templates_questoes")
        
        if not os.path.exists(templates_dir):
            messagebox.showinfo("Aviso", "Nenhum template encontrado.")
            return
        
        templates = [f[:-5] for f in os.listdir(templates_dir) if f.endswith('.json')]
        
        if not templates:
            messagebox.showinfo("Aviso", "Nenhum template encontrado.")
            return
        
        # Diálogo de seleção
        janela = tk.Toplevel(self.janela)
        janela.title("Carregar Template")
        janela.geometry("400x300")
        janela.grab_set()
        
        tk.Label(janela, text="Selecione um template:").pack(pady=10)
        
        listbox = tk.Listbox(janela, height=10)
        listbox.pack(fill="both", expand=True, padx=10, pady=10)
        
        for template in templates:
            listbox.insert(tk.END, template)
        
        def aplicar():
            seleção = listbox.curselection()
            if not seleção:
                return
            
            nome_template = listbox.get(seleção[0])
            self._aplicar_template(nome_template)
            janela.destroy()
        
        tk.Button(
            janela, text="Carregar", command=aplicar,
            bg=self.co2, fg="white", padx=20
        ).pack(pady=10)
    
    def _aplicar_template(self, nome: str):
        """Aplica um template aos campos."""
        import os
        import json
        
        try:
            arquivo = os.path.join("config", "templates_questoes", f"{nome}.json")
            with open(arquivo, 'r', encoding='utf-8') as f:
                template = json.load(f)
            
            # Aplicar valores
            if hasattr(self, 'cad_componente') and template.get('componente'):
                self.cad_componente.set(template['componente'])
                self._atualizar_habilidades_cadastro()
            
            if hasattr(self, 'cad_ano') and template.get('ano'):
                self.cad_ano.set(template['ano'])
                self._atualizar_habilidades_cadastro()
            
            if hasattr(self, 'cad_tipo') and template.get('tipo'):
                self.cad_tipo.set(template['tipo'])
                self.atualizar_campos_tipo()
            
            if hasattr(self, 'cad_dificuldade') and template.get('dificuldade'):
                self.cad_dificuldade.set(template['dificuldade'])
            
            # Alternativas
            if hasattr(self, 'cad_alternativas') and template.get('alternativas'):
                for letra, texto in template['alternativas'].items():
                    if letra in self.cad_alternativas:
                        self.cad_alternativas[letra].delete(0, tk.END)
                        self.cad_alternativas[letra].insert(0, texto)
            
            messagebox.showinfo("Sucesso", f"Template '{nome}' aplicado!")
            
        except Exception as e:
            logger.error(f"Erro ao carregar template: {e}")
            messagebox.showerror("Erro", f"Erro ao carregar template:\\n{e}")
    
    # =========================================================================
    # IMPORTAÇÃO EM LOTE - EXCEL/CSV
    # =========================================================================
    
    def importar_questoes_excel(self):
        """Importa questões de um arquivo Excel com validação robusta."""
        from tkinter import filedialog
        import os
        
        arquivo = filedialog.askopenfilename(
            title="Selecionar arquivo Excel",
            filetypes=[
                ("Arquivos Excel", "*.xlsx *.xls"),
                ("Todos os arquivos", "*.*")
            ]
        )
        
        if not arquivo:
            return
        
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror(
                "Erro",
                "Biblioteca openpyxl não instalada.\\n\\n"
                "Execute: pip install openpyxl"
            )
            return
        
        try:
            from banco_questoes.services import QuestaoService
            from banco_questoes.models import (
                Questao, QuestaoAlternativa, ComponenteCurricular, AnoEscolar,
                TipoQuestao, DificuldadeQuestao, StatusQuestao, VisibilidadeQuestao
            )
            
            wb = openpyxl.load_workbook(arquivo)
            ws = wb.active
            
            # Validar cabeçalho
            headers_esperados = [
                'componente', 'ano', 'habilidade_bncc', 'tipo', 'dificuldade',
                'enunciado', 'alt_a', 'alt_b', 'alt_c', 'alt_d', 'alt_e', 'gabarito'
            ]
            
            headers = [cell.value.lower() if cell.value else '' for cell in ws[1]]
            
            if headers[:6] != headers_esperados[:6]:
                messagebox.showerror(
                    "Erro",
                    "Formato de arquivo inválido.\\n\\n"
                    "Cabeçalhos esperados:\\n" +
                    ", ".join(headers_esperados)
                )
                return
            
            # Processar linhas
            questoes_importadas = 0
            erros = []
            
            janela_progresso = tk.Toplevel(self.janela)
            janela_progresso.title("Importando Questões")
            janela_progresso.geometry("500x200")
            janela_progresso.grab_set()
            
            tk.Label(janela_progresso, text="Importando questões...", font=("Arial", 12)).pack(pady=10)
            texto_log = tk.Text(janela_progresso, height=8, wrap="word")
            texto_log.pack(fill="both", expand=True, padx=10, pady=10)
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Validar campos obrigatórios
                    if not all([row[0], row[1], row[2], row[3], row[4], row[5]]):
                        erros.append(f"Linha {idx}: Campos obrigatórios vazios")
                        continue
                    
                    # Criar objeto Questão
                    questao = Questao(
                        componente_curricular=ComponenteCurricular(row[0]),
                        ano_escolar=AnoEscolar(row[1]),
                        habilidade_bncc_codigo=row[2],
                        tipo=TipoQuestao(row[3]),
                        dificuldade=DificuldadeQuestao(row[4]),
                        enunciado=row[5],
                        status=StatusQuestao.RASCUNHO,
                        visibilidade=VisibilidadeQuestao.ESCOLA,
                        autor_id=self.funcionario_id,
                        escola_id=config.ESCOLA_ID
                    )
                    
                    # Alternativas se for múltipla escolha
                    if row[3] == 'multipla_escolha':
                        alternativas = []
                        for i, letra in enumerate(['A', 'B', 'C', 'D', 'E']):
                            if row[6+i]:  # alt_a, alt_b, etc
                                alt = QuestaoAlternativa(
                                    letra=letra,
                                    texto=row[6+i],
                                    correta=(letra == row[11]),  # gabarito
                                    ordem=i+1
                                )
                                alternativas.append(alt)
                        
                        if len(alternativas) < 2:
                            erros.append(f"Linha {idx}: Mínimo 2 alternativas necessárias")
                            continue
                        
                        questao.alternativas = alternativas
                        questao.gabarito_letra = row[11]
                    
                    # Salvar no banco
                    questao_id = QuestaoService.criar(questao)
                    
                    if questao_id:
                        questoes_importadas += 1
                        texto_log.insert(tk.END, f"✓ Linha {idx}: Questão #{questao_id} importada\\n")
                        texto_log.see(tk.END)
                        janela_progresso.update()
                    else:
                        erros.append(f"Linha {idx}: Falha ao salvar no banco")
                    
                except Exception as e:
                    erros.append(f"Linha {idx}: {str(e)}")
            
            # Relatório final
            janela_progresso.destroy()
            
            mensagem = f"✅ Importação concluída!\\n\\n"
            mensagem += f"Questões importadas: {questoes_importadas}\\n"
            
            if erros:
                mensagem += f"Erros: {len(erros)}\\n\\n"
                mensagem += "Primeiros erros:\\n"
                mensagem += "\\n".join(erros[:5])
                if len(erros) > 5:
                    mensagem += f"\\n... e mais {len(erros)-5} erros"
                
                # Salvar log completo
                from datetime import datetime
                log_file = os.path.join("logs", f"importacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
                os.makedirs("logs", exist_ok=True)
                with open(log_file, 'w', encoding='utf-8') as f:
                    f.write("\\n".join(erros))
                mensagem += f"\\n\\nLog completo salvo em: {log_file}"
            
            messagebox.showinfo("Importação Concluída", mensagem)
            
        except Exception as e:
            logger.error(f"Erro ao importar questões: {e}")
            messagebox.showerror("Erro", f"Erro ao importar questões:\\n{e}")


def abrir_banco_questoes(janela_principal=None):
    """
    Função auxiliar para abrir o banco de questões.
    
    Args:
        janela_principal: Referência à janela principal (opcional)
    """
    try:
        # Esconder a janela principal para focar na interface do banco de questões
        if janela_principal:
            try:
                janela_principal.withdraw()
            except Exception:
                # Ignorar se não for possível ocultar
                pass

        interface = InterfaceBancoQuestoes(janela_principal=janela_principal)
        return interface
    except Exception as e:
        logger.error(f"Erro ao abrir banco de questões: {e}")
        messagebox.showerror("Erro", f"Não foi possível abrir o banco de questões: {e}")
        # Restaurar janela principal caso tenha sido escondida
        if janela_principal:
            try:
                janela_principal.deiconify()
            except Exception:
                pass
        return None
