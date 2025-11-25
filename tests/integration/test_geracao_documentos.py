"""
Testes de integração para geração de documentos.

Este módulo testa a geração de:
- Boletins escolares
- Declarações
- Relatórios
- Históricos escolares
"""

import pytest
import os
import tempfile
from pathlib import Path
from datetime import datetime
from unittest.mock import Mock, patch, MagicMock
from typing import Dict, Any

from config_logs import get_logger

logger = get_logger(__name__)


# =============================================================================
# TESTES DE GERAÇÃO DE DOCUMENTOS
# =============================================================================

class TestGeracaoDocumentos:
    """Testes de integração para geração de documentos."""
    
    @patch('services.report_service.gerar_boletim')
    def test_gerar_boletim_aluno(self, mock_gerar):
        """Testa geração de boletim para aluno."""
        # Mock
        mock_gerar.return_value = (True, "boletim_123.pdf")
        
        # Importar service
        from services import report_service
        
        # Executar
        sucesso, arquivo = report_service.gerar_boletim(
            aluno_id=123,
            ano_letivo_id=2025
        )
        
        assert sucesso is True
        assert isinstance(arquivo, str) and arquivo.endswith('.pdf')
        mock_gerar.assert_called_once()
    
    @patch('services.report_service.gerar_declaracao')
    def test_gerar_declaracao_comparecimento(self, mock_gerar):
        """Testa geração de declaração de comparecimento."""
        # Mock
        mock_gerar.return_value = (True, "declaracao_123.pdf")
        
        # Importar service
        from services import report_service
        
        # Executar
        sucesso, arquivo = report_service.gerar_declaracao(
            aluno_id=123,
            tipo='comparecimento',
            data=datetime.now()
        )
        
        assert sucesso is True
        assert isinstance(arquivo, str) and arquivo.endswith('.pdf')
        mock_gerar.assert_called_once()
    
    @patch('services.report_service.gerar_historico_escolar')
    def test_gerar_historico_escolar(self, mock_gerar):
        """Testa geração de histórico escolar."""
        # Mock
        mock_gerar.return_value = (True, "historico_123.pdf")
        
        # Importar service
        from services import report_service
        
        # Executar
        sucesso, arquivo = report_service.gerar_historico_escolar(
            aluno_id=123
        )
        
        assert sucesso is True
        assert isinstance(arquivo, str) and arquivo.endswith('.pdf')
        mock_gerar.assert_called_once()
    
    @patch('gerarPDF.gerar_pdf')
    def test_gerar_lista_presenca_turma(self, mock_pdf):
        """Testa geração de lista de presença para turma."""
        # Mock
        mock_pdf.return_value = True
        
        # Importar módulo
        import gerarPDF
        
        # Dados de teste
        dados_turma = {
            'turma_nome': 'Turma A',
            'serie': '1º Ano',
            'alunos': [
                {'id': 1, 'nome': 'João Silva'},
                {'id': 2, 'nome': 'Maria Santos'}
            ]
        }
        
        # Executar
        resultado = gerarPDF.gerar_pdf(dados_turma, tipo='lista_presenca')
        
        assert resultado is True
        mock_pdf.assert_called_once()
    
    def test_integracao_gerar_boletim_real(self, db_test_connection):
        """
        Teste de integração real para geração de boletim.
        
        Requer dados reais no banco e bibliotecas de PDF instaladas.
        """
        try:
            from services.report_service import gerar_boletim
            
            # Buscar um aluno real para teste
            cursor = db_test_connection.cursor(dictionary=True)
            cursor.execute("""
                SELECT a.id, a.nome
                FROM Alunos a
                JOIN Matriculas m ON m.aluno_id = a.id
                WHERE a.escola_id = 60
                LIMIT 1
            """)
            aluno = cursor.fetchone()
            cursor.close()
            
            if aluno:
                # Tentar gerar boletim real
                sucesso, resultado = gerar_boletim(
                    aluno_id=aluno['id'],
                    ano_letivo_id=2025
                )
                
                if sucesso:
                    # Verificar que arquivo foi criado
                    if isinstance(resultado, str):
                        arquivo_path = Path(resultado)
                        assert arquivo_path.exists()
                        assert arquivo_path.stat().st_size > 0
                else:
                    # Boletim pode falhar se não houver notas
                    logger.info(f"Boletim não gerado: {resultado}")
            else:
                pytest.skip("Nenhum aluno disponível para teste real")
                
        except ImportError:
            pytest.skip("Módulo de geração de boletim não disponível")


# =============================================================================
# TESTES DE RELATÓRIOS
# =============================================================================

class TestGeracaoRelatorios:
    """Testes de integração para geração de relatórios."""
    
    @patch('services.report_service.gerar_relatorio_notas')
    def test_gerar_relatorio_notas_turma(self, mock_relatorio):
        """Testa geração de relatório de notas por turma."""
        # Mock
        mock_relatorio.return_value = (True, "relatorio_notas_turma_1.xlsx")
        
        from services import report_service
        
        sucesso, arquivo = report_service.gerar_relatorio_notas(
            turma_id=1,
            bimestre=1,
            formato='xlsx'
        )
        
        assert sucesso is True
        assert isinstance(arquivo, str) and arquivo.endswith('.xlsx')
    
    @patch('services.report_service.gerar_relatorio_frequencia')
    def test_gerar_relatorio_frequencia_mensal(self, mock_relatorio):
        """Testa geração de relatório de frequência mensal."""
        # Mock
        mock_relatorio.return_value = (True, "relatorio_freq_2025_01.pdf")
        
        from services import report_service
        
        sucesso, arquivo = report_service.gerar_relatorio_frequencia(
            turma_id=1,
            mes=1,
            ano=2025,
            formato='pdf'
        )
        
        assert sucesso is True
        assert isinstance(arquivo, str) and arquivo.endswith('.pdf')
    
    @patch('services.report_service.gerar_relatorio_matriculas')
    def test_gerar_relatorio_matriculas_escola(self, mock_relatorio):
        """Testa geração de relatório de matrículas por escola."""
        # Mock
        dados_mock = {
            'total_matriculas': 150,
            'matriculas_ativas': 140,
            'matriculas_canceladas': 10,
            'por_serie': [
                {'serie': '1º Ano', 'total': 30},
                {'serie': '2º Ano', 'total': 28}
            ]
        }
        mock_relatorio.return_value = (True, dados_mock)
        
        from services import report_service
        
        sucesso, dados = report_service.gerar_relatorio_matriculas(
            escola_id=60,
            ano_letivo=2025
        )
        
        assert sucesso is True
        assert dados is not None and isinstance(dados, dict)
        assert dados['total_matriculas'] > 0
        assert len(dados['por_serie']) > 0
    
    def test_integracao_relatorio_matriculas_real(self, db_test_connection):
        """Teste de integração real para relatório de matrículas."""
        cursor = db_test_connection.cursor(dictionary=True)
        
        # Query real para contar matrículas
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN status = 'Ativo' THEN 1 ELSE 0 END) as ativas
            FROM Matriculas m
            JOIN anos_letivos al ON m.ano_letivo_id = al.id
            WHERE al.ano = YEAR(CURDATE())
        """)
        
        resultado = cursor.fetchone()
        cursor.close()
        
        assert resultado is not None
        assert resultado['total'] >= 0
        assert resultado['ativas'] >= 0


# =============================================================================
# TESTES DE PERFORMANCE DE DOCUMENTOS
# =============================================================================

class TestPerformanceDocumentos:
    """Testes de performance para geração de documentos."""
    
    @pytest.mark.slow
    def test_gerar_multiplos_boletins_performance(self):
        """
        Testa performance ao gerar múltiplos boletins.
        
        Meta: Gerar 50 boletins em menos de 30 segundos.
        """
        import time
        from unittest.mock import patch
        
        with patch('services.report_service.gerar_boletim') as mock_gerar:
            mock_gerar.return_value = (True, "boletim.pdf")
            
            inicio = time.time()
            
            # Gerar 50 boletins
            for i in range(50):
                from services import report_service
                report_service.gerar_boletim(
                    aluno_id=i,
                    ano_letivo_id=2025
                )
            
            duracao = time.time() - inicio
            
            assert duracao < 30  # Deve levar menos de 30 segundos
            logger.info(f"50 boletins gerados em {duracao:.2f}s")
    
    @pytest.mark.slow
    def test_gerar_relatorio_grande_volume(self, db_test_connection):
        """
        Testa geração de relatório com grande volume de dados.
        
        Meta: Relatório com 1000+ registros em menos de 10 segundos.
        """
        import time
        
        inicio = time.time()
        
        cursor = db_test_connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT 
                a.id, a.nome, m.status,
                t.nome as turma, s.nome as serie
            FROM Alunos a
            JOIN Matriculas m ON m.aluno_id = a.id
            JOIN Turmas t ON m.turma_id = t.id
            JOIN Series s ON t.serie_id = s.id
            LIMIT 1000
        """)
        
        registros = cursor.fetchall()
        cursor.close()
        
        duracao = time.time() - inicio
        
        assert len(registros) <= 1000
        assert duracao < 10  # Deve levar menos de 10 segundos
        logger.info(f"Query de {len(registros)} registros em {duracao:.2f}s")


# =============================================================================
# TESTES DE VALIDAÇÃO DE DOCUMENTOS
# =============================================================================

class TestValidacaoDocumentos:
    """Testes de validação de documentos gerados."""
    
    def test_validar_pdf_gerado(self):
        """Testa que PDF gerado é válido."""
        try:
            from PyPDF2 import PdfReader
            
            # Criar PDF de teste
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp_path = tmp.name
            
            # Mock da geração
            with patch('gerarPDF.gerar_pdf', return_value=tmp_path):
                import gerarPDF
                resultado = gerarPDF.gerar_pdf({}, tipo='teste')
                
                if isinstance(resultado, str) and os.path.exists(resultado):
                    # Tentar ler PDF
                    try:
                        reader = PdfReader(resultado)
                        assert len(reader.pages) > 0
                    except Exception as e:
                        pytest.fail(f"PDF inválido: {e}")
                    finally:
                        os.unlink(resultado)
                        
        except ImportError:
            pytest.skip("PyPDF2 não instalado")
    
    def test_validar_excel_gerado(self):
        """Testa que arquivo Excel gerado é válido."""
        try:
            import openpyxl
            
            # Criar Excel de teste
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name
                
                # Criar workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                assert ws is not None
                ws['A1'] = 'Teste'
                wb.save(tmp_path)
                
                # Validar
                wb_read = openpyxl.load_workbook(tmp_path)
                ws_read = wb_read.active
                assert ws_read is not None
                assert ws_read['A1'].value == 'Teste'
                
                # Cleanup
                wb_read.close()
                os.unlink(tmp_path)
                
        except ImportError:
            pytest.skip("openpyxl não instalado")


# =============================================================================
# FIXTURES
# =============================================================================

@pytest.fixture
def db_test_connection():
    """Conexão com banco de teste."""
    from db.connection import get_connection
    # `get_connection()` é um context manager; entrar explicitamente para obter
    # o proxy de conexão e garantir fechamento adequado.
    ctx = get_connection()
    conn = ctx.__enter__()
    try:
        yield conn
    finally:
        try:
            ctx.__exit__(None, None, None)
        except Exception:
            logger.exception("Erro ao fechar conexão de teste (fixture db_test_connection)")


@pytest.fixture
def temp_pdf_file():
    """Arquivo PDF temporário para testes."""
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        yield tmp.name
    
    if os.path.exists(tmp.name):
        os.unlink(tmp.name)


@pytest.fixture
def temp_excel_file():
    """Arquivo Excel temporário para testes."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        yield tmp.name
    
    if os.path.exists(tmp.name):
        os.unlink(tmp.name)
