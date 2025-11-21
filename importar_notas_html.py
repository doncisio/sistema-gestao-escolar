from config_logs import get_logger
logger = get_logger(__name__)
"""
Script para extrair notas do HTML exportado do GEDUC e preencher template Excel
"""

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re
import os
from tkinter import messagebox, filedialog
import tkinter as tk

def extrair_informacoes_html(html_path):
    """
    Extrai informações da página HTML do GEDUC
    Retorna: turma, disciplina, bimestre, lista de alunos com notas
    """
    try:
        with open(html_path, 'r', encoding='utf-8') as file:
            html_content = file.read()
        
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Extrair turma selecionada
        turma_select = soup.find('select', {'name': 'IDTURMA'})
        turma = "Desconhecida"
        if turma_select:
            turma_option = turma_select.find('option', {'selected': '1'})
            if turma_option:
                turma = turma_option.text.strip()
        
        # Extrair disciplina selecionada
        disciplina_select = soup.find('select', {'name': 'IDTURMASDISP'})
        disciplina = "Desconhecida"
        if disciplina_select:
            # Primeira option é geralmente a selecionada
            disciplina_option = disciplina_select.find('option')
            if disciplina_option:
                disciplina = disciplina_option.text.strip()
        
        # Extrair bimestre/período selecionado
        bimestre_radios = soup.find_all('input', {'name': 'IDAVALIACOES'})
        bimestre = "1º"
        for radio in bimestre_radios:
            if radio.get('checked'):
                label = soup.find('label', {'for': radio.get('id')})
                if label:
                    texto_bimestre = label.text.strip()
                    # Extrair número do período (1º, 2º, 3º, 4º)
                    match = re.search(r'(\d+)º', texto_bimestre)
                    if match:
                        bimestre = f"{match.group(1)}º"
        
        # Extrair alunos e notas da tabela
        alunos_notas = []
        tbody = soup.find('tbody', {'class': 'tdatagrid_body'})
        
        if tbody:
            rows = tbody.find_all('tr', {'class': ['tdatagrid_row_odd', 'tdatagrid_row_even']})
            
            for row in rows:
                cells = row.find_all('td', {'class': 'tdatagrid_cell'})
                
                if len(cells) >= 2:
                    # Primeira célula: ordem
                    ordem_text = cells[0].text.strip()
                    
                    # Segunda célula: nome do aluno
                    nome_aluno = cells[1].text.strip()
                    
                    # Se a ordem não for vazia e o nome não for vazio (não é o cabeçalho)
                    if ordem_text and nome_aluno and ordem_text.isdigit():
                        # Terceira célula contém as notas em inputs
                        if len(cells) >= 3:
                            nota_inputs = cells[2].find_all('input', {'class': 'tfield'})
                            notas = []
                            
                            for input_nota in nota_inputs:
                                valor_nota = input_nota.get('value', '')
                                if valor_nota:
                                    try:
                                        # Converte para float
                                        nota_float = float(valor_nota)
                                        notas.append(nota_float)
                                    except ValueError:
                                        notas.append(0.0)
                                else:
                                    notas.append(0.0)
                            
                            # Calcula média das notas disponíveis e multiplica por 10
                            notas_validas = [n for n in notas if isinstance(n, float) and n > 0]
                            if notas_validas:
                                media = sum(notas_validas) / len(notas_validas)
                                nota_final: float = media * 10  # Multiplica por 10
                            else:
                                nota_final = 0.0
                            
                            alunos_notas.append({
                                'ordem': int(ordem_text),
                                'nome': nome_aluno,
                                'nota': nota_final  # Apenas a nota final
                            })
        
        return {
            'turma': turma,
            'disciplina': disciplina,
            'bimestre': bimestre,
            'alunos': alunos_notas
        }
    
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao extrair informações do HTML: {e}")
        return None


def criar_excel_com_notas(dados, output_path):
    """
    Cria arquivo Excel com as notas extraídas do HTML
    Formato: Template_Notas__-_MAT_{disciplina}_{bimestre}_bimestre.xlsx
    """
    try:
        # Criar novo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notas"
        
        # Definir estilos
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment_left = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalho - apenas 3 colunas
        ws['A1'] = 'Nº'
        ws['B1'] = 'Nome do Aluno'
        ws['C1'] = 'Nota'
        
        # Aplicar estilo ao cabeçalho
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 12
        
        # Preencher dados dos alunos
        row = 2
        for aluno in dados['alunos']:
            ws[f'A{row}'] = aluno['ordem']
            ws[f'B{row}'] = aluno['nome']
            
            # Preencher nota final (média * 10)
            if aluno['nota'] != '':
                ws[f'C{row}'] = round(aluno['nota'], 2)
            
            # Aplicar estilo às células
            ws[f'A{row}'].alignment = cell_alignment
            ws[f'B{row}'].alignment = cell_alignment_left
            ws[f'C{row}'].alignment = cell_alignment
            
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.font = cell_font
                cell.border = thin_border
            
            row += 1
        
        # Adicionar informações no topo
        ws.insert_rows(1, 3)
        ws['A1'] = f"Turma: {dados['turma']}"
        ws['A2'] = f"Disciplina: {dados['disciplina']}"
        ws['A3'] = f"Bimestre: {dados['bimestre']}"
        
        ws['A1'].font = Font(name='Arial', size=12, bold=True)
        ws['A2'].font = Font(name='Arial', size=12, bold=True)
        ws['A3'].font = Font(name='Arial', size=12, bold=True)
        
        # Salvar arquivo
        wb.save(output_path)
        return True
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao criar arquivo Excel: {e}")
        return False


def preencher_template_existente(dados, template_path, output_path):
    """
    Preenche um template Excel existente com as notas extraídas
    """
    try:
        # Carregar template existente
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Encontrar a linha inicial dos dados (após o cabeçalho)
        # Procura pela coluna "Nome do Aluno" ou similar
        linha_inicio = 2
        for row in range(1, 10):
            cell_value = ws.cell(row, 2).value
            if cell_value and ('Nome' in str(cell_value) or 'ALUNO' in str(cell_value).upper()):
                linha_inicio = row + 1
                break
        
        # Preencher notas
        for i, aluno in enumerate(dados['alunos']):
            row = linha_inicio + i
            
            # Preencher número de ordem
            ws.cell(row, 1, aluno['ordem'])
            
            # Preencher nome
            ws.cell(row, 2, aluno['nome'])
            
            # Preencher nota final (média * 10) na coluna 3
            if aluno['nota'] != '':
                ws.cell(row, 3, round(aluno['nota'], 2))
        
        # Salvar arquivo
        wb.save(output_path)
        return True
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao preencher template: {e}")
        return False


def interface_importacao(janela_pai=None):
    """
    Interface gráfica para importação de notas do HTML para Excel
    janela_pai: referência à janela principal (opcional)
    """
    # Criar janela
    if janela_pai:
        root = tk.Toplevel(janela_pai)
    else:
        root = tk.Tk()
    
    root.title("Importar Notas do HTML GEDUC")
    root.geometry("650x550")
    root.configure(bg="#F5F5F5")
    
    # Centralizar janela na tela
    root.update_idletasks()
    width = 650
    height = 550
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    # Configurar comportamento de fechamento
    def ao_fechar():
        root.destroy()
        if janela_pai:
            janela_pai.deiconify()
    
    root.protocol("WM_DELETE_WINDOW", ao_fechar)
    
    # Cores
    co0 = "#F5F5F5"
    co1 = "#003A70"
    co2 = "#77B341"
    co4 = "#4A86E8"
    
    # Variáveis
    html_path_var = tk.StringVar()
    template_path_var = tk.StringVar()
    usar_template_var = tk.BooleanVar(value=False)
    
    # Frame principal
    frame_principal = tk.Frame(root, bg=co0)
    frame_principal.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
    
    # Título
    titulo = tk.Label(
        frame_principal,
        text="Importar Notas do GEDUC",
        font=("Arial", 16, "bold"),
        bg=co0,
        fg=co1
    )
    titulo.pack(pady=(0, 15))
    
    # Seleção do arquivo HTML
    frame_html = tk.Frame(frame_principal, bg=co0)
    frame_html.pack(fill=tk.X, pady=8)
    
    tk.Label(
        frame_html,
        text="Arquivo HTML:",
        font=("Arial", 11, "bold"),
        bg=co0,
        fg=co1
    ).pack(anchor="w")
    
    frame_html_input = tk.Frame(frame_html, bg=co0)
    frame_html_input.pack(fill=tk.X, pady=5)
    
    entry_html = tk.Entry(frame_html_input, textvariable=html_path_var, font=("Arial", 10), width=50)
    entry_html.pack(side=tk.LEFT, fill=tk.X, expand=True)
    
    def selecionar_html():
        filename = filedialog.askopenfilename(
            title="Selecione o arquivo HTML",
            filetypes=[("Arquivos HTML", "*.html *.htm"), ("Todos os arquivos", "*.*")],
            parent=root
        )
        if filename:
            html_path_var.set(filename)
            logger.info(f"Arquivo selecionado: {filename}")
    
    btn_html = tk.Button(
        frame_html_input,
        text="Procurar",
        command=selecionar_html,
        bg=co4,
        fg="white",
        font=("Arial", 10, "bold"),
        relief=tk.RAISED,
        cursor="hand2"
    )
    btn_html.pack(side=tk.LEFT, padx=(5, 0))
    
    # Opção de usar template
    check_template = tk.Checkbutton(
        frame_principal,
        text="Usar template Excel existente",
        variable=usar_template_var,
        font=("Arial", 10),
        bg=co0,
        fg=co1
    )
    check_template.pack(anchor="w", pady=8)
    
    # Seleção do template (condicional)
    frame_template = tk.Frame(frame_principal, bg=co0)
    frame_template.pack(fill=tk.X, pady=8)
    
    tk.Label(
        frame_template,
        text="Template Excel (opcional):",
        font=("Arial", 11, "bold"),
        bg=co0,
        fg=co1
    ).pack(anchor="w")
    
    frame_template_input = tk.Frame(frame_template, bg=co0)
    frame_template_input.pack(fill=tk.X, pady=5)
    
    entry_template = tk.Entry(frame_template_input, textvariable=template_path_var, font=("Arial", 10), width=50)
    entry_template.pack(side=tk.LEFT, fill=tk.X, expand=True)
    
    def selecionar_template():
        filename = filedialog.askopenfilename(
            title="Selecione o template Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")],
            parent=root
        )
        if filename:
            template_path_var.set(filename)
            logger.info(f"Template selecionado: {filename}")
    
    btn_template = tk.Button(
        frame_template_input,
        text="Procurar",
        command=selecionar_template,
        bg=co4,
        fg="white",
        font=("Arial", 10, "bold"),
        relief=tk.RAISED,
        cursor="hand2"
    )
    btn_template.pack(side=tk.LEFT, padx=(5, 0))
    
    # Informações
    info_label = tk.Label(
        frame_principal,
        text="O arquivo será salvo com o nome:\nTemplate_Notas__-_MAT_{disciplina}_{bimestre}_bimestre.xlsx",
        font=("Arial", 9),
        bg=co0,
        fg="#666666",
        justify=tk.LEFT
    )
    info_label.pack(pady=15)
    
    # Botão processar
    def processar():
        html_path = html_path_var.get()
        
        if not html_path:
            messagebox.showerror("Erro", "Selecione um arquivo HTML!", parent=root)
            return
        
        if not os.path.exists(html_path):
            messagebox.showerror("Erro", f"Arquivo HTML não encontrado!\n\n{html_path}", parent=root)
            return
        
        # Extrair dados do HTML
        dados = extrair_informacoes_html(html_path)
        
        if not dados or not dados['alunos']:
            messagebox.showerror("Erro", "Não foi possível extrair dados do HTML ou nenhum aluno encontrado!")
            return
        
        # Gerar nome do arquivo de saída baseado no nome do HTML
        html_filename = os.path.splitext(os.path.basename(html_path))[0]  # Remove extensão .html
        output_filename = f"{html_filename}.xlsx"
        
        # Diretório de saída
        output_dir = os.path.dirname(html_path)
        output_path = os.path.join(output_dir, output_filename)
        
        # Verificar se deve usar template
        usar_template = usar_template_var.get()
        template_path = template_path_var.get()
        
        sucesso = False
        if usar_template and template_path and os.path.exists(template_path):
            sucesso = preencher_template_existente(dados, template_path, output_path)
        else:
            sucesso = criar_excel_com_notas(dados, output_path)
        
        if sucesso:
            messagebox.showinfo(
                "Sucesso",
                f"Arquivo Excel criado com sucesso!\n\n"
                f"Turma: {dados['turma']}\n"
                f"Disciplina: {dados['disciplina']}\n"
                f"Bimestre: {dados['bimestre']}\n"
                f"Total de alunos: {len(dados['alunos'])}\n\n"
                f"Arquivo salvo em:\n{output_path}",
                parent=root
            )
            # Limpar campos para nova importação
            html_path_var.set("")
            template_path_var.set("")
            usar_template_var.set(False)
    
    btn_processar = tk.Button(
        frame_principal,
        text="PROCESSAR E GERAR EXCEL",
        command=processar,
        bg=co2,
        fg="white",
        font=("Arial", 12, "bold"),
        relief=tk.RAISED,
        cursor="hand2",
        height=2,
        width=30
    )
    btn_processar.pack(pady=15, side=tk.BOTTOM)
    
    root.mainloop()


if __name__ == "__main__":
    interface_importacao()