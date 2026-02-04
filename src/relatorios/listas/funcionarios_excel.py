"""
Módulo para exportação de dados de funcionários administrativos para Excel.
"""

import os
from datetime import datetime
from typing import Optional, List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from tkinter import messagebox

from src.core.config_logs import get_logger
from db.connection import get_cursor

logger = get_logger(__name__)


def buscar_funcionarios_administrativos(escola_id: int = 60) -> List[Dict[str, Any]]:
    """
    Busca os funcionários administrativos da escola (exclui professores e tutores).
    
    Args:
        escola_id: ID da escola (padrão: 60)
        
    Returns:
        Lista de dicionários com dados dos funcionários
    """
    query = """
    SELECT 
        f.nome,
        f.cpf,
        f.data_nascimento,
        f.email,
        f.telefone,
        f.endereco_cep AS cep,
        CONCAT_WS(', ',
            NULLIF(f.endereco_logradouro, ''),
            NULLIF(f.endereco_numero, ''),
            NULLIF(f.endereco_complemento, ''),
            NULLIF(f.endereco_bairro, ''),
            NULLIF(f.endereco_cidade, ''),
            NULLIF(f.endereco_estado, '')
        ) AS endereco_completo,
        f.cargo,
        f.funcao,
        f.vinculo
    FROM 
        funcionarios f
    WHERE 
        f.escola_id = %s
        AND f.cargo NOT IN ('Professor@', 'Tutor/Cuidador')
    ORDER BY 
        f.nome
    """
    
    try:
        with get_cursor() as cursor:
            cursor.execute(query, (escola_id,))
            resultados = cursor.fetchall()
            
            logger.info(f"Encontrados {len(resultados)} funcionários administrativos")
            return resultados
            
    except Exception as e:
        logger.exception(f"Erro ao buscar funcionários administrativos: {e}")
        return []


def formatar_data(data: Any) -> str:
    """Formata data para exibição."""
    if data is None:
        return ""
    if isinstance(data, str):
        return data
    return data.strftime("%d/%m/%Y")


def formatar_cpf(cpf: Optional[str]) -> str:
    """Formata CPF para exibição."""
    if not cpf:
        return ""
    # Remove caracteres não numéricos
    cpf_numeros = ''.join(filter(str.isdigit, cpf))
    if len(cpf_numeros) == 11:
        return f"{cpf_numeros[:3]}.{cpf_numeros[3:6]}.{cpf_numeros[6:9]}-{cpf_numeros[9:]}"
    return cpf


def aplicar_estilo_cabecalho(ws, row: int = 1):
    """Aplica estilo ao cabeçalho da planilha."""
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 11):  # 10 colunas
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border


def aplicar_estilo_dados(ws, max_row: int, max_col: int):
    """Aplica estilo às células de dados."""
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            
            # Alinhamento específico por coluna
            if col in [1, 7, 8, 9]:  # NOME, ENDERECO, FUNÇÃO, REGIME
                cell.alignment = alignment_left
            else:  # Demais colunas centralizadas
                cell.alignment = alignment_center


def ajustar_largura_colunas(ws):
    """Ajusta a largura das colunas."""
    larguras = {
        'A': 35,  # NOME
        'B': 15,  # CPF
        'C': 18,  # DATA DE NASCIMENTO
        'D': 30,  # EMAIL
        'E': 16,  # TELEFONE
        'F': 25,  # NATURALIDADE (vazio por enquanto)
        'G': 12,  # CEP
        'H': 45,  # ENDEREÇO
        'I': 30,  # FUNÇÃO
        'J': 15,  # REGIME
    }
    
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura


def gerar_excel_funcionarios_administrativos(
    escola_id: int = 60,
    caminho_saida: Optional[str] = None
) -> Optional[str]:
    """
    Gera arquivo Excel com dados dos funcionários administrativos.
    
    Args:
        escola_id: ID da escola (padrão: 60)
        caminho_saida: Caminho para salvar o arquivo (opcional)
        
    Returns:
        Caminho do arquivo gerado ou None em caso de erro
    """
    try:
        # Buscar dados
        funcionarios = buscar_funcionarios_administrativos(escola_id)
        
        if not funcionarios:
            messagebox.showwarning(
                "Aviso",
                "Nenhum funcionário administrativo encontrado."
            )
            return None
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Funcionários Administrativos"
        
        # Definir cabeçalhos
        headers = [
            "NOME",
            "CPF",
            "DATA DE NASCIMENTO",
            "EMAIL",
            "TELEFONE",
            "NATURALIDADE",
            "CEP",
            "ENDEREÇO",
            "FUNÇÃO",
            "REGIME"
        ]
        
        # Adicionar cabeçalhos
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Aplicar estilo ao cabeçalho
        aplicar_estilo_cabecalho(ws)
        
        # Adicionar dados
        for idx, func in enumerate(funcionarios, start=2):
            nome = func.get('nome', '')
            ws.cell(row=idx, column=1, value=nome.upper() if nome else '')
            ws.cell(row=idx, column=2, value=formatar_cpf(func.get('cpf')))
            ws.cell(row=idx, column=3, value=formatar_data(func.get('data_nascimento')))
            
            email = func.get('email', '')
            ws.cell(row=idx, column=4, value=email.upper() if email else '')
            
            telefone = func.get('telefone', '')
            ws.cell(row=idx, column=5, value=telefone.upper() if telefone else '')
            
            ws.cell(row=idx, column=6, value='')  # NATURALIDADE (não disponível no banco)
            
            cep = func.get('cep', '')
            ws.cell(row=idx, column=7, value=cep.upper() if cep else '')
            
            endereco = func.get('endereco_completo', '')
            ws.cell(row=idx, column=8, value=endereco.upper() if endereco else '')
            
            # FUNÇÃO - usar funcao se disponível, senão usar cargo
            funcao = func.get('funcao') or func.get('cargo', '')
            ws.cell(row=idx, column=9, value=funcao.upper() if funcao else '')
            
            # REGIME - mapear vinculo
            vinculo = func.get('vinculo', '')
            ws.cell(row=idx, column=10, value=vinculo.upper() if vinculo else '')
        
        # Aplicar estilo aos dados
        aplicar_estilo_dados(ws, len(funcionarios) + 1, len(headers))
        
        # Ajustar largura das colunas
        ajustar_largura_colunas(ws)
        
        # Definir caminho de saída
        if not caminho_saida:
            # Criar diretório se não existir
            dir_documentos = os.path.join(os.getcwd(), 'documentos_gerados')
            os.makedirs(dir_documentos, exist_ok=True)
            
            # Nome do arquivo com timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"funcionarios_administrativos_{timestamp}.xlsx"
            caminho_saida = os.path.join(dir_documentos, nome_arquivo)
        
        # Salvar arquivo
        wb.save(caminho_saida)
        logger.info(f"Arquivo Excel gerado: {caminho_saida}")
        
        # Mensagem de sucesso
        messagebox.showinfo(
            "Sucesso",
            f"Arquivo gerado com sucesso!\n\n"
            f"Total de funcionários: {len(funcionarios)}\n"
            f"Local: {caminho_saida}"
        )
        
        return caminho_saida
        
    except Exception as e:
        logger.exception(f"Erro ao gerar Excel de funcionários: {e}")
        messagebox.showerror(
            "Erro",
            f"Erro ao gerar arquivo Excel:\n{str(e)}"
        )
        return None


if __name__ == "__main__":
    # Teste do módulo
    print("Testando geração de Excel de funcionários administrativos...")
    arquivo = gerar_excel_funcionarios_administrativos()
    if arquivo:
        print(f"✅ Arquivo gerado: {arquivo}")
    else:
        print("❌ Falha ao gerar arquivo")
