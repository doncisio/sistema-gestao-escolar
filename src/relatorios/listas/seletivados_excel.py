"""
Módulo para exportação de dados dos profissionais seletivados para Excel.

Gera planilha com professores, tutores e cuidadores provenientes
de processos seletivos (vínculo = 'Seletivo').
"""

import os
from datetime import datetime
from typing import Optional, List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from tkinter import messagebox

from src.core.config_logs import get_logger
from db.connection import get_cursor

logger = get_logger(__name__)


def buscar_seletivados(escola_id: int = 60) -> List[Dict[str, Any]]:
    """
    Busca professores, tutores e cuidadores com vínculo 'Seletivo'.

    Args:
        escola_id: ID da escola (padrão: 60)

    Returns:
        Lista de dicionários com dados dos profissionais seletivados.
    """
    query = """
    SELECT
        f.nome,
        e.nome AS escola_lotacao,
        f.cargo,
        f.funcao,
        f.turno,
        f.carga_horaria
    FROM
        funcionarios f
        LEFT JOIN escolas e ON e.id = f.escola_id
    WHERE
        f.escola_id = %s
        AND f.vinculo = 'Seletivo'
        AND f.cargo IN (
            'Professor@',
            'Tutor',
            'Professora de Atendimento Educacional Especializado (AEE)',
            'Professor de Atendimento Educacional Especializado (AEE)'
        )
    ORDER BY
        f.cargo, f.nome
    """

    try:
        with get_cursor() as cursor:
            cursor.execute(query, (escola_id,))
            resultados = cursor.fetchall()
            logger.info(f"Encontrados {len(resultados)} profissionais seletivados")
            return resultados
    except Exception as e:
        logger.exception(f"Erro ao buscar seletivados: {e}")
        return []


def gerar_excel_seletivados(
    escola_id: int = 60,
    caminho_saida: Optional[str] = None,
) -> Optional[str]:
    """
    Gera arquivo Excel com a relação dos profissionais seletivados.

    Colunas:
        Nº | Nome do Profissional | Escola de Lotação | Cargo/Função |
        Turno | Carga Horária Semanal | Situação (Em efetivo exercício?)

    Args:
        escola_id: ID da escola (padrão: 60)
        caminho_saida: Caminho para salvar o arquivo (opcional)

    Returns:
        Caminho do arquivo gerado ou None em caso de erro.
    """
    try:
        funcionarios = buscar_seletivados(escola_id)

        if not funcionarios:
            messagebox.showwarning(
                "Aviso",
                "Nenhum profissional seletivado encontrado.",
            )
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Profissionais Seletivados"

        # ── Cabeçalhos ──────────────────────────────────────────────
        headers = [
            "Nº",
            "Nome do Profissional",
            "Escola de Lotação",
            "Cargo/Função",
            "Turno",
            "Carga Horária Semanal",
            "Situação (Em efetivo exercício?)",
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # ── Estilo do cabeçalho ─────────────────────────────────────
        header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # ── Dados ───────────────────────────────────────────────────
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="center")

        for idx, func in enumerate(funcionarios, start=1):
            row = idx + 1

            # Nº
            ws.cell(row=row, column=1, value=idx)

            # Nome do Profissional
            nome = func.get("nome", "")
            ws.cell(row=row, column=2, value=nome.upper() if nome else "")

            # Escola de Lotação
            escola = func.get("escola_lotacao", "")
            ws.cell(row=row, column=3, value=escola.upper() if escola else "")

            # Cargo/Função – usa função quando disponível, senão o cargo
            cargo = func.get("cargo", "")
            funcao = func.get("funcao", "")
            cargo_funcao = funcao if funcao else cargo
            # Abreviações de funções longas
            _abreviacoes = {
                "Professora de Atendimento Educacional Especializado (AEE)": "Professora de AEE",
                "Professor de Atendimento Educacional Especializado (AEE)": "Professor de AEE",
                "Docente de AEE - Sala de Recursos Multifuncionais (SRM)": "Professora de AEE",
            }
            cargo_funcao = _abreviacoes.get(cargo_funcao, cargo_funcao)
            ws.cell(row=row, column=4, value=cargo_funcao.upper() if cargo_funcao else "")

            # Turno
            turno = func.get("turno", "")
            ws.cell(row=row, column=5, value=turno.upper() if turno else "")

            # Carga Horária Semanal
            carga = func.get("carga_horaria", "")
            ws.cell(row=row, column=6, value=carga if carga else "")

            # Situação – padrão "Sim"
            ws.cell(row=row, column=7, value="Sim")

            # Estilo das células de dados
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col in (2, 3, 4):  # Nome, Escola, Cargo
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center

        # ── Largura das colunas ─────────────────────────────────────
        larguras = {
            "A": 6,   # Nº
            "B": 40,  # Nome do Profissional
            "C": 40,  # Escola de Lotação
            "D": 25,  # Cargo/Função
            "E": 22,  # Turno
            "F": 22,  # Carga Horária Semanal
            "G": 32,  # Situação
        }
        for col_letter, largura in larguras.items():
            ws.column_dimensions[col_letter].width = largura

        # ── Salvar ──────────────────────────────────────────────────
        if not caminho_saida:
            from src.services.utils.pdf import get_output_path
            caminho_saida = get_output_path("Lista Profissionais Seletivados", ".xlsx", "Listas")

        wb.save(caminho_saida)
        logger.info(f"Arquivo Excel de seletivados gerado: {caminho_saida}")

        messagebox.showinfo(
            "Sucesso",
            f"Arquivo gerado com sucesso!\n\n"
            f"Total de profissionais: {len(funcionarios)}\n"
            f"Local: {caminho_saida}",
        )

        return caminho_saida

    except Exception as e:
        logger.exception(f"Erro ao gerar Excel de seletivados: {e}")
        messagebox.showerror(
            "Erro",
            f"Erro ao gerar arquivo Excel:\n{str(e)}",
        )
        return None


if __name__ == "__main__":
    print("Gerando Excel de profissionais seletivados...")
    arquivo = gerar_excel_seletivados()
    if arquivo:
        print(f"✅ Arquivo gerado: {arquivo}")
    else:
        print("❌ Falha ao gerar arquivo")
