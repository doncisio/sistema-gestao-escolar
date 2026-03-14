"""
Script para automatizar a extração de notas do GEDUC
Faz login, navega pelas páginas e extrai todas as notas automaticamente
"""

from selenium import webdriver
from typing import Optional
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException

from src.core.config_logs import get_logger

logger = get_logger(__name__)

# Tentar importar webdriver-manager (facilita instalação do ChromeDriver)
try:
    from webdriver_manager.chrome import ChromeDriverManager
    WEBDRIVER_MANAGER_DISPONIVEL = True
except ImportError:
    WEBDRIVER_MANAGER_DISPONIVEL = False
    logger.warning("⚠ webdriver-manager não instalado. Instale com: pip install webdriver-manager")
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import time
import os
import re
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, ttk
import json


class AutomacaoGEDUC:
    """
    Classe para automatizar extração de notas do GEDUC
    """

    # Mapeamento: número da série → IDCURSOORI no GEduc
    # Fonte: opções do select#IDCURSOORI extraídas do DeclaracaoForm
    _GEDUC_SERIE_IDS: dict = {
        '1': '4',   # 1º ANO
        '2': '7',   # 2º ANO
        '3': '8',   # 3º ANO
        '4': '10',  # 4º ANO
        '5': '12',  # 5º ANO
        '6': '13',  # 6º ANO
        '7': '14',  # 7º ANO
        '8': '15',  # 8º ANO
        '9': '16',  # 9º ANO
    }

    def __init__(self, headless=False):
        """
        Inicializa o navegador
        headless: Se True, executa sem abrir janela do navegador
        """
        # Tipo opcional: pode ser None até o navegador ser iniciado
        self.driver: Optional[webdriver.Chrome] = None
        self.headless = headless
        self.url_base = "https://semed.geduc.com.br"
        self.dados_extraidos = []
        
    def iniciar_navegador(self):
        """
        Configura e inicia o navegador Chrome
        """
        try:
            chrome_options = Options()
            
            if self.headless:
                chrome_options.add_argument("--headless")
            
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # Tentar diferentes métodos de inicialização
            driver_iniciado = False
            metodo_usado = ""
            
            # MÉTODO 1: Tentar usar ChromeDriver do PATH (mais rápido)
            if not driver_iniciado:
                try:
                    logger.info("→ Tentando usar ChromeDriver do sistema...")
                    self.driver = webdriver.Chrome(options=chrome_options)
                    driver_iniciado = True
                    metodo_usado = "ChromeDriver do sistema (PATH)"
                except Exception as e1:
                    logger.warning("  ✗ ChromeDriver do sistema não encontrado")
            
            # MÉTODO 2: Usar webdriver-manager com cache
            if not driver_iniciado and WEBDRIVER_MANAGER_DISPONIVEL:
                try:
                    logger.info("→ Tentando usar webdriver-manager (cache)...")
                    service = Service(ChromeDriverManager().install())
                    self.driver = webdriver.Chrome(service=service, options=chrome_options)
                    driver_iniciado = True
                    metodo_usado = "webdriver-manager (cache)"
                except Exception as e2:
                    logger.exception("  ✗ Erro com webdriver-manager: %s", e2)
                    
                    # Se erro foi de conexão, tentar cache offline
                    if "Could not reach host" in str(e2) or "offline" in str(e2).lower():
                        try:
                            logger.info("→ Tentando usar cache offline do webdriver-manager...")
                            # Forçar uso de cache existente
                            import webdriver_manager.chrome as wm_chrome
                            cache_path = os.path.join(os.path.expanduser("~"), ".wdm", "drivers", "chromedriver")
                            
                            if os.path.exists(cache_path):
                                # Procurar executável no cache
                                for root, dirs, files in os.walk(cache_path):
                                    for file in files:
                                        if file == "chromedriver.exe" or file == "chromedriver":
                                            driver_path = os.path.join(root, file)
                                            logger.info("  → Encontrado no cache: %s", driver_path)
                                            service = Service(driver_path)
                                            self.driver = webdriver.Chrome(service=service, options=chrome_options)
                                            driver_iniciado = True
                                            metodo_usado = "webdriver-manager (cache offline)"
                                            break
                                    if driver_iniciado:
                                        break
                        except Exception as e3:
                            logger.exception("  ✗ Cache offline não disponível: %s", e3)
            
            # MÉTODO 3: Procurar chromedriver.exe na pasta do script
            if not driver_iniciado:
                try:
                    logger.info("→ Procurando chromedriver.exe na pasta do script...")
                    script_dir = os.path.dirname(os.path.abspath(__file__))
                    local_chromedriver = os.path.join(script_dir, "chromedriver.exe")
                    
                    if os.path.exists(local_chromedriver):
                        logger.info("  → Encontrado: %s", local_chromedriver)
                        service = Service(local_chromedriver)
                        self.driver = webdriver.Chrome(service=service, options=chrome_options)
                        driver_iniciado = True
                        metodo_usado = "ChromeDriver local (pasta do script)"
                    else:
                        logger.warning("  ✗ Não encontrado em: %s", local_chromedriver)
                except Exception as e4:
                    logger.exception("  ✗ Erro ao usar ChromeDriver local: %s", e4)
            
            # Se nenhum método funcionou
            if not driver_iniciado:
                raise Exception("Nenhum método de inicialização funcionou")
            
            # Sucesso!
            assert self.driver is not None, "navegador não iniciado"
            self.driver.set_page_load_timeout(30)
            logger.info("✓ Navegador iniciado com sucesso usando: %s", metodo_usado)
            return True
            
        except Exception as e:
            logger.exception("✗ Erro ao iniciar navegador: %s", e)
            
            # Mensagem de erro personalizada
            mensagem_erro = "❌ Não foi possível iniciar o navegador Chrome.\n\n"
            mensagem_erro += "📝 SOLUÇÕES DISPONÍVEIS:\n\n"
            
            mensagem_erro += "1️⃣  SOLUÇÃO RÁPIDA (Recomendada):\n"
            mensagem_erro += "   • Baixe chromedriver.exe manualmente\n"
            mensagem_erro += "   • Site: https://googlechromelabs.github.io/chrome-for-testing/\n"
            mensagem_erro += f"   • Copie para: {os.path.dirname(os.path.abspath(__file__))}\n\n"
            
            mensagem_erro += "2️⃣  VERIFICAR CHROME:\n"
            mensagem_erro += "   • Certifique-se de ter Google Chrome instalado\n"
            mensagem_erro += "   • Atualize para a versão mais recente\n\n"
            
            mensagem_erro += "3️⃣  CONEXÃO DE INTERNET:\n"
            mensagem_erro += "   • Verifique sua conexão\n"
            mensagem_erro += "   • Desative proxy/firewall temporariamente\n\n"
            
            mensagem_erro += f"🔍 Erro técnico: {str(e)[:200]}"
            
            messagebox.showerror("Erro ao Iniciar Navegador", mensagem_erro)
            return False
    
    def fazer_login(self, usuario, senha, timeout_recaptcha=60):
        """
        Faz login no sistema GEDUC.
        Preenche credenciais e clica no botão de login automaticamente.
        Se o reCAPTCHA bloquear, aguarda resolução manual e clica novamente.
        """
        if self.driver is None:
            logger.error("✗ Erro: navegador não iniciado. Chame iniciar_navegador() antes de fazer login.")
            return False

        try:
            logger.info("→ Acessando página de login...")
            self.driver.get(f"{self.url_base}/index.php?class=LoginForm")

            wait = WebDriverWait(self.driver, 15)
            campo_usuario = wait.until(
                EC.presence_of_element_located((By.NAME, "login"))
            )

            campo_usuario.clear()
            campo_usuario.send_keys(usuario)
            logger.info("  ✓ Usuário preenchido: %s", usuario)

            campo_senha = self.driver.find_element(By.NAME, "password")
            campo_senha.clear()
            campo_senha.send_keys(senha)
            logger.info("  ✓ Senha preenchida")

            # ── Clicar no botão de login automaticamente ──────────────────
            def _clicar_botao_login():
                """Tenta clicar no botão de submit do formulário de login."""
                seletores = [
                    (By.CSS_SELECTOR, "button[type='submit']"),
                    (By.CSS_SELECTOR, "input[type='submit']"),
                    (By.CSS_SELECTOR, "button.btn-primary"),
                    (By.XPATH, "//button[contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'ENTRAR') or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'LOGIN')]"),
                    (By.XPATH, "//input[@type='submit']"),
                    (By.CSS_SELECTOR, "form button"),
                ]
                for by, sel in seletores:
                    try:
                        btn = self.driver.find_element(by, sel)
                        if btn.is_displayed() and btn.is_enabled():
                            btn.click()
                            return True
                    except Exception:
                        continue
                # Fallback: submit via JS no primeiro formulário
                try:
                    self.driver.execute_script(
                        "document.querySelector('form').submit();"
                    )
                    return True
                except Exception:
                    pass
                return False

            clicou = _clicar_botao_login()
            if clicou:
                logger.info("  ✓ Botão de login clicado automaticamente")
            else:
                logger.warning("  ⚠ Não foi possível clicar no botão de login automaticamente")

            # ── Aguardar resultado: login ok ou reCAPTCHA pendente ────────
            time.sleep(2)
            if "LoginForm" not in self.driver.current_url:
                logger.info("✓ Login realizado com sucesso!")
                return True

            # reCAPTCHA provavelmente apareceu — aguardar resolução manual
            logger.info("\n%s", "="*60)
            logger.info("⚠️  reCAPTCHA detectado — resolva manualmente no navegador")
            logger.info("→ Marque 'Não sou um robô' e, se necessário, complete o desafio")
            logger.info("→ O sistema clicará em Login automaticamente após a resolução")
            logger.info("→ Você tem %s segundos", timeout_recaptcha)
            logger.info("%s\n", "="*60)

            tempo_inicio = time.time()
            while True:
                time.sleep(1.5)

                # Saiu da página de login → sucesso
                if "LoginForm" not in self.driver.current_url:
                    logger.info("✓ Login realizado com sucesso!")
                    return True

                # Verificar se o reCAPTCHA foi resolvido (token disponível)
                try:
                    token = self.driver.execute_script(
                        "return (typeof grecaptcha !== 'undefined' && grecaptcha.getResponse) "
                        "? grecaptcha.getResponse() : '';"
                    )
                    if token:
                        logger.info("  ✓ reCAPTCHA resolvido — clicando em login...")
                        _clicar_botao_login()
                        time.sleep(2)
                        if "LoginForm" not in self.driver.current_url:
                            logger.info("✓ Login realizado com sucesso!")
                            return True
                except Exception:
                    pass

                tempo_decorrido = time.time() - tempo_inicio
                if tempo_decorrido > timeout_recaptcha:
                    logger.error("✗ Timeout de %ds expirado sem login concluído", timeout_recaptcha)
                    return False

                tempo_restante = int(timeout_recaptcha - tempo_decorrido)
                if tempo_restante % 10 == 0:
                    logger.info("  ⏳ Aguardando reCAPTCHA... (%ds restantes)", tempo_restante)

        except TimeoutException:
            logger.error("✗ Timeout ao carregar página de login")
            return False
        except Exception as e:
            logger.exception("✗ Erro durante login: %s", e)
            return False
    
    def mudar_ano_letivo(self, ano_letivo: int = 2025):
        """
        Muda o ano letivo no GEDUC após o login
        
        Args:
            ano_letivo: Ano letivo desejado (ex: 2025, 2026)
        """
        if self.driver is None:
            logger.error("✗ Erro: navegador não iniciado")
            return False
        
        try:
            logger.info(f"→ Mudando para ano letivo {ano_letivo}...")
            
            # Acessar página de mudança de ano
            self.driver.get(f"{self.url_base}/index.php?class=AltAnoletivo")
            
            # Aguardar carregamento
            wait = WebDriverWait(self.driver, 10)
            time.sleep(2)
            
            ano_selecionado = False
            
            # ESTRATÉGIA 1: Procurar campo de input/select para preencher
            try:
                logger.info(f"  → Tentativa 1: Procurando campo de entrada para digitar o ano...")
                
                # Tentar encontrar input de texto
                inputs_texto = self.driver.find_elements(By.XPATH, "//input[@type='text' or @type='number' or not(@type)]")
                if inputs_texto:
                    logger.info(f"    ✓ Encontrados {len(inputs_texto)} campo(s) de entrada")
                    for idx, campo in enumerate(inputs_texto):
                        try:
                            # Limpar campo e preencher com o ano
                            campo.clear()
                            campo.send_keys(str(ano_letivo))
                            logger.info(f"    ✓ Ano {ano_letivo} digitado no campo {idx+1}")
                            
                            # Pressionar ENTER
                            from selenium.webdriver.common.keys import Keys
                            campo.send_keys(Keys.RETURN)
                            logger.info(f"    ✓ ENTER pressionado")
                            time.sleep(3)
                            
                            ano_selecionado = True
                            break
                        except Exception as e:
                            logger.warning(f"    ⚠ Erro no campo {idx+1}: {e}")
                            continue
                
                # Se não encontrou input, tentar select
                if not ano_selecionado:
                    selects = self.driver.find_elements(By.TAG_NAME, "select")
                    if selects:
                        logger.info(f"    ✓ Encontrados {len(selects)} campo(s) select")
                        for idx, select_elem in enumerate(selects):
                            try:
                                select = Select(select_elem)
                                # Tentar selecionar por valor
                                try:
                                    select.select_by_value(str(ano_letivo))
                                    logger.info(f"    ✓ Ano {ano_letivo} selecionado no select {idx+1} (por valor)")
                                    time.sleep(2)
                                    ano_selecionado = True
                                    break
                                except:
                                    # Tentar selecionar por texto visível
                                    select.select_by_visible_text(str(ano_letivo))
                                    logger.info(f"    ✓ Ano {ano_letivo} selecionado no select {idx+1} (por texto)")
                                    time.sleep(2)
                                    ano_selecionado = True
                                    break
                            except Exception as e:
                                logger.warning(f"    ⚠ Erro no select {idx+1}: {e}")
                                continue
            except Exception as e:
                logger.warning(f"  ✗ Estratégia 1 falhou: {e}")
            
            # ESTRATÉGIA 2: Procurar links diretos com o ano
            if not ano_selecionado:
                try:
                    logger.info(f"  → Tentativa 2: Procurando link direto com ano {ano_letivo}...")
                    xpath_link = f"//a[contains(text(), '{ano_letivo}')]"
                    elementos = self.driver.find_elements(By.XPATH, xpath_link)
                    
                    if elementos:
                        logger.info(f"    ✓ Encontrados {len(elementos)} link(s) com o ano")
                        for idx, elemento in enumerate(elementos):
                            try:
                                texto = elemento.text.strip()
                                logger.info(f"    → Tentando clicar em: '{texto}'")
                                
                                # Tentar click normal
                                try:
                                    elemento.click()
                                    logger.info(f"    ✓ Click bem-sucedido")
                                    time.sleep(2)
                                    ano_selecionado = True
                                    break
                                except:
                                    # Tentar JavaScript click
                                    self.driver.execute_script("arguments[0].click();", elemento)
                                    logger.info(f"    ✓ Click via JavaScript bem-sucedido")
                                    time.sleep(2)
                                    ano_selecionado = True
                                    break
                            except Exception as e:
                                logger.warning(f"    ⚠ Erro ao clicar no elemento {idx+1}: {e}")
                                continue
                except Exception as e:
                    logger.warning(f"  ✗ Estratégia 2 falhou: {e}")
            
            # ESTRATÉGIA 3: Procurar botões/divs clicáveis
            if not ano_selecionado:
                try:
                    logger.info(f"  → Tentativa 3: Procurando outros elementos clicáveis...")
                    xpath_geral = f"//*[contains(text(), '{ano_letivo}') and (self::div or self::button or self::span)]"
                    elementos = self.driver.find_elements(By.XPATH, xpath_geral)
                    
                    if elementos:
                        logger.info(f"    ✓ Encontrados {len(elementos)} elemento(s) clicáveis")
                        for elemento in elementos:
                            try:
                                self.driver.execute_script("arguments[0].click();", elemento)
                                logger.info(f"    ✓ Click bem-sucedido em: {elemento.tag_name}")
                                time.sleep(2)
                                ano_selecionado = True
                                break
                            except Exception as e:
                                logger.warning(f"    ⚠ Erro ao clicar: {e}")
                                continue
                except Exception as e:
                    logger.warning(f"  ✗ Estratégia 3 falhou: {e}")
            
            if ano_selecionado:
                logger.info(f"✓ Ano letivo alterado para {ano_letivo}")
                return True
            else:
                logger.error(f"✗ Não foi possível selecionar o ano {ano_letivo}")
                # Salvar screenshot para debug
                try:
                    screenshot_path = f"erro_ano_letivo_{ano_letivo}.png"
                    self.driver.save_screenshot(screenshot_path)
                    logger.info(f"  Screenshot salvo em: {screenshot_path}")
                except:
                    pass
                return False
                
        except Exception as e:
            logger.exception(f"✗ Erro ao mudar ano letivo: %s", e)
            return False
    
    def acessar_registro_notas(self):
        """
        Navega até a página de registro de notas
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            logger.info("→ Navegando para registro de notas...")
            
            # URL da página de registro de notas
            self.driver.get(f"{self.url_base}/index.php?class=RegNotasForm")
            
            # Aguardar carregamento
            wait = WebDriverWait(self.driver, 15)
            wait.until(EC.presence_of_element_located((By.NAME, "IDTURMA")))
            
            logger.info("✓ Página de registro de notas carregada")
            return True
            
        except Exception as e:
            logger.exception("✗ Erro ao acessar registro de notas: %s", e)
            return False
    
    def acessar_recuperacao_bimestral(self):
        """
        Navega até a página de recuperação bimestral
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            logger.info("→ Navegando para recuperação bimestral...")
            
            # URL da página de recuperação bimestral
            self.driver.get(f"{self.url_base}/index.php?class=RegNotasbimForm")
            
            # Aguardar carregamento
            wait = WebDriverWait(self.driver, 15)
            wait.until(EC.presence_of_element_located((By.NAME, "IDTURMA")))
            
            logger.info("✓ Página de recuperação bimestral carregada")
            return True
            
        except Exception as e:
            logger.exception("✗ Erro ao acessar recuperação bimestral: %s", e)
            return False
    
    def acessar_notas_finais(self):
        """
        Navega até a página de notas finais (recuperação anual)
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            logger.info("→ Navegando para notas finais...")
            
            # URL da página de notas finais
            self.driver.get(f"{self.url_base}/index.php?class=RegNotasFinaisForm")
            
            # Aguardar carregamento
            wait = WebDriverWait(self.driver, 15)
            wait.until(EC.presence_of_element_located((By.NAME, "IDTURMA")))
            
            logger.info("✓ Página de notas finais carregada")
            return True
            
        except Exception as e:
            logger.exception("✗ Erro ao acessar notas finais: %s", e)
            return False
    
    def obter_opcoes_select(self, select_name):
        """
        Obtém todas as opções de um elemento select
        Retorna lista de dicionários com {'value': valor, 'text': texto}
        """
        try:
            assert self.driver is not None, "navegador não iniciado"

            # Estratégia 1: encontrar por NAME (mais comum)
            try:
                select_element = Select(self.driver.find_element(By.NAME, select_name))
            except Exception:
                select_element = None

            # Estratégia 2: encontrar por CSS parcial no name/id
            if select_element is None:
                try:
                    select_elem = self.driver.find_element(By.CSS_SELECTOR, f"select[name*='{select_name}']")
                    select_element = Select(select_elem)
                except Exception:
                    select_element = None

            if select_element is None:
                try:
                    select_elem = self.driver.find_element(By.CSS_SELECTOR, f"select[id*='{select_name}']")
                    select_element = Select(select_elem)
                except Exception:
                    select_element = None

            # Estratégia 3: vasculhar todos os <select> e tentar identificar pelo atributo name/id
            if select_element is None:
                try:
                    selects = self.driver.find_elements(By.TAG_NAME, 'select')
                    for s in selects:
                        name_attr = s.get_attribute('name') or ''
                        id_attr = s.get_attribute('id') or ''
                        if select_name.lower() in name_attr.lower() or select_name.lower() in id_attr.lower():
                            select_element = Select(s)
                            break
                except Exception:
                    select_element = None

            # Estratégia 4: fallback com BeautifulSoup (quando Selenium não encontra por atributos)
            opcoes = []
            if select_element is None:
                try:
                    html = self.driver.page_source
                    soup = BeautifulSoup(html, 'html.parser')
                    # Procurar selects cujo name/id contenha select_name
                    candidates = []
                    for s in soup.find_all('select'):
                        name_attr = s.get('name') or ''
                        id_attr = s.get('id') or ''
                        if select_name.lower() in name_attr.lower() or select_name.lower() in id_attr.lower():
                            candidates.append(s)

                    # Se não encontrou por atributo, pegar selects com opções não-vazias
                    if not candidates:
                        for s in soup.find_all('select'):
                            opts = s.find_all('option')
                            if opts and len(opts) > 1:
                                candidates.append(s)

                    # Extrair opções do primeiro candidato
                    if candidates:
                        s = candidates[0]
                        for opt in s.find_all('option'):
                            valor = opt.get('value')
                            texto = (opt.text or '').strip()
                            if valor and valor != '' and not texto.startswith('---'):
                                opcoes.append({'value': valor, 'text': texto})
                        return opcoes
                except Exception as e:
                    logger.exception("✗ Erro no fallback BeautifulSoup para select '%s': %s", select_name, e)

            # Se temos um Select via Selenium, extrair opções
            if select_element is not None:
                for option in select_element.options:
                    valor = option.get_attribute('value')
                    texto = option.text.strip()
                    # Pular opções vazias ou de instrução
                    if valor and valor != '' and not texto.startswith('---'):
                        opcoes.append({'value': valor, 'text': texto})

            return opcoes

        except Exception as e:
            logger.exception("✗ Erro ao obter opções do select '%s': %s", select_name, e)
            return []
    
    def selecionar_opcao(self, select_name, valor):
        """
        Seleciona uma opção em um elemento select pelo valor
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            select_element = Select(self.driver.find_element(By.NAME, select_name))
            select_element.select_by_value(str(valor))
            time.sleep(0.5)  # Aguardar processamento
            return True
        except Exception as e:
            logger.exception("✗ Erro ao selecionar opção '%s' em '%s': %s", valor, select_name, e)
            return False
    
    def selecionar_bimestre(self, numero_bimestre):
        """
        Seleciona o bimestre pelos radio buttons
        numero_bimestre: 1, 2, 3 ou 4
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            # Buscar todos os radio buttons de avaliação
            radios = self.driver.find_elements(By.NAME, "IDAVALIACOES")
            
            for radio in radios:
                radio_id = radio.get_attribute('id')
                # Buscar label associada
                try:
                    label = self.driver.find_element(By.CSS_SELECTOR, f"label[for='{radio_id}']")
                    texto_label = label.text.strip()
                    
                    # Verificar se contém o número do bimestre
                    if f"{numero_bimestre}º" in texto_label or f"{numero_bimestre}°" in texto_label:
                        # Clicar no radio button
                        self.driver.execute_script("arguments[0].click();", radio)
                        time.sleep(0.5)
                        return True
                except:
                    continue
            
            logger.warning("✗ Bimestre %sº não encontrado", numero_bimestre)
            return False
            
        except Exception as e:
            logger.exception("✗ Erro ao selecionar bimestre: %s", e)
            return False
    
    def clicar_exibir_alunos(self):
        """
        Clica no botão para exibir os alunos
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            # Buscar botão de exibir alunos
            botoes = self.driver.find_elements(By.TAG_NAME, "button")
            
            for botao in botoes:
                texto = botao.text.strip().upper()
                if "EXIBIR" in texto or "ALUNOS" in texto or "CARREGAR" in texto:
                    self.driver.execute_script("arguments[0].click();", botao)
                    time.sleep(2)  # Aguardar carregamento da tabela
                    return True
            
            # Tentar também por input type=button
            inputs = self.driver.find_elements(By.CSS_SELECTOR, "input[type='button']")
            for input_btn in inputs:
                valor = input_btn.get_attribute('value')
                if valor and ("EXIBIR" in valor.upper() or "ALUNOS" in valor.upper()):
                    self.driver.execute_script("arguments[0].click();", input_btn)
                    time.sleep(2)
                    return True
            
            logger.warning("✗ Botão 'Exibir Alunos' não encontrado")
            return False
            
        except Exception as e:
            logger.exception("✗ Erro ao clicar em exibir alunos: %s", e)
            return False
    
    def extrair_medias_bimestre(self, bimestre_numero):
        """
        Extrai apenas as MÉDIAS dos alunos para o bimestre especificado.
        Retorna lista de alunos com nome e média.
        
        Args:
            bimestre_numero: Número do bimestre (1, 2, 3 ou 4)
        
        Returns:
            Lista de dicionários: [{'nome': 'NOME ALUNO', 'media': 85.0}, ...]
        """
        try:
            # Extrair todas as notas
            dados = self.extrair_notas_pagina_atual(bimestre_numero=bimestre_numero)
            
            if not dados or not dados.get('alunos'):
                return []
            
            # Retornar apenas nome e média
            medias = []
            for aluno in dados['alunos']:
                medias.append({
                    'nome': aluno.get('nome', ''),
                    'media': aluno.get('media', '')
                })
            
            return medias
            
        except Exception as e:
            logger.exception("Erro ao extrair médias do bimestre %s: %s", bimestre_numero, e)
            return []
    
    def extrair_notas_pagina_atual(self, turma_nome=None, disciplina_nome=None, bimestre_numero=None):
        """
        Extrai notas da página atual usando BeautifulSoup
        Retorna dicionário com turma, disciplina, bimestre e lista de alunos
        
        turma_nome: Nome da turma (se None, tenta extrair do HTML)
        disciplina_nome: Nome da disciplina (se None, tenta extrair do HTML)
        bimestre_numero: Número do bimestre (se None, tenta extrair do HTML)
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            # Obter HTML da página
            html_content = self.driver.page_source
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Usar parâmetros passados ou tentar extrair do HTML
            if turma_nome is None:
                turma = "Desconhecida"
                turma_select = soup.find('select', {'name': 'IDTURMA'})
                if turma_select:
                    turma_option = turma_select.find('option', {'selected': True})
                    if turma_option:
                        turma = turma_option.text.strip()
            else:
                turma = turma_nome
            
            if disciplina_nome is None:
                disciplina = "Desconhecida"
                disciplina_select = soup.find('select', {'name': 'IDTURMASDISP'})
                if disciplina_select:
                    disciplina_option = disciplina_select.find('option', {'selected': True})
                    if disciplina_option:
                        disciplina = disciplina_option.text.strip()
            else:
                disciplina = disciplina_nome
            
            if bimestre_numero is None:
                bimestre = "1º"
                bimestre_radios = soup.find_all('input', {'name': 'IDAVALIACOES'})
                for radio in bimestre_radios:
                    if radio.get('checked'):
                        radio_id = radio.get('id')
                        label = soup.find('label', {'for': radio_id})
                        if label:
                            texto_bimestre = label.text.strip()
                            match = re.search(r'(\d+)º', texto_bimestre)
                            if match:
                                bimestre = f"{match.group(1)}º"
            else:
                bimestre = f"{bimestre_numero}º"
            
            # Extrair alunos e notas da tabela
            alunos_notas = []
            tbody = soup.find('tbody', {'class': 'tdatagrid_body'})
            
            if tbody:
                rows = tbody.find_all('tr', {'class': ['tdatagrid_row_odd', 'tdatagrid_row_even']})
                
                for row in rows:
                    cells = row.find_all('td', {'class': 'tdatagrid_cell'})
                    
                    if len(cells) >= 2:
                        ordem_text = cells[0].text.strip()
                        nome_aluno = cells[1].text.strip()
                        
                        if ordem_text and nome_aluno and ordem_text.isdigit():
                            # Extrair notas
                            if len(cells) >= 3:
                                nota_inputs = cells[2].find_all('input', {'class': 'tfield'})
                                notas = []
                                
                                for input_nota in nota_inputs:
                                    valor_nota = input_nota.get('value', '')
                                    if valor_nota:
                                        try:
                                            nota_float = float(valor_nota)
                                            notas.append(nota_float)
                                        except ValueError:
                                            pass
                                
                                # Calcular média
                                if notas:
                                    media = sum(notas) / len(notas)
                                    nota_final = media * 10
                                else:
                                    nota_final = ''
                                    notas = []  # Lista vazia se não houver notas
                                
                                alunos_notas.append({
                                    'ordem': int(ordem_text),
                                    'nome': nome_aluno,
                                    'notas_individuais': notas,  # Guardar notas individuais
                                    'media': nota_final  # Média final (média × 10)
                                })
            
            # Log para depuração: listar nomes extraídos e checar presença de aluna específica
            try:
                nomes_extraidos = [a.get('nome', '').upper() for a in alunos_notas]
                logger.debug(f"[DEBUG_GEDUC] Alunos extraídos ({len(nomes_extraidos)}): {', '.join(nomes_extraidos[:20])}{'...' if len(nomes_extraidos)>20 else ''}")
                # Verificar presença de 'MARIA CECILYA' aproximada
                encontradas = [n for n in nomes_extraidos if 'MARIA' in n and ('CECIL' in n or 'CEC' in n or 'CECI' in n)]
                if encontradas:
                    logger.debug(f"[DEBUG_GEDUC] Possíveis ocorrências de Maria Cecilya encontradas: {encontradas}")
                else:
                    logger.debug("[DEBUG_GEDUC] Nenhuma ocorrência clara de 'Maria Cecilya' encontrada entre os nomes extraídos.")
            except Exception:
                logger.exception("Erro ao gerar debug dos nomes extraídos do GEDUC")

            return {
                'turma': turma,
                'disciplina': disciplina,
                'bimestre': bimestre,
                'alunos': alunos_notas,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
        except Exception as e:
            logger.exception("✗ Erro ao extrair notas: %s", e)
            return None
    
    def extrair_recuperacao_pagina_atual(self):
        """
        Extrai dados de recuperação da página RegNotasbimForm
        
        Estrutura REAL da tabela (descoberta via debug 01/11/2025):
        - Coluna 0: Ordem
        - Coluna 1: Matrícula
        - Coluna 2: Alunos (NOME)
        - Coluna 3: (INPUT) Média Atual campo 1
        - Coluna 4: (INPUT) Média Atual campo 2
        - Coluna 5: (INPUT) Média Atual campo 3
        - Coluna 6: (INPUT) código
        - Coluna 7: (INPUT) valor
        - Coluna 8: Situação (texto "Aprovado"/"Reprovado")
        - Coluna 9: (INPUT) RECUPERAÇÃO ← COLUNA CORRETA!
        - Coluna 10: Média Gravada
        
        Returns:
            Lista de dicts: [{'nome': 'ALUNO', 'recuperacao': 7.0}, ...]
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            # Aguardar tabela carregar
            wait = WebDriverWait(self.driver, 10)
            tabela = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table.table, table.tdatagrid_table"))
            )
            
            # Buscar todas as linhas da tabela
            linhas = tabela.find_elements(By.TAG_NAME, "tr")
            
            dados = []
            
            # Processar linhas (pular cabeçalho)
            for idx, linha in enumerate(linhas):
                # Pular cabeçalho (primeira linha geralmente)
                if idx == 0:
                    continue
                
                try:
                    colunas = linha.find_elements(By.TAG_NAME, "td")
                    
                    # Verificar se tem colunas suficientes (precisa ter pelo menos 10 colunas)
                    if len(colunas) < 10:
                        continue
                    
                    # Estrutura REAL da tabela (descoberta via debug):
                    # [0] = Ordem
                    # [1] = Matrícula
                    # [2] = Alunos (nome)
                    # [3] = (INPUT) Média Atual campo 1
                    # [4] = (INPUT) Média Atual campo 2
                    # [5] = (INPUT) Média Atual campo 3
                    # [6] = (INPUT) código
                    # [7] = (INPUT) valor
                    # [8] = Situação (texto)
                    # [9] = (INPUT) Recuperação ← COLUNA CORRETA!
                    # [10] = Média Gravada
                    
                    # Extrair nome do aluno (coluna 2)
                    nome = colunas[2].text.strip()
                    
                    if not nome:
                        continue
                    
                    # Extrair recuperação (coluna 9 - CORRIGIDO!)
                    recuperacao = None
                    try:
                        # Tentar pegar de input na coluna 9
                        recup_input = colunas[9].find_element(By.TAG_NAME, "input")
                        recup_str = recup_input.get_attribute("value")
                    except:
                        # Se não for input, pegar texto da coluna 9
                        recup_str = colunas[9].text.strip()
                    
                    # Converter para float
                    if recup_str and recup_str != '':
                        try:
                            # Substituir vírgula por ponto
                            recup_str = recup_str.replace(',', '.')
                            recuperacao = float(recup_str)
                        except ValueError:
                            recuperacao = None
                    
                    # Adicionar aos dados (mesmo se recuperação for None)
                    dados.append({
                        'nome': nome,
                        'recuperacao': recuperacao
                    })
                
                except Exception as e:
                    # Erro em linha específica, continuar para próxima
                    continue
            
            logger.info("✓ Extraídos %d registros de recuperação", len(dados))
            return dados
            
        except Exception as e:
            logger.exception("✗ Erro ao extrair dados de recuperação: %s", e)
            return []
    
    def extrair_notas_finais_pagina_atual(self):
        """
        Extrai notas finais da página RegNotasFinaisForm (recuperação anual)
        
        Esta função extrai as médias finais anuais da tabela de notas finais do GEDUC:
        - Média Calculada: média dos 4 bimestres (coluna 3)
        - Recuperação: nota da prova de recuperação anual se houver (coluna 4 - input)
        - Resultado Final (Gravada): nota final já calculada pelo GEDUC (coluna 5)
        
        A estrutura da tabela é:
        [0] Ordem | [1] Matrícula | [2] Alunos | [3] Média Calculada | [4] Recuperação | [5] Resultado Final
        
        Returns:
            Lista de dicts: [{'nome': 'ALUNO', 'media_atual': 7.5, 'recuperacao_final': 8.0, 'resultado_final': 8.0}, ...]
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            # Aguardar tabela carregar
            wait = WebDriverWait(self.driver, 10)
            tabela = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table.table, table.tdatagrid_table"))
            )
            
            # Buscar todas as linhas da tabela
            linhas = tabela.find_elements(By.TAG_NAME, "tr")
            
            dados = []
            
            # Processar linhas (pular cabeçalho)
            for idx, linha in enumerate(linhas):
                # Pular cabeçalho (primeira linha geralmente)
                if idx == 0:
                    continue
                
                try:
                    colunas = linha.find_elements(By.TAG_NAME, "td")
                    
                    # Verificar se tem colunas suficientes
                    if len(colunas) < 3:
                        continue
                    
                    # Estrutura da tabela de notas finais (RegNotasFinaisForm):
                    # [0] = Ordem
                    # [1] = Matrícula
                    # [2] = Alunos (nome)
                    # [3] = Média Calculada (texto - média dos 4 bimestres)
                    # [4] = Recuperação (input dentro de table - pode estar vazio)
                    # [5] = Resultado Final (Gravada) (texto - nota final)
                    
                    # Extrair nome do aluno
                    nome = colunas[2].text.strip() if len(colunas) > 2 else ""
                    
                    if not nome:
                        continue
                    
                    # Inicializar variáveis
                    media_atual = None
                    recuperacao_final = None
                    resultado_final = None
                    
                    # Extrair Média Calculada (coluna 3) - texto direto
                    if len(colunas) > 3:
                        try:
                            texto_media = colunas[3].text.strip().replace(',', '.')
                            if texto_media:
                                media_atual = float(texto_media)
                        except (ValueError, AttributeError):
                            pass
                    
                    # Extrair Recuperação (coluna 4) - input dentro de table
                    if len(colunas) > 4:
                        try:
                            # Buscar input name="RECF[]" que contém o valor da recuperação
                            inputs = colunas[4].find_elements(By.CSS_SELECTOR, "input[name='RECF[]']")
                            if inputs and inputs[0].get_attribute("value"):
                                valor_rec = inputs[0].get_attribute("value").strip().replace(',', '.')
                                if valor_rec:
                                    recuperacao_final = float(valor_rec)
                        except (ValueError, AttributeError, IndexError):
                            pass
                    
                    # Extrair Resultado Final (coluna 5) - texto direto
                    if len(colunas) > 5:
                        try:
                            texto_resultado = colunas[5].text.strip().replace(',', '.')
                            if texto_resultado:
                                resultado_final = float(texto_resultado)
                        except (ValueError, AttributeError):
                            pass
                    
                    # Adicionar aos dados
                    # Nota: usamos media_atual (coluna 3), mas o sistema deve usar
                    # resultado_final (coluna 5) se não houver recuperação,
                    # ou recuperacao_final (coluna 4) se existir
                    dados.append({
                        'nome': nome,
                        'media_atual': media_atual,
                        'recuperacao_final': recuperacao_final,
                        'resultado_final': resultado_final
                    })
                
                except Exception as e:
                    # Erro em linha específica, continuar para próxima
                    logger.debug("Erro ao processar linha %d: %s", idx, e)
                    continue
            
            logger.info("✓ Extraídos %d registros de notas finais", len(dados))
            return dados
            
        except Exception as e:
            logger.exception("✗ Erro ao extrair notas finais: %s", e)
            return []
    
    def extrair_todas_notas(self, turmas_selecionadas=None, bimestres=[1, 2, 3, 4], diretorio_saida="notas_extraidas", callback_progresso=None):
        """
        Extrai todas as notas de todas as combinações de turma/disciplina/bimestre
        NOVO: Cria 1 arquivo Excel por TURMA/BIMESTRE com múltiplas planilhas (uma por disciplina)
        Salva IMEDIATAMENTE após processar cada turma/bimestre
        
        turmas_selecionadas: Lista de IDs de turmas (None = todas)
        bimestres: Lista de números de bimestres para extrair
        diretorio_saida: Diretório onde salvar os arquivos
        callback_progresso: Função para atualizar progresso (opcional)
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            logger.info("\n%s", "="*60)
            logger.info("INICIANDO EXTRAÇÃO AUTOMÁTICA DE NOTAS")
            logger.info("%s\n", "="*60)
            
            # Criar diretório de saída
            if not os.path.exists(diretorio_saida):
                os.makedirs(diretorio_saida)
                logger.info("✓ Diretório criado: %s", diretorio_saida)
            
            # Acessar página de registro de notas
            if not self.acessar_registro_notas():
                return False
            
            # Obter todas as turmas disponíveis
            logger.info("→ Obtendo lista de turmas...")
            turmas = self.obter_opcoes_select('IDTURMA')
            
            if not turmas:
                logger.error("✗ Nenhuma turma encontrada")
                return False
            
            logger.info("  ✓ %d turmas encontradas", len(turmas))
            
            # Filtrar turmas se especificado
            if turmas_selecionadas:
                turmas = [t for t in turmas if t['value'] in turmas_selecionadas]
            
            total_arquivos_criados = 0
            total_planilhas_criadas = 0
            
            # Iterar por cada turma
            for idx_turma, turma in enumerate(turmas, 1):
                logger.info("\n[%d/%d] TURMA: %s", idx_turma, len(turmas), turma['text'])
                logger.info("%s", "-" * 60)
                
                # Selecionar turma
                if not self.selecionar_opcao('IDTURMA', turma['value']):
                    continue
                
                time.sleep(1)  # Aguardar carregamento de disciplinas
                
                # Obter disciplinas da turma
                disciplinas = self.obter_opcoes_select('IDTURMASDISP')
                
                if not disciplinas:
                    logger.warning("  ⚠ Nenhuma disciplina encontrada para esta turma")
                    continue
                
                logger.info("  → %d disciplinas encontradas", len(disciplinas))
                
                # Processar cada bimestre para esta turma
                for bimestre in bimestres:
                    logger.info("\n  → PROCESSANDO BIMESTRE %sº", bimestre)
                    
                    # Dicionário para armazenar dados de todas as disciplinas deste bimestre
                    dados_turma_bimestre = {
                        'turma': turma['text'],
                        'bimestre': bimestre,
                        'disciplinas': []
                    }
                    
                    # Iterar por cada disciplina
                    for idx_disc, disciplina in enumerate(disciplinas, 1):
                        logger.info("    [%d/%d] %s", idx_disc, len(disciplinas), disciplina['text'])
                        
                        # Selecionar disciplina
                        if not self.selecionar_opcao('IDTURMASDISP', disciplina['value']):
                            logger.error("✗ Erro ao selecionar")
                            continue
                        
                        time.sleep(0.5)
                        
                        # Selecionar bimestre
                        if not self.selecionar_bimestre(bimestre):
                            logger.error("✗ Erro ao selecionar bimestre")
                            continue
                        
                        # Clicar em exibir alunos
                        if not self.clicar_exibir_alunos():
                            logger.error("✗ Erro ao carregar alunos")
                            continue
                        
                        # Extrair notas (passando nome da turma, disciplina e bimestre)
                        dados = self.extrair_notas_pagina_atual(
                            turma_nome=turma['text'],
                            disciplina_nome=disciplina['text'],
                            bimestre_numero=bimestre
                        )
                        
                        if dados and dados['alunos']:
                            dados_turma_bimestre['disciplinas'].append(dados)
                            logger.info("✓ %d alunos", len(dados['alunos']))
                        else:
                            logger.warning("⚠ Sem notas")
                        
                        # Atualizar callback de progresso
                        if callback_progresso:
                            callback_progresso(total_planilhas_criadas + 1, len(turmas) * len(bimestres) * len(disciplinas))
                    
                    # SALVAR IMEDIATAMENTE após processar todas as disciplinas deste bimestre
                    if dados_turma_bimestre['disciplinas']:
                        arquivo_criado = self._salvar_turma_bimestre(dados_turma_bimestre, diretorio_saida)
                        if arquivo_criado:
                            total_arquivos_criados += 1
                            total_planilhas_criadas += len(dados_turma_bimestre['disciplinas'])
                            logger.info("\n  ✓ ARQUIVO SALVO: %s", os.path.basename(arquivo_criado))
                            logger.info("    → %d planilhas (disciplinas)", len(dados_turma_bimestre['disciplinas']))
                    else:
                        logger.warning("\n  ⚠ Nenhuma nota encontrada para o %sº bimestre", bimestre)
            
            logger.info("\n%s", "="*60)
            logger.info("EXTRAÇÃO CONCLUÍDA!")
            logger.info("Total de arquivos Excel criados: %d", total_arquivos_criados)
            logger.info("Total de planilhas (disciplinas): %d", total_planilhas_criadas)
            logger.info("Localização: %s", os.path.abspath(diretorio_saida))
            logger.info("%s\n", "="*60)
            
            return True
            
        except Exception as e:
            logger.exception("✗ Erro durante extração: %s", e)
            return False
    
    def _salvar_turma_bimestre(self, dados_turma_bimestre, diretorio_saida):
        """
        Salva um arquivo Excel com múltiplas planilhas (uma por disciplina)
        para uma turma/bimestre específico
        
        dados_turma_bimestre: {
            'turma': 'Nome da turma',
            'bimestre': número do bimestre,
            'disciplinas': [lista de dados de disciplinas]
        }
        """
        try:
            turma = dados_turma_bimestre['turma']
            bimestre = dados_turma_bimestre['bimestre']
            disciplinas = dados_turma_bimestre['disciplinas']
            
            if not disciplinas:
                return None
            
            # Limpar nome da turma para usar no arquivo
            turma_limpa = re.sub(r'[^\w\s-]', '', turma).strip().replace(' ', '_')
            
            # Nome do arquivo: Notas_TURMA_Xbim.xlsx
            nome_arquivo = f"Notas_{turma_limpa}_{bimestre}bim.xlsx"
            caminho_arquivo = os.path.join(diretorio_saida, nome_arquivo)
            
            # Criar workbook
            wb = openpyxl.Workbook()
            
            # Criar uma planilha para cada disciplina
            for idx, dados_disciplina in enumerate(disciplinas):
                # Nome da planilha (limitado a 31 caracteres do Excel)
                nome_disciplina = dados_disciplina['disciplina'][:31]
                
                # Criar planilha
                if idx == 0:
                    # Renomear a planilha ativa padrão
                    ws = wb.active
                    assert ws is not None, "Planilha padrão não encontrada"
                    ws.title = nome_disciplina
                else:
                    # Criar nova planilha
                    ws = wb.create_sheet(title=nome_disciplina)
                    assert ws is not None, "Falha ao criar nova planilha"
                
                # Preencher planilha com dados da disciplina
                self._preencher_planilha(ws, dados_disciplina)
            
            # Salvar arquivo
            wb.save(caminho_arquivo)
            return caminho_arquivo
            
        except Exception as e:
            logger.exception("\n  ✗ Erro ao salvar arquivo: %s", e)
            return None
    
    def _preencher_planilha(self, ws, dados):
        """
        Preenche uma planilha do Excel com os dados de uma disciplina
        Agora inclui colunas para cada nota individual + coluna de média
        """
        try:
            # Estilos
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            cell_font = Font(name='Arial', size=10)
            cell_alignment = Alignment(horizontal='center', vertical='center')
            cell_alignment_left = Alignment(horizontal='left', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Informações no topo
            ws['A1'] = f"Turma: {dados['turma']}"
            ws['A2'] = f"Disciplina: {dados['disciplina']}"
            ws['A3'] = f"Bimestre: {dados['bimestre']}"
            
            ws['A1'].font = Font(name='Arial', size=12, bold=True)
            ws['A2'].font = Font(name='Arial', size=12, bold=True)
            ws['A3'].font = Font(name='Arial', size=12, bold=True)
            
            # Descobrir quantas notas individuais existem (máximo entre todos os alunos)
            max_notas = 0
            for aluno in dados['alunos']:
                num_notas = len(aluno.get('notas_individuais', []))
                if num_notas > max_notas:
                    max_notas = num_notas
            
            # Cabeçalho da tabela (linha 5)
            ws['A5'] = 'Nº'
            ws['B5'] = 'Nome do Aluno'
            
            # Criar colunas para cada nota individual
            col_letra = 'C'
            for i in range(1, max_notas + 1):
                col = chr(ord('C') + i - 1)
                ws[f'{col}5'] = f'Nota {i}'
                col_letra = col
            
            # Coluna da média (depois das notas individuais)
            col_media = chr(ord(col_letra) + 1)
            ws[f'{col_media}5'] = 'Média'
            
            # Aplicar estilo ao cabeçalho
            for col_idx in range(ord('A'), ord(col_media) + 1):
                col = chr(col_idx)
                cell = ws[f'{col}5']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 45
            for i in range(max_notas):
                col = chr(ord('C') + i)
                ws.column_dimensions[col].width = 10
            ws.column_dimensions[col_media].width = 12
            
            # Dados dos alunos (começando na linha 6)
            row = 6
            for aluno in dados['alunos']:
                # Número e nome
                ws[f'A{row}'] = aluno['ordem']
                ws[f'B{row}'] = aluno['nome']
                
                # Notas individuais
                notas_individuais = aluno.get('notas_individuais', [])
                for i, nota in enumerate(notas_individuais):
                    col = chr(ord('C') + i)
                    ws[f'{col}{row}'] = round(nota, 2)
                    ws[f'{col}{row}'].alignment = cell_alignment
                    ws[f'{col}{row}'].font = cell_font
                    ws[f'{col}{row}'].border = thin_border
                
                # Preencher colunas vazias se este aluno tiver menos notas que o máximo
                for i in range(len(notas_individuais), max_notas):
                    col = chr(ord('C') + i)
                    ws[f'{col}{row}'].alignment = cell_alignment
                    ws[f'{col}{row}'].font = cell_font
                    ws[f'{col}{row}'].border = thin_border
                
                # Média
                media = aluno.get('media', '')
                if media != '':
                    ws[f'{col_media}{row}'] = round(media, 2)
                
                # Aplicar estilo às células de número, nome e média
                ws[f'A{row}'].alignment = cell_alignment
                ws[f'B{row}'].alignment = cell_alignment_left
                ws[f'{col_media}{row}'].alignment = cell_alignment
                
                for col_idx in range(ord('A'), ord(col_media) + 1):
                    col = chr(col_idx)
                    cell = ws[f'{col}{row}']
                    cell.font = cell_font
                    cell.border = thin_border
                
                row += 1
            
        except Exception as e:
            logger.exception("✗ Erro ao preencher planilha: %s", e)
    
    def salvar_dados_excel(self, diretorio_saida="notas_extraidas"):
        """
        DEPRECATED: Use extrair_todas_notas() que já salva automaticamente
        Mantido para compatibilidade com código antigo
        """
        logger.warning("⚠️ Aviso: Os arquivos já foram salvos automaticamente durante a extração")
        return []
    
    def acessar_lista_horarios(self):
        """
        Navega até a página de lista de horários de turmas
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            logger.info("→ Navegando para lista de horários...")
            
            # URL da página de lista de horários
            self.driver.get(f"{self.url_base}/index.php?class=TurmaHorariosList")
            
            # Aguardar mais tempo após mudança de ano
            time.sleep(3)
            
            # Aguardar carregamento da tabela com timeout maior
            wait = WebDriverWait(self.driver, 30)
            
            # Tentar localizar a tabela de diferentes formas
            tabela_encontrada = False
            
            try:
                # Tentar pela classe tdatagrid_body
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, "tdatagrid_body")))
                tabela_encontrada = True
                logger.info("✓ Tabela encontrada pela classe 'tdatagrid_body'")
            except TimeoutException:
                logger.warning("⚠ Classe 'tdatagrid_body' não encontrada, tentando outras formas...")
                
                # Tentar por tag table
                try:
                    wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
                    tabela_encontrada = True
                    logger.info("✓ Tabela encontrada pela tag <table>")
                except TimeoutException:
                    logger.warning("⚠ Nenhuma tabela encontrada")
            
            if tabela_encontrada:
                logger.info("✓ Página de lista de horários carregada")
                return True
            else:
                # Verificar se há mensagem de "sem dados"
                html = self.driver.page_source
                if "Nenhum registro encontrado" in html or "sem dados" in html.lower():
                    logger.warning("⚠ Página carregada mas não há horários cadastrados para este ano")
                else:
                    logger.error("✗ Não foi possível localizar tabela de horários")
                    
                # Salvar screenshot para análise
                try:
                    screenshot_path = "erro_lista_horarios.png"
                    self.driver.save_screenshot(screenshot_path)
                    logger.info(f"  Screenshot salvo em: {screenshot_path}")
                except:
                    pass
                
                return False
            
        except Exception as e:
            logger.exception("✗ Erro ao acessar lista de horários: %s", e)
            # Salvar screenshot do erro
            try:
                if self.driver:
                    screenshot_path = "erro_lista_horarios_exception.png"
                    self.driver.save_screenshot(screenshot_path)
                    logger.info(f"  Screenshot salvo em: {screenshot_path}")
            except:
                pass
            return False
    
    def extrair_horario_turma(self, turma_nome: str) -> Optional[dict]:
        """
        Extrai o horário de uma turma específica do GEDUC
        
        Args:
            turma_nome: Nome da turma a buscar (ex: "1º ANO-MATU", "6º ANO-VESP - A")
            
        Returns:
            Dict com estrutura:
            {
                'turma_nome': str,
                'turma_id': int ou None,
                'horarios': [
                    {
                        'dia': str,  # Segunda, Terça, etc
                        'horario': str,  # Linha da tabela (1-6)
                        'disciplina': str,  # Nome da disciplina
                        'professor': str ou None  # Nome do professor se disponível
                    }
                ]
            }
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            
            # Acessar lista de horários se ainda não estiver nela
            if "TurmaHorariosList" not in self.driver.current_url:
                if not self.acessar_lista_horarios():
                    return None
            
            wait = WebDriverWait(self.driver, 10)
            
            # Procurar turma na lista e obter link
            logger.info(f"→ Procurando turma: {turma_nome}")
            html = self.driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            
            # Encontrar linha com nome da turma na coluna correta
            turma_encontrada = False
            turma_id = None
            link_horario = None
            
            # Procurar na tabela
            tbody = soup.find('tbody')
            if tbody:
                for tr in tbody.find_all('tr'):
                    tds = tr.find_all('td', class_='tdatagrid_cell')
                    
                    # A estrutura é: [acao1, acao2, codigo, serie, TURMA, turno, escola, ano]
                    # Índice 4 = coluna "Turma" (0-based: acao1=0, acao2=1, cod=2, serie=3, turma=4)
                    if len(tds) >= 5:
                        coluna_turma = tds[4]  # 5ª coluna (índice 4)
                        texto_turma = coluna_turma.get_text(strip=True)
                        
                        # Comparar com nome procurado
                        if turma_nome.upper() in texto_turma.upper() or texto_turma.upper() in turma_nome.upper():
                            # Extrair ID do href
                            href = coluna_turma.get('href', '')
                            if 'IDTURMA=' in href:
                                import re
                                match = re.search(r'IDTURMA=(\d+)', href)
                                if match:
                                    turma_id = match.group(1)
                                    # Construir URL do formulário de horário
                                    link_horario = f"{self.url_base}/index.php?class=QuadhorariosemanalList&method=onReload&key={turma_id}&IDTURMA={turma_id}"
                                    logger.info(f"  ✓ Turma encontrada: '{texto_turma}' (ID: {turma_id})")
                                    turma_encontrada = True
                                    break
            
            if not turma_encontrada:
                logger.warning(f"  ✗ Turma não encontrada: {turma_nome}")
                # Listar turmas disponíveis para debug
                logger.info("  Turmas disponíveis na página:")
                if tbody:
                    for tr in tbody.find_all('tr')[:5]:  # Primeiras 5 turmas
                        tds = tr.find_all('td', class_='tdatagrid_cell')
                        if len(tds) >= 5:
                            logger.info(f"    - {tds[4].get_text(strip=True)}")
                return None
            
            # Passo 1: Acessar lista de semanas para a turma
            logger.info(f"  → Acessando lista de semanas da turma...")
            self.driver.get(link_horario)
            time.sleep(3)
            
            # Passo 2: Procurar link de edição na lista de semanas
            html = self.driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            
            # Procurar links de edição (ícone de editar ou visualizar)
            link_editar = None
            for a in soup.find_all('a', href=True):
                href = a['href']
                # Procurar link que tenha QuadhorariosemanalForm com onEdit ou onView
                if 'QuadhorariosemanalForm' in href and ('onEdit' in href or 'onView' in href or 'onReload' in href):
                    if 'IDHORARIOSEMANAL' in href or 'key=' in href:
                        link_editar = href
                        if not link_editar.startswith('http'):
                            if link_editar.startswith('index.php'):
                                link_editar = f"{self.url_base}/{link_editar}"
                            else:
                                link_editar = f"{self.url_base}/{link_editar}"
                        logger.info(f"  → Link de edição encontrado: {link_editar[:100]}...")
                        break
            
            # Se não encontrou link na lista, tentar pegar a primeira semana da tabela
            if not link_editar:
                logger.warning("  ⚠ Link de edição não encontrado na lista de semanas")
                # Procurar tabela com semanas
                tbody = soup.find('tbody')
                if tbody:
                    primeira_linha = tbody.find('tr')
                    if primeira_linha:
                        # Procurar link de ação (ícone de editar/visualizar)
                        link_acao = primeira_linha.find('a', href=True)
                        if link_acao:
                            href = link_acao['href']
                            # Este deve ser o link para visualizar os horários daquela semana
                            if not href.startswith('http'):
                                if href.startswith('index.php'):
                                    href = f"{self.url_base}/{href}"
                                else:
                                    href = f"{self.url_base}/{href}"
                            logger.info(f"  → Usando link da primeira semana: {href[:100]}...")
                            link_editar = href
            
            if not link_editar:
                logger.error("  ✗ Não foi possível encontrar link para horários")
                # Salvar HTML para debug
                try:
                    with open('geduc_lista_semanas.html', 'w', encoding='utf-8') as f:
                        f.write(html)
                    logger.info("  → HTML da lista de semanas salvo em geduc_lista_semanas.html")
                except:
                    pass
                return None
            
            # Passo 3: Acessar formulário de horário semanal
            logger.info(f"  → Acessando formulário de horário...")
            self.driver.get(link_editar)
            time.sleep(3)
            
            # Extrair dados da tabela de horários
            html = self.driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            
            # Salvar HTML para debug
            try:
                with open('geduc_horario_formulario.html', 'w', encoding='utf-8') as f:
                    f.write(html)
                logger.info("  → HTML do formulário salvo em geduc_horario_formulario.html")
            except:
                pass
            
            # Encontrar tabela (com border="1px")
            tabela = soup.find('table', border=True)
            if not tabela:
                logger.warning("  ✗ Tabela de horários não encontrada")
                # Tentar encontrar qualquer tabela
                tabelas = soup.find_all('table')
                logger.info(f"  → Encontradas {len(tabelas)} tabelas no total")
                if tabelas:
                    tabela = tabelas[0]
                    logger.info("  → Usando primeira tabela encontrada")
                else:
                    return None
            
            horarios_extraidos = []
            
            # A primeira linha tem os dias da semana
            linhas = tabela.find_all('tr')
            if len(linhas) < 2:
                logger.warning("  ✗ Tabela vazia")
                return None
            
            # Extrair dias da primeira linha
            primeira_linha = linhas[0]
            dias_semana = []
            for td in primeira_linha.find_all('td'):
                dia = td.get_text(strip=True)
                if dia:
                    dias_semana.append(dia)
            
            logger.info(f"  → Dias encontrados: {dias_semana}")
            
            # Debug: mostrar número de linhas na tabela
            logger.info(f"  → Total de linhas na tabela: {len(linhas)}")
            
            # Mapear dias do GEDUC para nomes padrão
            mapa_dias = {
                'Domingo': 'Domingo',
                'Segunda': 'Segunda',
                'Terça': 'Terça',
                'Quarta': 'Quarta',
                'Quinta': 'Quinta',
                'Sexta': 'Sexta',
                'Sábado': 'Sábado'
            }
            
            # Processar linhas de horário (índices 1 em diante)
            for idx_linha, linha in enumerate(linhas[1:], 1):
                celulas = linha.find_all('td')
                logger.info(f"  → Linha {idx_linha}: {len(celulas)} células")
                
                if not celulas or len(celulas) != 7:
                    logger.warning(f"    ⚠ Linha {idx_linha} pulada: esperadas 7 células, encontradas {len(celulas)}")
                    continue
                
                # Para cada dia (célula)
                for idx_dia, celula in enumerate(celulas):
                    if idx_dia >= len(dias_semana):
                        break
                    
                    dia_original = dias_semana[idx_dia]
                    dia = mapa_dias.get(dia_original, dia_original)
                    
                    # Ignorar Domingo e Sábado
                    if dia in ['Domingo', 'Sábado']:
                        continue
                    
                    # Tentar extrair disciplina
                    disciplina = None
                    
                    # 1. Link (disciplina já cadastrada)
                    link = celula.find('a')
                    if link:
                        disciplina = link.get_text(strip=True)
                        logger.info(f"    → Linha {idx_linha}, Dia {dia}: encontrado link com '{disciplina}'")
                        
                        # Se encontrou disciplina válida, adicionar
                        if disciplina and disciplina not in ['', '-', 'Intervalo']:
                            horarios_extraidos.append({
                                'dia': dia,
                                'horario': f"Linha {idx_linha}",  # Usar índice da linha
                                'disciplina': disciplina,
                                'professor': None
                            })
                            logger.info(f"    ✓ Horário adicionado: {dia} - {disciplina}")
                    else:
                        # Apenas loga se não tem link
                        logger.info(f"    → Linha {idx_linha}, Dia {dia}: sem disciplina cadastrada")
            
            logger.info(f"  ✓ Extraídos {len(horarios_extraidos)} horários")
            
            return {
                'turma_nome': turma_nome,
                'turma_id': int(turma_id) if turma_id and str(turma_id).isdigit() else None,
                'horarios': horarios_extraidos,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
        except Exception as e:
            logger.exception("✗ Erro ao extrair horário: %s", e)
            return None
    
    def listar_turmas_disponiveis(self) -> list:
        """
        Lista todas as turmas disponíveis na página de lista de horários
        
        Returns:
            Lista de dicts com 'id' e 'nome' das turmas
        """
        try:
            assert self.driver is not None, "navegador não iniciado"
            
            # Acessar lista de horários se ainda não estiver nela
            if "TurmaHorariosList" not in self.driver.current_url:
                if not self.acessar_lista_horarios():
                    return []
            
            # Extrair turmas da tabela
            html = self.driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            
            turmas = []
            for td in soup.find_all('td', class_='tdatagrid_cell'):
                href = td.get('href', '')
                if 'IDTURMA=' in href:
                    import re
                    match = re.search(r'IDTURMA=(\d+)', href)
                    if match:
                        turma_id = match.group(1)
                        turma_nome = td.get_text(strip=True)
                        
                        # Evitar duplicatas
                        if not any(t['id'] == turma_id for t in turmas):
                            turmas.append({
                                'id': turma_id,
                                'nome': turma_nome
                            })
            
            logger.info(f"✓ Encontradas {len(turmas)} turmas disponíveis")
            return turmas
            
        except Exception as e:
            logger.exception("✗ Erro ao listar turmas: %s", e)
            return []

    # =========================================================================
    # GERAÇÃO DE DECLARAÇÕES NO GEDUC
    # =========================================================================

    def _obter_serie_id_geduc(self, nome_serie: str) -> Optional[str]:
        """
        Converte o nome da série local para o IDCURSOORI correspondente no GEduc.
        Ex: '9º Ano' ou '9º ANO' → '16'
        """
        import re as _re
        match = _re.search(r'(\d+)', nome_serie)
        if match:
            return self._GEDUC_SERIE_IDS.get(match.group(1))
        return None

    def _construir_texto_declaracao_html(
        self,
        nome_aluno: str,
        sexo: str,
        data_nasc_formatada: str,
        naturalidade: str,
        matricula: str,
        responsaveis: list,
        nome_escola: str,
        municipio_escola: str,
        nome_serie: str,
        nivel_ensino: str,
        turno: str,
        tipo_declaracao: str,
        motivo_outros: str = '',
        ano_letivo: int = 2026,
        data_documento: str = '',
    ) -> str:
        """
        Constrói o HTML da declaração com texto correto para injetar no Summernote.

        Corrige em relação ao texto padrão do GEduc:
        - Concordância de gênero (nascido/nascida, filho/filha, matriculado/matriculada)
        - Filiação com 0, 1 ou 2 responsáveis
        - Bloco "Declaração para fins de:" com tipo (Bolsa Família, Trabalho, Outros)
        - Assinatura como GESTOR(A) em vez do nome da gestora
        - Data e município corretos
        """
        gen = 'a' if (sexo or 'M').upper() == 'F' else 'o'

        # --- Filiação (dados em CAIXA ALTA, sem negrito) ---
        resp1 = responsaveis[0][0] if len(responsaveis) > 0 and responsaveis[0] else None
        resp2 = responsaveis[1][0] if len(responsaveis) > 1 and responsaveis[1] else None
        _r1 = resp1.upper() if resp1 else None
        _r2 = resp2.upper() if resp2 else None
        if _r1 and _r2:
            filiacao = f"filh{gen} de {_r1} e {_r2}"
        elif _r1:
            filiacao = f"filh{gen} de {_r1}"
        elif _r2:
            filiacao = f"filh{gen} de {_r2}"
        else:
            filiacao = f"filh{gen} de pais não cadastrados"

        # --- Turno ---
        turno_upper = (turno or '').upper()
        if 'MAT' in turno_upper:
            turno_texto = 'Matutino'
        elif 'VES' in turno_upper or 'VESP' in turno_upper:
            turno_texto = 'Vespertino'
        else:
            turno_texto = turno.capitalize() if turno else ''

        # --- Tipo de declaração (caixa alta, sem negrito) ---
        if tipo_declaracao == 'Outros' and motivo_outros:
            tipo_texto = f'OUTROS: {motivo_outros.upper()}'
        else:
            tipo_texto = tipo_declaracao.upper()

        # --- Complementos opcionais (caixa alta, sem negrito) ---
        natural_bloco = (
            f', natural de {naturalidade.upper()}'
            if naturalidade else ''
        )
        matricula_bloco = (
            f', com matrícula {matricula}'
            if matricula else ''
        )

        # --- Data do documento ---
        if not data_documento:
            from src.utils.dates import formatar_data_extenso as _formatar_data
            import datetime as _dt
            data_documento = _formatar_data(_dt.datetime.now())

        municipio_doc = municipio_escola if municipio_escola else 'PAÇO DO LUMIAR / MA'

        # {QRCODE} é um token processado pelo servidor GEduc ao gerar o PDF.
        # Deve ser mantido no final do HTML para o QR code de autenticação aparecer.
        _s = 'font-family:Arial;font-size:12pt'
        html = (
            '<p><br></p>'
            '<p style="font-family:Arial;font-size:18pt;text-align:center"><strong>DECLARAÇÃO</strong></p>'
            '<p><br></p>'
            '<p><br></p>'
            f'<p style="{_s};text-align:justify">'
            f'Declaro para os devidos fins de direito, que {nome_aluno.upper()}'
            f'{matricula_bloco}, nascid{gen} no dia {data_nasc_formatada}'
            f'{natural_bloco}, {filiacao}, está regularmente matriculad{gen} na '
            f'{nome_escola.upper()}, cursando o {nome_serie.upper()} '
            f'DO {nivel_ensino.upper()}, no ano letivo de {ano_letivo}, '
            f'no turno {turno_texto}.'
            '</p>'
            f'<p style="{_s};text-align:justify">'
            'Situação Acadêmica: Cursando'
            '</p>'
            '<p><br></p>'
            f'<p style="{_s};text-align:justify">'
            f'Declaração para fins de: {tipo_texto}'
            '</p>'
            f'<p style="{_s};text-align:justify">'
            'Por ser a expressão da verdade, dato e assino a presente declaração, '
            'para que surta os devidos efeitos legais.'
            '</p>'
            '<p><br></p>'
            f'<p style="{_s};text-align:right">{municipio_doc}, {data_documento}</p>'
            '<p><br></p>'
            f'<p style="{_s};text-align:center">___________________________________</p>'
            f'<p style="{_s};text-align:center">GESTOR(A)</p>'
            '<p><br></p>'
            '<p><br></p>'
            '<p>{QRCODE}</p>'
        )
        return html

    def gerar_declaracao(
        self,
        aluno_id: int,
        tipo_declaracao: str = 'Bolsa Família',
        motivo_outros: str = '',
    ) -> bool:
        """
        Gera uma declaração autenticada no GEduc para o aluno especificado.

        Fluxo:
          1. Consulta dados do aluno no BD local (nome, sexo, série, turma, responsáveis)
          2. Navega para DeclaracaoForm no GEduc
          3. Seleciona série → turma → aluno → modelo "Declaração Escolar"
          4. Clica "Carregar Dados" para obter o texto padrão do GEduc
          5. Extrai matrícula e naturalidade do texto carregado
          6. Substitui o conteúdo do editor Summernote com texto corrigido
          7. Clica "Gerar Declaração" para obter o PDF autenticado com QR code

        Args:
            aluno_id:        ID do aluno no banco local
            tipo_declaracao: 'Bolsa Família' | 'Trabalho' | 'Outros'
                             (Transferência não é suportada via este fluxo)
            motivo_outros:   Descrição do motivo quando tipo_declaracao='Outros'

        Returns:
            True se a declaração foi enviada ao GEduc com sucesso
        """
        if self.driver is None:
            logger.error(
                "✗ Navegador não iniciado. Chame iniciar_navegador() e fazer_login() primeiro."
            )
            return False

        import re as _re
        import json as _json
        import datetime as _dt
        import pandas as _pd
        from src.core.conexao import conectar_bd
        from src.core.config import ANO_LETIVO_ATUAL
        from src.relatorios.declaracao_aluno import (
            obter_dados_escola,
            obter_dados_aluno,
            obter_responsaveis,
        )
        from src.utils.dates import formatar_data_extenso as formatar_data

        # ── 1. Buscar dados do aluno no BD local ──────────────────────────
        logger.info("→ Consultando dados do aluno #%s no banco local...", aluno_id)
        conn = conectar_bd()
        if conn is None:
            logger.error("✗ Não foi possível conectar ao banco de dados.")
            return False
        cursor = conn.cursor()
        try:
            dados_escola = obter_dados_escola(cursor, 60)
            dados_aluno = obter_dados_aluno(cursor, aluno_id)
            responsaveis = obter_responsaveis(cursor, aluno_id)
        finally:
            cursor.close()
            conn.close()

        if not dados_aluno:
            logger.error("✗ Aluno #%s não encontrado no banco.", aluno_id)
            return False

        nome_aluno, nascimento, sexo, nome_serie, nome_turma, turno, nivel_ensino = dados_aluno
        nome_escola = dados_escola[1] if dados_escola else "ESCOLA"
        municipio_escola = dados_escola[5] if dados_escola else "PAÇO DO LUMIAR / MA"

        data_nasc_formatada = (
            _pd.to_datetime(nascimento).strftime("%d/%m/%Y")
            if nascimento and _pd.notnull(nascimento) else ""
        )
        serie_turma_str = f"{nome_serie} {nome_turma}".strip() if nome_serie else ""
        nivel_str = nivel_ensino or "ENSINO FUNDAMENTAL"

        logger.info(
            "  ✓ Aluno: %s | Série: %s | Turma: %s | Sexo: %s",
            nome_aluno, nome_serie, nome_turma, sexo
        )

        # ── 2. Navegar para DeclaracaoForm ────────────────────────────────
        logger.info("→ Acessando formulário de declaração no GEduc...")
        assert self.driver is not None
        self.driver.get(f"{self.url_base}/index.php?class=DeclaracaoForm")
        wait = WebDriverWait(self.driver, 20)
        try:
            wait.until(EC.presence_of_element_located((By.NAME, "IDCURSOORI")))
        except TimeoutException:
            logger.error("✗ Formulário DeclaracaoForm não carregou.")
            return False
        time.sleep(1)

        # ── 3a. Selecionar Série ──────────────────────────────────────────
        id_geduc_serie = self._obter_serie_id_geduc(nome_serie or "")
        if not id_geduc_serie:
            logger.error(
                "✗ Não foi possível mapear a série '%s' para um ID GEduc. "
                "Verifique _GEDUC_SERIE_IDS.",
                nome_serie,
            )
            return False

        logger.info("→ Selecionando série '%s' (ID GEduc: %s)...", nome_serie, id_geduc_serie)
        try:
            Select(self.driver.find_element(By.NAME, "IDCURSOORI")).select_by_value(id_geduc_serie)
        except Exception as e:
            logger.exception("✗ Erro ao selecionar série: %s", e)
            return False

        # Aguardar AJAX carregar as turmas
        logger.info("→ Aguardando carregamento das turmas via AJAX...")
        try:
            wait.until(
                lambda d: len(Select(d.find_element(By.NAME, "IDTURMAORI")).options) > 1
            )
        except TimeoutException:
            logger.warning("⚠ Timeout aguardando turmas. Continuando mesmo assim...")
        time.sleep(0.8)

        # ── 3b. Selecionar Turma ──────────────────────────────────────────
        # Montar nome esperado no GEduc: "{serie}{turma}-{turno}" ou "{serie}-{turno}"
        # Exemplos: "8º ANO A-VESP", "8º ANO-VESP", "1º ANO B-MAT"
        _turno_upper = (turno or '').upper()
        _sufixo_turno = 'VESP' if 'VES' in _turno_upper else 'MAT'
        _serie_upper = (nome_serie or '').upper()
        if nome_turma and nome_turma.strip():
            _nome_turma_geduc = f"{_serie_upper} {nome_turma.strip().upper()}-{_sufixo_turno}"
        else:
            _nome_turma_geduc = f"{_serie_upper}-{_sufixo_turno}"

        logger.info("→ Selecionando turma (esperado: '%s')...", _nome_turma_geduc)
        try:
            select_turma = Select(self.driver.find_element(By.NAME, "IDTURMAORI"))
            # Opções reais = opções com value não-vazio (exclui placeholder)
            opcoes_reais = [
                o for o in select_turma.options if o.get_attribute("value")
            ]
            turma_selecionada = False

            # Normaliza espaços ao redor do traço para comparação tolerante
            # ex: "7º ANO A -VESP" e "7º ANO A-VESP" → ambos viram "7º ANO A-VESP"
            import re as _re_local
            def _normalizar(s: str) -> str:
                return _re_local.sub(r'\s*-\s*', '-', s.strip().upper())

            _nome_esperado_norm = _normalizar(_nome_turma_geduc)

            # Tentativa 1: match normalizado
            for option in opcoes_reais:
                if _nome_esperado_norm in _normalizar(option.text):
                    select_turma.select_by_value(option.get_attribute("value"))
                    turma_selecionada = True
                    logger.info("  ✓ Turma selecionada por nome: %s", option.text.strip())
                    break

            # Tentativa 2: turma vazia no banco local + só uma turma no GEduc → seleciona automaticamente
            if not turma_selecionada and len(opcoes_reais) == 1:
                option = opcoes_reais[0]
                select_turma.select_by_value(option.get_attribute("value"))
                turma_selecionada = True
                logger.info(
                    "  ✓ Turma selecionada automaticamente (única disponível): %s",
                    option.text.strip()
                )

            if not turma_selecionada:
                nomes_opcoes = [o.text.strip() for o in opcoes_reais]
                logger.error(
                    "✗ Turma '%s' não encontrada. Opções disponíveis: %s",
                    _nome_turma_geduc, nomes_opcoes
                )
                return False
        except Exception as e:
            logger.exception("✗ Erro ao selecionar turma: %s", e)
            return False

        # Aguardar AJAX carregar alunos
        logger.info("→ Aguardando carregamento dos alunos via AJAX...")
        try:
            wait.until(
                lambda d: len(Select(d.find_element(By.NAME, "IDALUNO")).options) > 1
            )
        except TimeoutException:
            logger.warning("⚠ Timeout aguardando alunos. Continuando mesmo assim...")
        time.sleep(0.8)

        # ── 3c. Selecionar Aluno (Select2) ────────────────────────────────
        logger.info("→ Selecionando aluno '%s'...", nome_aluno)
        aluno_selecionado = False

        import unicodedata as _ucd

        def _sem_acento(s: str) -> str:
            """Remove acentos e converte para maiúsculas — GEduc não usa acentuação."""
            return ''.join(
                c for c in _ucd.normalize('NFD', (s or '').upper())
                if _ucd.category(c) != 'Mn'
            )

        nome_busca = _sem_acento(nome_aluno)
        # Primeiro nome sem acento para digitar no campo de busca do Select2
        primeiro_nome_busca = _sem_acento(nome_aluno.split()[0]) if nome_aluno else ""

        # Tentativa 1: pelo select nativo (quando já carregado)
        try:
            select_aluno_elem = self.driver.find_element(By.NAME, "IDALUNO")
            select_aluno = Select(select_aluno_elem)
            for option in select_aluno.options:
                if option.get_attribute("value") and nome_busca in _sem_acento(option.text):
                    select_aluno.select_by_value(option.get_attribute("value"))
                    # Disparar evento change para o Select2 reconhecer a seleção
                    self.driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));",
                        select_aluno_elem,
                    )
                    aluno_selecionado = True
                    logger.info("  ✓ Aluno selecionado (select nativo): %s", option.text.strip())
                    break
        except Exception as e:
            logger.warning("  ⚠ Seleção nativa falhou: %s", e)

        # Tentativa 2: via Select2 (busca por texto)
        if not aluno_selecionado:
            try:
                select2_container = self.driver.find_element(By.CSS_SELECTOR, ".select2-container")
                select2_container.click()
                time.sleep(0.5)
                search_input = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".select2-search__field"))
                )
                # Digitar o primeiro nome (sem acento) para filtrar
                search_input.send_keys(primeiro_nome_busca)
                time.sleep(1.5)
                resultados = self.driver.find_elements(
                    By.CSS_SELECTOR, ".select2-results__option"
                )
                for resultado in resultados:
                    if nome_busca in _sem_acento(resultado.text):
                        resultado.click()
                        aluno_selecionado = True
                        logger.info("  ✓ Aluno selecionado (Select2): %s", resultado.text)
                        break
            except Exception as e:
                logger.exception("  ✗ Seleção via Select2 falhou: %s", e)

        if not aluno_selecionado:
            logger.error("✗ Não foi possível selecionar o aluno '%s'.", nome_aluno)
            return False

        # ── 3d. Selecionar modelo "Declaração Escolar" (value=4) ──────────
        logger.info("→ Selecionando modelo 'DECLARAÇÃO ESCOLAR'...")
        try:
            Select(self.driver.find_element(By.NAME, "IDMODELODOC")).select_by_value("4")
            logger.info("  ✓ Modelo selecionado")
        except Exception as e:
            logger.exception("✗ Erro ao selecionar modelo: %s", e)
            return False

        # ── 4. Clicar "Carregar Dados" ────────────────────────────────────
        logger.info("→ Clicando em 'Carregar Dados'...")
        try:
            btn_carregar = self.driver.find_element(By.NAME, "BtnCarrAlunos")
            self.driver.execute_script("arguments[0].click();", btn_carregar)
        except Exception as e:
            logger.exception("✗ Botão 'Carregar Dados' não encontrado: %s", e)
            return False

        # Aguardar o editor Summernote ser preenchido (pelo menos 50 chars)
        logger.info("→ Aguardando preenchimento do editor Summernote...")
        try:
            wait.until(lambda d: len(
                d.execute_script(
                    "return document.querySelector('.note-editable') "
                    "&& document.querySelector('.note-editable').innerText || '';"
                ).strip()
            ) > 50)
        except TimeoutException:
            logger.warning("⚠ Timeout aguardando editor. Verificando conteúdo mesmo assim...")
        time.sleep(2)

        # ── 5. Extrair matrícula e naturalidade do texto do GEduc ─────────
        logger.info("→ Extraindo matrícula e naturalidade do texto gerado pelo GEduc...")
        texto_geduc = ""
        try:
            texto_geduc = self.driver.execute_script(
                "var el = document.querySelector('.note-editable');"
                "return el ? el.innerText : '';"
            ) or ""
        except Exception as e:
            logger.warning("⚠ Não foi possível ler o texto do GEduc: %s", e)

        matricula_str = ""
        match_mat = _re.search(r'matr[íi]cula[^\d]*(\d+)', texto_geduc, _re.IGNORECASE)
        if match_mat:
            matricula_str = match_mat.group(1)
            logger.info("  ✓ Matrícula GEduc extraída: %s", matricula_str)

        naturalidade_str = ""
        match_nat = _re.search(r'natural de ([^,\n<]+)', texto_geduc, _re.IGNORECASE)
        if match_nat:
            naturalidade_str = match_nat.group(1).strip()
            logger.info("  ✓ Naturalidade extraída: %s", naturalidade_str)

        # ── 6. Construir e injetar texto corrigido no Summernote ──────────
        logger.info("→ Construindo HTML da declaração corrigida...")

        # Preservar o cabeçalho original do GEduc (tabela com logotipo e nome da escola)
        html_atual = ""
        try:
            html_atual = self.driver.execute_script(
                "return $('[name=\"TEXTO\"]').summernote('code') || '';"
            ) or ""
        except Exception as _e:
            logger.warning("⚠ Não foi possível extrair HTML atual do editor: %s", _e)

        # Ponto de corte: fim da última </table> (contém imagem e cabeçalho)
        prefixo_html = ""
        _match_tabela = list(_re.finditer(r'</table>', html_atual, _re.IGNORECASE))
        if _match_tabela:
            prefixo_html = html_atual[:_match_tabela[-1].end()]
            logger.info("  ✓ Cabeçalho preservado (%d chars)", len(prefixo_html))
        else:
            # Fallback: tudo antes do parágrafo que inicia com DECLARAÇÃO
            _match_decl = _re.search(
                r'<p[^>]*>(?:<[^>]+>)*\s*DECLARA', html_atual, _re.IGNORECASE
            )
            if _match_decl:
                prefixo_html = html_atual[:_match_decl.start()]
                logger.info("  ✓ Cabeçalho preservado via fallback (%d chars)", len(prefixo_html))

        html_declaracao = self._construir_texto_declaracao_html(
            nome_aluno=nome_aluno or "",
            sexo=sexo or "M",
            data_nasc_formatada=data_nasc_formatada,
            naturalidade=naturalidade_str,
            matricula=matricula_str,
            responsaveis=list(responsaveis),
            nome_escola=nome_escola,
            municipio_escola=municipio_escola,
            nome_serie=nome_serie or "",
            nivel_ensino=nivel_str,
            turno=turno or "",
            tipo_declaracao=tipo_declaracao,
            motivo_outros=motivo_outros,
            ano_letivo=ANO_LETIVO_ATUAL,
            data_documento=formatar_data(_dt.datetime.now()),
        )

        html_completo = prefixo_html + html_declaracao

        logger.info("→ Injetando texto corrigido no Summernote...")
        try:
            # json.dumps garante escape correto de aspas e caracteres especiais
            html_json = _json.dumps(html_completo)
            self.driver.execute_script(
                f"$('[name=\"TEXTO\"]').summernote('code', {html_json});"
            )
            time.sleep(0.5)
            logger.info("  ✓ Texto injetado com sucesso")
        except Exception as e:
            logger.exception("✗ Erro ao injetar texto no Summernote: %s", e)
            return False

        # ── 7. Clicar "Gerar Declaração" ──────────────────────────────────
        logger.info("→ Clicando em 'Gerar Declaração'...")
        try:
            # Tenta localizar pelo nome (pode conter ã); fallback pelo texto do botão
            btn_gerar = None
            try:
                btn_gerar = self.driver.find_element(By.NAME, "btn_gerar_declaração")
            except NoSuchElementException:
                botoes = self.driver.find_elements(By.TAG_NAME, "button")
                for b in botoes:
                    if "GERAR DECLARA" in (b.text or "").upper():
                        btn_gerar = b
                        break

            if btn_gerar is None:
                logger.error("✗ Botão 'Gerar Declaração' não encontrado na página.")
                return False

            self.driver.execute_script("arguments[0].click();", btn_gerar)
            logger.info("  ✓ Declaração enviada ao GEduc para geração!")
            time.sleep(3)
            return True

        except Exception as e:
            logger.exception("✗ Erro ao clicar em 'Gerar Declaração': %s", e)
            return False

    def fechar(self):
        """
        Fecha o navegador
        """
        if self.driver:
            self.driver.quit()
            logger.info("✓ Navegador fechado")


def interface_automacao():
    """
    Interface gráfica para automação de extração de notas
    """
    root = tk.Tk()
    root.title("Automação de Extração de Notas - GEDUC")
    root.geometry("700x600")
    root.configure(bg="#F5F5F5")
    
    # Centralizar janela
    root.update_idletasks()
    width = 700
    height = 600
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    # Cores
    co0 = "#F5F5F5"
    co1 = "#003A70"
    co2 = "#77B341"
    co4 = "#4A86E8"
    
    # Variáveis
    usuario_var = tk.StringVar()
    senha_var = tk.StringVar()
    headless_var = tk.BooleanVar(value=False)
    timeout_var = tk.IntVar(value=120)  # 2 minutos por padrão
    bim1_var = tk.BooleanVar(value=True)
    bim2_var = tk.BooleanVar(value=True)
    bim3_var = tk.BooleanVar(value=True)
    bim4_var = tk.BooleanVar(value=True)
    
    # Frame principal
    frame_principal = tk.Frame(root, bg=co0)
    frame_principal.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
    
    # Título
    titulo = tk.Label(
        frame_principal,
        text="Automação de Extração de Notas GEDUC",
        font=("Arial", 16, "bold"),
        bg=co0,
        fg=co1
    )
    titulo.pack(pady=(0, 10))
    
    # Aviso sobre reCAPTCHA
    aviso_recaptcha = tk.Label(
        frame_principal,
        text="⚠️ IMPORTANTE: Você precisará resolver o reCAPTCHA manualmente no navegador",
        font=("Arial", 9, "italic"),
        bg=co0,
        fg="#E65100"
    )
    aviso_recaptcha.pack(pady=(0, 15))
    
    # Credenciais
    frame_credenciais = tk.LabelFrame(
        frame_principal,
        text="Credenciais de Acesso",
        font=("Arial", 11, "bold"),
        bg=co0,
        fg=co1,
        padx=15,
        pady=10
    )
    frame_credenciais.pack(fill=tk.X, pady=10)
    
    # Usuário
    tk.Label(frame_credenciais, text="Usuário:", font=("Arial", 10), bg=co0).grid(row=0, column=0, sticky="w", pady=5)
    entry_usuario = tk.Entry(frame_credenciais, textvariable=usuario_var, font=("Arial", 10), width=40)
    entry_usuario.grid(row=0, column=1, pady=5, padx=(10, 0))
    
    # Senha
    tk.Label(frame_credenciais, text="Senha:", font=("Arial", 10), bg=co0).grid(row=1, column=0, sticky="w", pady=5)
    entry_senha = tk.Entry(frame_credenciais, textvariable=senha_var, font=("Arial", 10), width=40, show="*")
    entry_senha.grid(row=1, column=1, pady=5, padx=(10, 0))
    
    # Opções
    frame_opcoes = tk.LabelFrame(
        frame_principal,
        text="Opções de Extração",
        font=("Arial", 11, "bold"),
        bg=co0,
        fg=co1,
        padx=15,
        pady=10
    )
    frame_opcoes.pack(fill=tk.X, pady=10)
    
    # Bimestres
    tk.Label(frame_opcoes, text="Bimestres a extrair:", font=("Arial", 10, "bold"), bg=co0).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 5))
    
    tk.Checkbutton(frame_opcoes, text="1º Bimestre", variable=bim1_var, font=("Arial", 9), bg=co0).grid(row=1, column=0, sticky="w", padx=(20, 0))
    tk.Checkbutton(frame_opcoes, text="2º Bimestre", variable=bim2_var, font=("Arial", 9), bg=co0).grid(row=1, column=1, sticky="w")
    tk.Checkbutton(frame_opcoes, text="3º Bimestre", variable=bim3_var, font=("Arial", 9), bg=co0).grid(row=1, column=2, sticky="w")
    tk.Checkbutton(frame_opcoes, text="4º Bimestre", variable=bim4_var, font=("Arial", 9), bg=co0).grid(row=1, column=3, sticky="w")
    
    # Modo headless
    aviso_headless = tk.Label(
        frame_opcoes,
        text="⚠️ Não recomendado: você precisa ver o reCAPTCHA",
        font=("Arial", 8, "italic"),
        bg=co0,
        fg="#666666"
    )
    aviso_headless.grid(row=3, column=0, columnspan=4, sticky="w", pady=(0, 0), padx=(20, 0))
    
    # Timeout do reCAPTCHA
    tk.Label(
        frame_opcoes,
        text="Tempo para resolver reCAPTCHA (segundos):",
        font=("Arial", 9),
        bg=co0
    ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(10, 5))
    
    timeout_spinbox = tk.Spinbox(
        frame_opcoes,
        from_=30,
        to=300,
        textvariable=timeout_var,
        font=("Arial", 9),
        width=10
    )
    timeout_spinbox.grid(row=4, column=2, sticky="w", pady=(10, 5))
    
    tk.Checkbutton(
        frame_opcoes,
        text="Modo silencioso (sem abrir navegador)",
        variable=headless_var,
        font=("Arial", 9),
        bg=co0
    ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(10, 0))
    
    # Log de progresso
    frame_log = tk.LabelFrame(
        frame_principal,
        text="Progresso",
        font=("Arial", 11, "bold"),
        bg=co0,
        fg=co1,
        padx=10,
        pady=10
    )
    frame_log.pack(fill=tk.BOTH, expand=True, pady=10)
    
    # Barra de progresso
    progress_bar = ttk.Progressbar(frame_log, mode='indeterminate')
    progress_bar.pack(fill=tk.X, pady=(0, 10))
    
    # Texto de log
    text_log = tk.Text(frame_log, height=10, font=("Consolas", 9), bg="white", fg="black")
    text_log.pack(fill=tk.BOTH, expand=True)
    
    scrollbar = tk.Scrollbar(frame_log, command=text_log.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    text_log.config(yscrollcommand=scrollbar.set)
    
    def adicionar_log(mensagem):
        """Adiciona mensagem ao log"""
        try:
            # Agendar atualização no main thread via after para garantir thread-safety
            try:
                text_log.after(0, lambda m=mensagem: (text_log.insert(tk.END, f"{m}\n"), text_log.see(tk.END)))
            except Exception:
                # Fallback: tentar inserir diretamente (último recurso)
                text_log.insert(tk.END, f"{mensagem}\n")
                text_log.see(tk.END)
        except Exception:
            # Silenciar erros de log para não interromper a automação
            pass
    
    def iniciar_extracao():
        """Inicia o processo de extração"""
        usuario = usuario_var.get().strip()
        senha = senha_var.get().strip()
        
        if not usuario or not senha:
            messagebox.showerror("Erro", "Preencha usuário e senha!", parent=root)
            return
        
        # Obter bimestres selecionados
        bimestres = []
        if bim1_var.get(): bimestres.append(1)
        if bim2_var.get(): bimestres.append(2)
        if bim3_var.get(): bimestres.append(3)
        if bim4_var.get(): bimestres.append(4)
        
        if not bimestres:
            messagebox.showerror("Erro", "Selecione pelo menos um bimestre!", parent=root)
            return
        
        # Desabilitar botão
        btn_iniciar.config(state=tk.DISABLED)
        progress_bar.start()
        text_log.delete(1.0, tk.END)
        
        # Executar em thread separada
        import threading
        
        def executar():
            automacao = None
            try:
                adicionar_log("="*60)
                adicionar_log("INICIANDO AUTOMAÇÃO DE EXTRAÇÃO DE NOTAS")
                adicionar_log("="*60)
                
                # Criar instância
                automacao = AutomacaoGEDUC(headless=headless_var.get())
                
                # Iniciar navegador
                adicionar_log("\n→ Iniciando navegador...")
                if not automacao.iniciar_navegador():
                    adicionar_log("✗ Erro ao iniciar navegador")
                    return
                adicionar_log("✓ Navegador iniciado")
                
                # Fazer login
                adicionar_log("\n→ Fazendo login no GEDUC...")
                adicionar_log("⚠️  ATENÇÃO: Você precisará resolver o reCAPTCHA manualmente!")
                adicionar_log("   Marque a caixa 'Não sou um robô' no navegador")
                adicionar_log("   Depois clique no botão LOGIN")
                adicionar_log("")
                
                timeout_recaptcha = timeout_var.get()
                if not automacao.fazer_login(usuario, senha, timeout_recaptcha=timeout_recaptcha):
                    adicionar_log("✗ Login falhou - verifique usuário e senha")
                    adicionar_log("   Ou o tempo para resolver o reCAPTCHA expirou")
                    return
                adicionar_log("✓ Login realizado com sucesso")
                
                # Extrair notas
                adicionar_log(f"\n→ Extraindo notas dos bimestres: {', '.join([f'{b}º' for b in bimestres])}")
                adicionar_log("   (Os arquivos serão salvos automaticamente após cada turma/bimestre)")
                adicionar_log("   (Este processo pode levar alguns minutos...)\n")
                
                def callback_progresso(processadas, total):
                    adicionar_log(f"   Progresso: {processadas}/{total} planilhas processadas")
                
                if not automacao.extrair_todas_notas(
                    bimestres=bimestres, 
                    diretorio_saida="notas_extraidas",
                    callback_progresso=callback_progresso
                ):
                    adicionar_log("✗ Erro durante extração")
                    return
                
                # Mensagem final
                adicionar_log("\n" + "="*60)
                adicionar_log("EXTRAÇÃO CONCLUÍDA COM SUCESSO!")
                adicionar_log(f"Arquivos salvos em: {os.path.abspath('notas_extraidas')}")
                adicionar_log("="*60)
                
                messagebox.showinfo(
                    "Sucesso",
                    f"Extração concluída!\n\n"
                    f"Os arquivos Excel foram salvos em:\n"
                    f"notas_extraidas/\n\n"
                    f"Formato: 1 arquivo por turma/bimestre\n"
                    f"Cada arquivo contém múltiplas planilhas (uma por disciplina)",
                    parent=root
                )
                
            except Exception as e:
                adicionar_log(f"\n✗ ERRO: {e}")
                import traceback
                adicionar_log(traceback.format_exc())
                messagebox.showerror("Erro", f"Erro durante extração:\n{e}", parent=root)
            
            finally:
                if automacao:
                    adicionar_log("\n→ Fechando navegador...")
                    automacao.fechar()
                
                progress_bar.stop()
                btn_iniciar.config(state=tk.NORMAL)
        
        thread = threading.Thread(target=executar, daemon=True)
        thread.start()
    
    # Botão iniciar
    btn_iniciar = tk.Button(
        frame_principal,
        text="INICIAR EXTRAÇÃO AUTOMÁTICA",
        command=iniciar_extracao,
        bg=co2,
        fg="white",
        font=("Arial", 12, "bold"),
        relief=tk.RAISED,
        cursor="hand2",
        height=2
    )
    btn_iniciar.pack(pady=10, fill=tk.X)
    
    root.mainloop()


if __name__ == "__main__":
    interface_automacao()
