import openpyxl
from pathlib import Path
p = Path(r"C:/gestao/teste_pendencias.xlsx")
if not p.exists():
    print(f"Arquivo não encontrado: {p}")
    raise SystemExit(1)
wb = openpyxl.load_workbook(p, data_only=True)
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\n' + '='*40)
    print(f"Sheet: {name}")
    # Column widths explicitly set in column_dimensions
    widths = {k: (v.width if getattr(v, 'width', None) is not None else None) for k,v in ws.column_dimensions.items()}
    print("Column widths (defined):", widths)
    # Freeze panes
    try:
        print("freeze_panes:", ws.freeze_panes)
    except Exception as e:
        print("freeze_panes: (error)", e)
    # merged cells
    merged = [str(r) for r in getattr(ws, 'merged_cells').ranges] if getattr(ws, 'merged_cells', None) is not None else []
    print("merged (first 20):", merged[:20])
    # header sample (first 10 columns)
    header = [ws.cell(row=1, column=i).value for i in range(1, min(12, ws.max_column+1))]
    print("Header sample:", header)
    print("max_row, max_col:", ws.max_row, ws.max_column)

# Basic summary: column widths presence across sheets
print('\n' + '='*40)
for name in wb.sheetnames:
    ws = wb[name]
    defined = bool(ws.column_dimensions)
    print(f"{name}: column_dimensions present: {defined}")
wb.close()
print('\nInspeção finalizada')
