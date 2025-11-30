#!/usr/bin/env python3
"""
Importador BNCC multi-sheet com modo dry-run.
Uso:
  python scripts/import_bncc_multi_sheet.py --file C:/gestao/MapasDeFocoBncc_Unificados.xlsx --dry-run

Opções de DB (quando não `--dry-run`): --user, --database, --host, --port

O script varre todas as planilhas (ou apenas a informada em --sheet), detecta colunas
comuns e extrai `codigo` e `descricao` obrigatórios. Em `--dry-run` imprime amostras
parseadas sem tocar no banco.
"""
import argparse
from pathlib import Path
from openpyxl import load_workbook
import sys
# garantir que o diretório do repositório esteja em sys.path para importar bncc_parser
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from bncc_parser import parse_bncc_code
import subprocess

CANDIDATES = {
    'codigo': ['código da habilidade', 'codigo da habilidade', 'código', 'codigo', 'codigo_habilidade', 'code', 'cod'],
    'descricao': ['texto da habilidade', 'texto', 'descrição', 'descricao', 'habilidade', 'descrição da habilidade'],
    'unidade': ['unidade temática','unidade','unidade temática'],
    'conhecimento_previo': ['conhecimento prévio','conhecimento previo','conhecimento','prévio'],
    'classificacao': ['classificação','classificacao'],
    'competencias': ['competências relacionadas','competencias relacionadas','competências','competencias','competência'],
    'habilidades_relacionadas': ['habilidades relacionadas']
}


def find_column_idx(header, candidates):
    header_l = [ (h or '').strip().lower() for h in header ]
    for cand in candidates:
        # exact
        for i,h in enumerate(header_l):
            if h == cand:
                return i
    for cand in candidates:
        for i,h in enumerate(header_l):
            if cand in h:
                return i
    return None


def inspect_file(path, sheet_name=None, max_samples=5):
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames
    out = []
    for name in sheets:
        if name not in wb.sheetnames:
            print(f"Sheet not found: {name}")
            continue
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            continue
        header = [ (h or '') for h in header ]
        mapping = {}
        for key, cands in CANDIDATES.items():
            idx = find_column_idx(header, cands)
            mapping[key] = idx
        samples = []
        for _ in range(max_samples):
            try:
                row = next(it)
            except StopIteration:
                break
            samples.append(row)
        out.append((name, header, mapping, samples))
    return out


def format_cell(v):
    return '' if v is None else str(v).strip()


def dry_run(path, sheet=None, max_samples=10):
    res = inspect_file(path, sheet_name=sheet, max_samples=max_samples)
    total_codes = 0
    for sheet_name, header, mapping, samples in res:
        print(f"Sheet: {sheet_name}")
        print(f"Header: {header}")
        print('Detected mapping (column index or None):')
        for k,v in mapping.items():
            print(f"  {k}: {v} -> {header[v] if v is not None and v < len(header) else None}")
        print('Sample parsed rows:')
        shown = 0
        for row in samples:
            codigo = None
            if mapping['codigo'] is not None and mapping['codigo'] < len(row):
                codigo = format_cell(row[mapping['codigo']])
            descricao = None
            if mapping['descricao'] is not None and mapping['descricao'] < len(row):
                descricao = format_cell(row[mapping['descricao']])
            if not codigo:
                continue
            parsed = parse_bncc_code(codigo)
            total_codes += 1
            print(f"  codigo={codigo} | descricao={(descricao or '')[:80]!r}")
            print(f"    parsed: etapa={parsed.get('etapa_sigla')} componente={parsed.get('componente_codigo')} em_comp={parsed.get('em_competencia')} em_seq={parsed.get('em_sequencia')} grupo_faixa={parsed.get('grupo_faixa')} campo={parsed.get('campo_experiencias')} ano_bloco={parsed.get('ano_bloco')}")
            shown += 1
            if shown >= max_samples:
                break
        print()
    print(f"Total sample codes found across sheets (approx): {total_codes}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--file', required=True)
    parser.add_argument('--sheet', default=None)
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--max-samples', type=int, default=5)
    parser.add_argument('--host', default='localhost')
    parser.add_argument('--port', type=int, default=3306)
    parser.add_argument('--user', default=None)
    parser.add_argument('--database', default=None)
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"Arquivo não encontrado: {path}")
        raise SystemExit(1)

    if args.dry_run:
        dry_run(path, sheet=args.sheet, max_samples=args.max_samples)
    else:
        # modo de gravação: delegar para importar_bncc_from_excel.py por sheet
        if not args.user or not args.database:
            print('Para aplicar (gravar) é necessário informar --user e --database')
            raise SystemExit(1)
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = [args.sheet] if args.sheet else wb.sheetnames
        script_path = Path(__file__).resolve().parents[1] / 'importar_bncc_from_excel.py'
        for sheet in sheets:
            print(f"Applying import for sheet: {sheet}")
            cmd = [sys.executable, str(script_path), '--file', str(path), '--sheet', sheet, '--user', args.user, '--database', args.database, '--host', args.host, '--port', str(args.port)]
            # do not pass password on command-line; the called script will prompt
            subprocess.run(cmd)
