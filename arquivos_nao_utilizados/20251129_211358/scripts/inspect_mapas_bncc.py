from openpyxl import load_workbook
from pathlib import Path
wb_path = Path(r"C:\gestao\MapasDeFocoBncc_Unificados.xlsx")
if not wb_path.exists():
    print(f"Arquivo não encontrado: {wb_path}")
    raise SystemExit(1)

wb = load_workbook(wb_path, read_only=True, data_only=True)
print(f"Workbook: {wb_path}\nSheets: {len(wb.sheetnames)}\n")
for name in wb.sheetnames:
    print(f"--- Sheet: {name}")
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        print("(vazia)")
        continue
    # normalize header for printing
    header_display = [str(h) if h is not None else '' for h in header]
    print("Header:", header_display)
    # print up to 5 sample rows
    samples = []
    for _ in range(5):
        try:
            row = next(it)
            samples.append([ (str(c) if c is not None else '') for c in row ])
        except StopIteration:
            break
    print("Samples:")
    for s in samples:
        print(s)
    print()

print("Inspeção concluída.")
