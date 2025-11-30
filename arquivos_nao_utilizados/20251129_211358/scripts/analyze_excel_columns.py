#!/usr/bin/env python3
"""
Análise detalhada das colunas do Excel vs estrutura do banco
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter

path = Path(r"C:\gestao\MapasDeFocoBncc_Unificados.xlsx")
wb = load_workbook(path, read_only=True, data_only=True)

all_headers = []
sheets_analyzed = []

for sheet_name in wb.sheetnames[:10]:  # analisar primeiras 10 sheets
    ws = wb[sheet_name]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        continue
    
    # normalizar headers
    normalized = []
    for h in header:
        if h and str(h).strip():
            normalized.append(str(h).strip())
    
    if normalized:
        all_headers.extend(normalized)
        sheets_analyzed.append((sheet_name, normalized))

# contar frequência de cada coluna
counter = Counter(all_headers)

print("=== ANÁLISE DE COLUNAS DO EXCEL ===\n")
print(f"Sheets analisadas: {len(sheets_analyzed)}\n")
print("Colunas mais comuns (top 20):")
for col, count in counter.most_common(20):
    print(f"  {count:2d}x | {col}")

print("\n=== AMOSTRA DE HEADERS POR SHEET ===")
for name, headers in sheets_analyzed[:5]:
    print(f"\n{name}:")
    for i, h in enumerate(headers[:15]):  # primeiros 15 headers
        print(f"  [{i}] {h}")
