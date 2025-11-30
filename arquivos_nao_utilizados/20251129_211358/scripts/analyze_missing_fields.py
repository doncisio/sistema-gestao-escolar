#!/usr/bin/env python3
"""
Análise de campos ausentes e amostras de dados do Excel
"""
from openpyxl import load_workbook
from pathlib import Path
import mysql.connector
import os
from dotenv import load_dotenv

load_dotenv('.env')

path = Path(r"C:\gestao\MapasDeFocoBncc_Unificados.xlsx")
wb = load_workbook(path, read_only=True, data_only=True)

# Analisar sheet específica com dados reais
ws = wb["7 ano MT"]
it = ws.iter_rows(values_only=True)
header = next(it)

print("=== ANÁLISE DETALHADA: 7 ano MT ===\n")
print("Headers:")
for i, h in enumerate(header[:12]):
    if h and str(h).strip():
        print(f"  [{i}] {h}")

print("\n=== AMOSTRA DE 3 REGISTROS ===")
for idx, row in enumerate(it):
    if idx >= 3:
        break
    print(f"\nRegistro {idx+1}:")
    for i in range(min(9, len(row))):
        h = header[i] if i < len(header) else f"col_{i}"
        val = row[i]
        if val and str(val).strip():
            print(f"  [{i}] {h}: {str(val)[:120]}")

# Verificar o que temos no banco
conn = mysql.connector.connect(
    host=os.getenv('DB_HOST','localhost'),
    user=os.getenv('DB_USER'),
    password=os.getenv('DB_PASSWORD'),
    database=os.getenv('DB_NAME','redeescola')
)
c = conn.cursor()

print("\n\n=== DADOS NO BANCO (sample) ===")
c.execute("""
    SELECT codigo, 
           LEFT(descricao, 80) as descricao_trunc,
           LEFT(conhecimento_previo, 80) as conhec_prev_trunc,
           etapa_sigla, componente_codigo, ano_bloco
    FROM bncc_habilidades 
    WHERE componente_codigo = 'MA' AND ano_bloco = '07'
    LIMIT 3
""")

for row in c.fetchall():
    print(f"\nCódigo: {row[0]}")
    print(f"  Descrição: {row[1]}")
    print(f"  Conhec.Prévio: {row[2]}")
    print(f"  Parsed: etapa={row[3]}, comp={row[4]}, ano={row[5]}")

conn.close()

print("\n\n=== CAMPOS AUSENTES NA TABELA ===")
campos_excel = [
    "Campo de atuação",
    "Classificação", 
    "Objetivos de aprendizagem",
    "Competências relacionadas",
    "Habilidades relacionadas",
    "Comentários"
]

campos_no_banco = [
    "codigo",
    "descricao", 
    "conhecimento_previo",
    "etapa_sigla",
    "ano_bloco",
    "componente_codigo",
    "grupo_faixa",
    "campo_experiencias",
    "em_competencia",
    "em_sequencia",
    "codigo_raw"
]

print("\nCampos do Excel NÃO armazenados:")
for campo in campos_excel:
    print(f"  ❌ {campo}")

print("\nCampos no banco:")
for campo in campos_no_banco:
    print(f"  ✓ {campo}")
