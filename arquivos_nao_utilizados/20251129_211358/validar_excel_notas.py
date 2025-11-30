"""
Script de validação e visualização de arquivos Excel de notas gerados
"""

import openpyxl
import os
from pathlib import Path
from typing import Any, cast
from config_logs import get_logger

logger = get_logger(__name__)

def validar_excel_notas(excel_path):
    """
    Valida a estrutura de um arquivo Excel de notas
    """
    try:
        logger.info("\n" + "="*70)
        logger.info(f"VALIDANDO: {os.path.basename(excel_path)}")
        logger.info("" + "="*70 + "\n")
        
        # Carregar workbook
        wb = openpyxl.load_workbook(excel_path)
        ws = cast(Any, wb.active)
        
        # Informações gerais
        logger.info(f"📋 Nome da planilha: {ws.title}")
        logger.info(f"📏 Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")
        
        # Ler informações do topo
        logger.info("\n📌 Informações Extraídas:")
        if ws['A1'].value:
            logger.info(f"   {ws['A1'].value}")
        if ws['A2'].value:
            logger.info(f"   {ws['A2'].value}")
        if ws['A3'].value:
            logger.info(f"   {ws['A3'].value}")
        
        # Encontrar linha de cabeçalho
        linha_header = 4
        for row in range(1, 10):
            if ws.cell(row, 1).value == 'Nº':
                linha_header = row
                break
        
        logger.info(f"\n📊 Cabeçalho (Linha {linha_header}):")
        for col in range(1, 8):
            valor = ws.cell(linha_header, col).value
            if valor:
                logger.info(f"   Col {col}: {valor}")
        
        # Contar alunos
        num_alunos = ws.max_row - linha_header
        logger.info(f"\n👥 Total de alunos: {num_alunos}")
        
        # Mostrar primeiros 5 alunos
        logger.info(f"\n📝 Amostra de Dados (primeiros 5 alunos):")
        logger.info(f"{'-'*70}")
        
        for row in range(linha_header + 1, min(linha_header + 6, ws.max_row + 1)):
            ordem = ws.cell(row, 1).value or ''
            nome = ws.cell(row, 2).value or ''
            nota = ws.cell(row, 3).value or '-'
            
            # Formatar valor
            if isinstance(nota, (int, float)):
                nota = f"{nota:.2f}"
            
            logger.info(f"\n{ordem}. {nome}")
            logger.info(f"   Nota Final: {nota}")
        
        # Estatísticas
        logger.info(f"\n📈 Estatísticas:")
        
        # Calcular estatísticas das notas finais
        notas = []
        for row in range(linha_header + 1, ws.max_row + 1):
            nota = ws.cell(row, 3).value
            if isinstance(nota, (int, float)):
                notas.append(nota)
        
        if notas:
            media_turma = sum(notas) / len(notas)
            maior_nota = max(notas)
            menor_nota = min(notas)
            
            logger.info(f"   Média da turma: {media_turma:.2f}")
            logger.info(f"   Maior nota: {maior_nota:.2f}")
            logger.info(f"   Menor nota: {menor_nota:.2f}")
            
            # Aprovados (nota >= 70.0, já que são médias * 10)
            aprovados = sum(1 for n in notas if n >= 70.0)
            logger.info(f"   Aprovados (≥70.0): {aprovados}/{len(notas)} ({aprovados/len(notas)*100:.1f}%)")
        
        logger.info(f"\n✅ Arquivo válido e estrutura correta!")
        logger.info(f"{'='*70}\n")
        
        wb.close()
        return True
        
    except Exception as e:
        logger.exception(f"\n❌ ERRO: {e}\n")
        return False


def listar_e_validar_todos():
    """
    Lista e valida todos os arquivos Excel de notas no diretório
    """
    logger.info("🔍 Procurando arquivos Excel de notas...\n")
    
    # Procurar arquivos
    arquivos = list(Path('.').glob('Template_Notas*.xlsx'))
    
    if not arquivos:
        logger.warning("❌ Nenhum arquivo encontrado com o padrão 'Template_Notas*.xlsx'")
        return
    
    logger.info(f"✅ Encontrados {len(arquivos)} arquivo(s):\n")
    for i, arquivo in enumerate(arquivos, 1):
        logger.info(f"   {i}. {arquivo.name}")
    
    # Validar cada arquivo
    logger.info(f"\n{'='*70}")
    logger.info("INICIANDO VALIDAÇÃO")
    logger.info(f"{'='*70}")
    
    for arquivo in arquivos:
        validar_excel_notas(str(arquivo))


def criar_visualizacao_ascii(excel_path):
    """
    Cria uma visualização ASCII da tabela de notas
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = cast(Any, wb.active)
        
        # Encontrar linha de cabeçalho
        linha_header = 4
        for row in range(1, 10):
            if ws.cell(row, 1).value == 'Nº':
                linha_header = row
                break
        
        logger.info(f"\n{'='*90}")
        logger.info(f"VISUALIZAÇÃO: {os.path.basename(excel_path)}")
        logger.info(f"{'='*90}\n")
        
        # Imprimir informações do topo
        for row in range(1, linha_header):
            valor = ws.cell(row, 1).value
            if valor:
                logger.info(f"  {valor}")
        
        logger.info(f"\n{'─'*90}")
        
        # Cabeçalho
        logger.info(f"│ {'Nº':^4} │ {'Nome do Aluno':<45} │ {'Nota':^12} │")
        logger.info(f"{'─'*90}")
        
        # Dados (primeiros 10)
        for row in range(linha_header + 1, min(linha_header + 11, ws.max_row + 1)):
            ordem = str(ws.cell(row, 1).value or '').strip()
            nome = str(ws.cell(row, 2).value or '')[:45]  # Limitar nome
            nota = ws.cell(row, 3).value
            
            # Formatar
            nota_str = f"{nota:.2f}" if isinstance(nota, (int, float)) else "-"
            
            logger.info(f"│ {ordem:>4} │ {nome:<45} │ {nota_str:^12} │")
        
        if ws.max_row > linha_header + 10:
            logger.info(f"│ {'...':<4} │ {'...':<45} │ {'...':<12} │")

        logger.info(f"{'─'*90}\n")
        
        wb.close()
        
    except Exception as e:
        logger.exception(f"Erro ao criar visualização: {e}")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        # Validar arquivo específico
        arquivo = sys.argv[1]
        if os.path.exists(arquivo):
            validar_excel_notas(arquivo)
            criar_visualizacao_ascii(arquivo)
        else:
            logger.error(f"Arquivo não encontrado: {arquivo}")
    else:
        # Validar todos os arquivos
        listar_e_validar_todos()
