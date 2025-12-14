from openpyxl import load_workbook
path = r"C:\gestao\MapasDeFocoBncc_Unificados.xlsx"
sheet = '1º ano LP'
wb = load_workbook(filename=path, read_only=True, data_only=True)
if sheet not in wb.sheetnames:
    print('Sheet not found:', sheet)
    print('Available:', wb.sheetnames)
    raise SystemExit(1)
ws = wb[sheet]
count = 0
for row in ws.iter_rows(min_row=1, max_row=120, values_only=True):
    # print only rows that have some content
    if any(cell is not None for cell in row):
        print(count+1, row)
        count += 1
    if count >= 40:
        break
print('\nDone')
