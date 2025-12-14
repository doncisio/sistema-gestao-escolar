import pandas as pd
import re
from pathlib import Path
import openpyxl
import zipfile
import shutil
import time
import errno
import os

IN_CSV = Path('pendencias_por_bimestre.csv')
# Allow overriding output path via environment variable `EXPORT_PENDENCIAS_OUT`.
# This enables callers (e.g. the GUI) to ask the user for a Save As path and
# provide it to this script without changing its CLI.
env_out = os.environ.get('EXPORT_PENDENCIAS_OUT')
if env_out:
    OUT_XLSX = Path(env_out)
else:
    OUT_XLSX = Path('pendencias_por_bimestre.xlsx')

# Ler CSV se existir; caso contrário, buscar dados diretamente via relatorio_pendencias
if IN_CSV.exists():
    try:
        df = pd.read_csv(IN_CSV, encoding='utf-8')
    except Exception:
        df = pd.read_csv(IN_CSV)
else:
    # construir via relatorio_pendencias.buscar_pendencias_notas
    try:
        from src.relatorios.relatorio_pendencias import buscar_pendencias_notas
    except Exception as e:
        raise SystemExit(f"Arquivo CSV não encontrado e módulo 'relatorio_pendencias' indisponível: {e}")

    bimestres = ['1º bimestre', '2º bimestre', '3º bimestre', '4º bimestre']
    niveis = ['iniciais', 'finais']
    rows = []
    ESOLA_ID = 60
    for b in bimestres:
        for n in niveis:
            try:
                pend = buscar_pendencias_notas(b, n, None, ESOLA_ID)
            except Exception as e:
                print(f"Erro ao buscar pendências para {b} / {n}: {e}")
                pend = {}

            for chave, info in pend.items():
                serie, turma, turno = chave
                disciplinas_sem_lanc = sorted(list(info.get('disciplinas_sem_lancamento', []))) if info.get('disciplinas_sem_lancamento') else []
                for aluno_id, aluno_info in info.get('alunos', {}).items():
                    disc_sem_nota = sorted(list(set(aluno_info.get('disciplinas_sem_nota', []))))
                    if not disc_sem_nota:
                        continue
                    rows.append({
                        'bimestre': b,
                        'nivel': n,
                        'serie': serie,
                        'turma': turma,
                        'turno': turno,
                        'aluno_id': aluno_id,
                        'aluno_nome': aluno_info.get('nome'),
                        'disciplinas_sem_nota': ';'.join(disc_sem_nota),
                        'disciplinas_sem_lancamento': ';'.join(disciplinas_sem_lanc)
                    })
                has_alunos_com_pend = any(len(a.get('disciplinas_sem_nota', [])) > 0 for a in info.get('alunos', {}).values())
                if disciplinas_sem_lanc and not has_alunos_com_pend:
                    rows.append({
                        'bimestre': b,
                        'nivel': n,
                        'serie': serie,
                        'turma': turma,
                        'turno': turno,
                        'aluno_id': '',
                        'aluno_nome': '',
                        'disciplinas_sem_nota': '',
                        'disciplinas_sem_lancamento': ';'.join(disciplinas_sem_lanc)
                    })

    df = pd.DataFrame(rows, columns=['bimestre','nivel','serie','turma','turno','aluno_id','aluno_nome','disciplinas_sem_nota','disciplinas_sem_lancamento'])

# Função para sanitizar nomes de abas (máx 31 chars, evitar caracteres inválidos)
def safe_sheet_name(name: str) -> str:
    # remover caracteres inválidos e limitar tamanho
    name = re.sub(r"[\\/*?:\[\]]", "_", str(name))
    name = name.strip()
    if len(name) > 31:
        # tentar encurtar preservando palavras
        parts = name.split()
        out = ''
        for p in parts:
            if len(out) + 1 + len(p) <= 31:
                out = (out + ' ' + p).strip()
            else:
                break
        if not out:
            out = name[:31]
        name = out
    return name

# Garantir colunas esperadas
expected = {'bimestre','nivel','serie','turma','turno','aluno_id','aluno_nome','disciplinas_sem_nota','disciplinas_sem_lancamento'}
missing = expected - set(df.columns)
if missing:
    print(f"Aviso: colunas faltando no CSV: {missing}")

# Preparar nomes de abas que serão gravadas
bimestre_names = []
for b in sorted(df['bimestre'].dropna().unique()):
    bimestre_names.append(safe_sheet_name(b))
sheets_to_write = set(['Resumo']) | set(bimestre_names)

# Se o arquivo já existir, remover apenas as abas que iremos atualizar
OUT_TARGET = OUT_XLSX
if OUT_XLSX.exists():
    try:
        wb = openpyxl.load_workbook(OUT_XLSX)
        # Garantir que exista pelo menos uma sheet visível
        try:
            any_visible = any(getattr(ws, 'sheet_state', 'visible') == 'visible' for ws in wb.worksheets)
            if not any_visible and len(wb.worksheets) > 0:
                wb.worksheets[0].sheet_state = 'visible'
        except Exception:
            pass

        removed = []
        for sheet in list(wb.sheetnames):
            if sheet in sheets_to_write:
                try:
                    std = wb[sheet]
                    wb.remove(std)
                    removed.append(sheet)
                except Exception:
                    pass
        if removed:
            # salvar workbook sem as abas removidas antes de reescrever
            try:
                wb.save(OUT_XLSX)
            except Exception as e:
                # Se falhar ao salvar (arquivo em uso), avisar e ajustar destino para não sobrescrever
                print(f"Aviso: não foi possível salvar alterações no workbook existente: {e}")
                OUT_TARGET = OUT_XLSX.with_name(OUT_XLSX.stem + f'.new.{int(time.time())}' + OUT_XLSX.suffix)
                print(f"Usando arquivo alternativo para saída: {OUT_TARGET}")
    except Exception as e:
        print(f"Aviso: falha ao abrir/atualizar workbook existente: {e}")

# Preparar modo de escrita. Se o arquivo existir, verificar se é um XLSX válido.
# Em caso de arquivo corrompido (ex.: falta [Content_Types].xml) faremos backup
# e criaremos um novo workbook para evitar erro de leitura do openpyxl/pandas.
mode = 'w'
if OUT_XLSX.exists():
    try:
        # Primeiro checar se é um zip (XLSX é um zip format)
        if not zipfile.is_zipfile(OUT_XLSX):
            # Não é um XLSX válido — mover para backup e criar novo
            bak = OUT_XLSX.with_suffix(OUT_XLSX.suffix + f'.bak.{int(time.time())}')
            try:
                shutil.move(str(OUT_XLSX), str(bak))
                print(f"Arquivo existente inválido movido para backup: {bak}")
            except Exception as e:
                # Se falhar ao mover (ex.: arquivo em uso), não interromper; escolher saída alternativa
                msg = str(e)
                print(f"Falha ao mover arquivo inválido para backup: {e}")
                if 'being used by another process' in msg or 'The process cannot access the file' in msg or getattr(e, 'winerror', None) == 32:
                    OUT_TARGET = OUT_XLSX.with_name(OUT_XLSX.stem + f'.new.{int(time.time())}' + OUT_XLSX.suffix)
                    print(f"Arquivo em uso: não será movido. Usando saída alternativa: {OUT_TARGET}")
            mode = 'w'
        else:
            # Tentar abrir com openpyxl — pode lançar se estiver corrompido
            try:
                wb = openpyxl.load_workbook(OUT_XLSX)
                # Garantir que exista pelo menos uma sheet visível
                try:
                    any_visible = any(getattr(ws, 'sheet_state', 'visible') == 'visible' for ws in wb.worksheets)
                    if not any_visible and len(wb.worksheets) > 0:
                        wb.worksheets[0].sheet_state = 'visible'
                except Exception:
                    pass

                removed = []
                for sheet in list(wb.sheetnames):
                    if sheet in sheets_to_write:
                        try:
                            std = wb[sheet]
                            wb.remove(std)
                            removed.append(sheet)
                        except Exception:
                            pass
                if removed:
                    try:
                        wb.save(OUT_XLSX)
                    except Exception as e_save:
                        print(f"Aviso: não foi possível salvar alterações no workbook existente: {e_save}")
                        OUT_TARGET = OUT_XLSX.with_name(OUT_XLSX.stem + f'.new.{int(time.time())}' + OUT_XLSX.suffix)
                        print(f"Usando arquivo alternativo para saída: {OUT_TARGET}")
                try:
                    wb.close()
                except Exception:
                    pass
                mode = 'a'
            except Exception as e:
                # Provavelmente corrompido (p.ex. falta [Content_Types].xml)
                bak = OUT_XLSX.with_suffix(OUT_XLSX.suffix + f'.bak.{int(time.time())}')
                try:
                    shutil.move(str(OUT_XLSX), str(bak))
                    print(f"Arquivo XLSX corrompido movido para backup: {bak} (erro: {e})")
                except Exception as e2:
                    msg = str(e2)
                    print(f"Falha ao mover arquivo corrompido para backup: {e2}")
                    if 'being used by another process' in msg or 'The process cannot access the file' in msg or getattr(e2, 'winerror', None) == 32:
                        OUT_TARGET = OUT_XLSX.with_name(OUT_XLSX.stem + f'.new.{int(time.time())}' + OUT_XLSX.suffix)
                        print(f"Arquivo em uso: não será movido. Usando saída alternativa: {OUT_TARGET}")
                mode = 'w'
    except Exception as e:
        print(f"Aviso inesperado ao verificar arquivo existente: {e}")

# Agora escrever diretamente com openpyxl para preservar dimensões/formatacoes de colunas
try:
    if OUT_XLSX.exists() and mode == 'a' and OUT_TARGET == OUT_XLSX:
        wb = openpyxl.load_workbook(OUT_XLSX)
    else:
        # criar novo workbook
        wb = openpyxl.Workbook()
        # remover sheet padrão se for o único
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
            std = wb['Sheet']
            wb.remove(std)

    def write_df_to_sheet(ws, dataframe, preserve_col_widths=True):
        # Apagar conteúdo existente (manter dimensões de coluna)
        try:
            if ws.max_row > 0:
                ws.delete_rows(1, ws.max_row)
        except Exception:
            pass

        # Escrever cabeçalho
        cols = list(dataframe.columns)
        for c_idx, col_name in enumerate(cols, start=1):
            ws.cell(row=1, column=c_idx, value=str(col_name))

        # Escrever linhas
        for r_idx, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Ajustar larguras apenas se a planilha for nova (preservar se existir)
        if not preserve_col_widths:
            for i, col in enumerate(cols, start=1):
                col_letter = openpyxl.utils.get_column_letter(i)
                try:
                    max_len = max(
                        [len(str(x)) if x is not None else 0 for x in dataframe[col].values] + [len(str(col))]
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
                except Exception:
                    pass

    # Gerar resumo como DataFrame
    try:
        resumo = df.groupby(['bimestre','nivel']).agg(
            total_linhas=pd.NamedAgg(column='aluno_id', aggfunc='count'),
            total_alunos_com_pendencias=pd.NamedAgg(column='aluno_id', aggfunc=lambda s: s[s.notna()].nunique())
        ).reset_index()
    except Exception as e:
        print('Falha ao gerar aba Resumo:', e)
        resumo = pd.DataFrame(columns=['bimestre','nivel','total_linhas','total_alunos_com_pendencias'])

    # Escrever/atualizar aba Resumo
    if 'Resumo' in wb.sheetnames:
        ws = wb['Resumo']
        preserve = True
    else:
        ws = wb.create_sheet('Resumo')
        preserve = False
    write_df_to_sheet(ws, resumo, preserve_col_widths=preserve)

    # Para cada bimestre, atualizar ou criar a aba
    for bimestre in sorted(df['bimestre'].dropna().unique()):
        sub = df[df['bimestre'] == bimestre].copy()
        sheet = safe_sheet_name(bimestre)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            preserve = True
        else:
            ws = wb.create_sheet(sheet)
            preserve = False
        write_df_to_sheet(ws, sub, preserve_col_widths=preserve)

    # Salvar workbook (usar OUT_TARGET se definido)
    try:
        wb.save(OUT_TARGET)
    except Exception as e_save_final:
        # Tentativa alternativa: se falhar ao salvar no destino (ex.: bloqueio), tentar um nome alternativo
        msg = str(e_save_final)
        print(f"Erro ao salvar arquivo XLSX no destino {OUT_TARGET}: {e_save_final}")
        alt = OUT_XLSX.with_name(OUT_XLSX.stem + f'.auto.{int(time.time())}' + OUT_XLSX.suffix)
        try:
            wb.save(alt)
            print(f"Arquivo salvo em saída alternativa: {alt}")
            OUT_TARGET = alt
        except Exception as e_alt:
            print(f"Falha ao salvar em arquivo alternativo também: {e_alt}")
            raise
    try:
        wb.close()
    except Exception:
        pass
except Exception as e:
    print(f"Erro ao gravar arquivo XLSX: {e}")

print(f"Arquivo gerado: {OUT_TARGET.resolve()}")
