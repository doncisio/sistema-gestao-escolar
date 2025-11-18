import pandas as pd
import re
from pathlib import Path
import openpyxl
import zipfile
import shutil
import time

IN_CSV = Path('pendencias_por_bimestre.csv')
OUT_XLSX = Path('pendencias_por_bimestre.xlsx')

if not IN_CSV.exists():
    raise SystemExit(f"Arquivo CSV não encontrado: {IN_CSV.resolve()}")

# Ler CSV
try:
    df = pd.read_csv(IN_CSV, encoding='utf-8')
except Exception:
    # tentar sem especificar encoding
    df = pd.read_csv(IN_CSV)

# Função para sanitizar nomes de abas (máx 31 chars, evitar caracteres inválidos)
def safe_sheet_name(name: str) -> str:
    # remover caracteres inválidos e limitar tamanho
    name = re.sub(r"[\\/*?:\[\]]", "_", str(name))
    name = name.strip()
    if len(name) > 31:
        # tentar encurtar preservando palavras
        parts = name.split()
        out = ''
        for p in parts:
            if len(out) + 1 + len(p) <= 31:
                out = (out + ' ' + p).strip()
            else:
                break
        if not out:
            out = name[:31]
        name = out
    return name

# Garantir colunas esperadas
expected = {'bimestre','nivel','serie','turma','turno','aluno_id','aluno_nome','disciplinas_sem_nota','disciplinas_sem_lancamento'}
missing = expected - set(df.columns)
if missing:
    print(f"Aviso: colunas faltando no CSV: {missing}")

# Preparar nomes de abas que serão gravadas
bimestre_names = []
for b in sorted(df['bimestre'].dropna().unique()):
    bimestre_names.append(safe_sheet_name(b))
sheets_to_write = set(['Resumo']) | set(bimestre_names)

# Se o arquivo já existir, remover apenas as abas que iremos atualizar
if OUT_XLSX.exists():
    try:
        wb = openpyxl.load_workbook(OUT_XLSX)
        removed = []
        for sheet in list(wb.sheetnames):
            if sheet in sheets_to_write:
                try:
                    std = wb[sheet]
                    wb.remove(std)
                    removed.append(sheet)
                except Exception:
                    pass
        if removed:
            # salvar workbook sem as abas removidas antes de reescrever
            wb.save(OUT_XLSX)
    except Exception as e:
        print(f"Aviso: falha ao abrir/atualizar workbook existente: {e}")

# Preparar modo de escrita. Se o arquivo existir, verificar se é um XLSX válido.
# Em caso de arquivo corrompido (ex.: falta [Content_Types].xml) faremos backup
# e criaremos um novo workbook para evitar erro de leitura do openpyxl/pandas.
mode = 'w'
if OUT_XLSX.exists():
    try:
        # Primeiro checar se é um zip (XLSX é um zip format)
        if not zipfile.is_zipfile(OUT_XLSX):
            # Não é um XLSX válido — mover para backup e criar novo
            bak = OUT_XLSX.with_suffix(OUT_XLSX.suffix + f'.bak.{int(time.time())}')
            try:
                shutil.move(str(OUT_XLSX), str(bak))
                print(f"Arquivo existente inválido movido para backup: {bak}")
            except Exception as e:
                print(f"Falha ao mover arquivo inválido para backup: {e}")
            mode = 'w'
        else:
            # Tentar abrir com openpyxl — pode lançar se estiver corrompido
            try:
                wb = openpyxl.load_workbook(OUT_XLSX)
                removed = []
                for sheet in list(wb.sheetnames):
                    if sheet in sheets_to_write:
                        try:
                            std = wb[sheet]
                            wb.remove(std)
                            removed.append(sheet)
                        except Exception:
                            pass
                if removed:
                    wb.save(OUT_XLSX)
                try:
                    wb.close()
                except Exception:
                    pass
                mode = 'a'
            except Exception as e:
                # Provavelmente corrompido (p.ex. falta [Content_Types].xml)
                bak = OUT_XLSX.with_suffix(OUT_XLSX.suffix + f'.bak.{int(time.time())}')
                try:
                    shutil.move(str(OUT_XLSX), str(bak))
                    print(f"Arquivo XLSX corrompido movido para backup: {bak} (erro: {e})")
                except Exception as e2:
                    print(f"Falha ao mover arquivo corrompido para backup: {e2}")
                mode = 'w'
    except Exception as e:
        print(f"Aviso inesperado ao verificar arquivo existente: {e}")

# Agora escrever diretamente com openpyxl para preservar dimensões/formatacoes de colunas
try:
    if OUT_XLSX.exists() and mode == 'a':
        wb = openpyxl.load_workbook(OUT_XLSX)
    else:
        # criar novo workbook
        wb = openpyxl.Workbook()
        # remover sheet padrão se for o único
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
            std = wb['Sheet']
            wb.remove(std)

    def write_df_to_sheet(ws, dataframe, preserve_col_widths=True):
        # Apagar conteúdo existente (manter dimensões de coluna)
        try:
            if ws.max_row > 0:
                ws.delete_rows(1, ws.max_row)
        except Exception:
            pass

        # Escrever cabeçalho
        cols = list(dataframe.columns)
        for c_idx, col_name in enumerate(cols, start=1):
            ws.cell(row=1, column=c_idx, value=str(col_name))

        # Escrever linhas
        for r_idx, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Ajustar larguras apenas se a planilha for nova (preservar se existir)
        if not preserve_col_widths:
            for i, col in enumerate(cols, start=1):
                col_letter = openpyxl.utils.get_column_letter(i)
                try:
                    max_len = max(
                        [len(str(x)) if x is not None else 0 for x in dataframe[col].values] + [len(str(col))]
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
                except Exception:
                    pass

    # Gerar resumo como DataFrame
    try:
        resumo = df.groupby(['bimestre','nivel']).agg(
            total_linhas=pd.NamedAgg(column='aluno_id', aggfunc='count'),
            total_alunos_com_pendencias=pd.NamedAgg(column='aluno_id', aggfunc=lambda s: s[s.notna()].nunique())
        ).reset_index()
    except Exception as e:
        print('Falha ao gerar aba Resumo:', e)
        resumo = pd.DataFrame(columns=['bimestre','nivel','total_linhas','total_alunos_com_pendencias'])

    # Escrever/atualizar aba Resumo
    if 'Resumo' in wb.sheetnames:
        ws = wb['Resumo']
        preserve = True
    else:
        ws = wb.create_sheet('Resumo')
        preserve = False
    write_df_to_sheet(ws, resumo, preserve_col_widths=preserve)

    # Para cada bimestre, atualizar ou criar a aba
    for bimestre in sorted(df['bimestre'].dropna().unique()):
        sub = df[df['bimestre'] == bimestre].copy()
        sheet = safe_sheet_name(bimestre)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            preserve = True
        else:
            ws = wb.create_sheet(sheet)
            preserve = False
        write_df_to_sheet(ws, sub, preserve_col_widths=preserve)

    # Salvar workbook
    wb.save(OUT_XLSX)
    try:
        wb.close()
    except Exception:
        pass
except Exception as e:
    print(f"Erro ao gravar arquivo XLSX: {e}")

print(f"Arquivo gerado: {OUT_XLSX.resolve()}")
